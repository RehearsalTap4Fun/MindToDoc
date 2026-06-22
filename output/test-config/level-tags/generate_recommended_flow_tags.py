# -*- coding: utf-8 -*-
"""生成 500 关推荐版心流曲线 tag。

推荐版目标:
- 以主生成器 10 个 tier 为骨架。
- 每 10 关一轮小循环:恢复/技巧/变化/压力/小高潮/恢复/复合/难点/备战/Boss。
- 60%-70% 关卡贴 tag,其余保留 tier 默认。
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
DEFAULT_INPUT = HERE / "LevelTagCfg.xlsx"

ROUNDS_TOTAL = 50
LEVELS_PER_ROUND = 10


def _tier(round_id: int) -> int:
    return (round_id - 1) // 5 + 1


def _base_tags(level_in_round: int, tier: int, round_id: int) -> tuple[list[str], str]:
    """返回单关 tag 与 Note。tag 顺序即 patch 顺序。"""
    phase = level_in_round

    if phase == 1:
        if tier <= 2:
            return ["short_match", "penalty_focus"], "恢复:短局点球确定性,给玩家重新进入心流"
        return ["short_match", "easy_minus"], "恢复:降低一档压力,避免连续挫败"
    if phase == 2:
        return ["set_piece"] if tier >= 3 else [], "技巧主题:任意球/点球基础复习"
    if phase == 3:
        if tier >= 6 and round_id % 2 == 0:
            return ["corner_focus"], "变化:角球变化,打破单一射门"
        return [], "默认:保留基础节奏,让玩家消化前一关技巧"
    if phase == 4:
        if tier >= 5:
            return ["hard_plus"], "压力:提高 AI 与对手星级,制造专注点"
        return [], "默认:压力预备,仅走 tier 基础难度"
    if phase == 5:
        if tier == 1:
            return [], "默认:首段小高潮留给基础 tier,避免过早复杂化"
        tags = ["set_piece", "hard_plus"] if tier >= 4 else ["set_piece"]
        return tags, "小高潮:定位球复合考验"
    if phase == 6:
        if round_id % 2 == 1:
            return ["short_match", "easy_minus"], "恢复:缩短局长并降低 AI 压力"
        return [], "默认恢复:不额外改写,降低 tag 密度"
    if phase == 7:
        if tier >= 6:
            return ["all_v2"], "复合:全 v2 切片,提升操作变化"
        return [], "默认:复合预备,不额外抬压"
    if phase == 8:
        if tier >= 7:
            return ["gk_test", "hard_plus"], "难点:守门考验叠加 AI 压力"
        return ["gk_test"], "难点:守门读秒与方向判断"
    if phase == 9:
        if tier >= 8:
            return ["long_match", "narrow_angle"], "备战:长局叠加收窄角度"
        if tier >= 5:
            return ["long_match", "hard_plus"], "备战:长局叠加对手压力"
        return [], "默认:轮末前呼吸位"

    # phase == 10,每轮结尾是检验点;每 5 轮大段结尾一定 boss。
    if round_id % 5 == 0:
        if tier >= 9:
            return ["long_match", "extreme_keeper", "boss"], "大段 Boss:终局长局+极限门将+全胜要求"
        if tier >= 6:
            return ["long_match", "boss"], "大段 Boss:长局全胜检验"
        return ["boss"], "大段 Boss:阶段能力检验"
    if tier >= 7:
        return ["must_win"], "轮末检验:无平局空间"
    if tier >= 4:
        return ["hard_plus"], "轮末检验:提高一档难度"
    return [], "默认检验:保留 tier 基础节奏"


def _adjust_for_macro_arc(tags: list[str], level_id: int, round_id: int, tier: int) -> list[str]:
    """按 500 关大曲线做少量宏观调整。"""
    out = list(tags)

    if level_id == 1:
        return ["tutorial", "penalty_focus"]

    if tier >= 8 and level_id % 25 == 0 and "boss" not in out:
        out.append("must_win")
    if tier >= 9 and level_id % 20 == 0 and "extreme_keeper" not in out and "narrow_angle" not in out:
        out.append("extreme_keeper")
    if tier <= 2 and "hard_plus" in out:
        out.remove("hard_plus")
    if tier <= 3 and "must_win" in out:
        out.remove("must_win")

    # 互斥兜底:每组只保留最符合本关强度的 tag。
    if "long_match" in out and "short_match" in out:
        out.remove("short_match")
    if "hard_plus" in out and "easy_minus" in out:
        out.remove("easy_minus" if tier >= 4 else "hard_plus")
    threshold_tags = [t for t in out if t in {"must_win", "lenient"}]
    if len(threshold_tags) > 1:
        out = [t for t in out if t != ("lenient" if tier >= 4 else "must_win")]
    modifier_tags = [t for t in out if t in {"extreme_keeper", "no_modifier", "narrow_angle"}]
    if len(modifier_tags) > 1:
        keep = "extreme_keeper" if tier >= 9 else modifier_tags[0]
        out = [t for t in out if t not in {"extreme_keeper", "no_modifier", "narrow_angle"} or t == keep]

    return out


def build_recommended_rows() -> list[dict]:
    rows: list[dict] = []
    for level_id in range(1, ROUNDS_TOTAL * LEVELS_PER_ROUND + 1):
        round_id = (level_id - 1) // LEVELS_PER_ROUND + 1
        level_in_round = (level_id - 1) % LEVELS_PER_ROUND + 1
        tier = _tier(round_id)
        tags, note = _base_tags(level_in_round, tier, round_id)
        tags = _adjust_for_macro_arc(tags, level_id, round_id, tier)
        rows.append({
            "ID": level_id,
            "Round": round_id,
            "LevelInRound": level_in_round,
            "Tier": tier,
            "Tags": tags,
            "Note": note,
        })
    return rows


def summarize(rows: list[dict]) -> dict:
    tags = Counter()
    tagged = 0
    for row in rows:
        if row["Tags"]:
            tagged += 1
            tags.update(row["Tags"])
    return {
        "levels_total": len(rows),
        "levels_with_tags": tagged,
        "coverage": round(tagged / len(rows), 3),
        "tag_counts": dict(sorted(tags.items())),
    }


def write_level_tag_cfg(path: Path = DEFAULT_INPUT) -> dict:
    rows = build_recommended_rows()
    wb = load_workbook(path)
    ws = wb["LevelTags"]
    by_id = {r["ID"]: r for r in rows}
    for row_idx in range(9, ws.max_row + 1):
        level_id = ws.cell(row_idx, 1).value
        if level_id not in by_id:
            continue
        item = by_id[level_id]
        ws.cell(row_idx, 5, ", ".join(item["Tags"]))
        ws.cell(row_idx, 6, item["Note"])
    wb.save(path)
    return summarize(rows)


def main() -> int:
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    rows = build_recommended_rows()
    summary = summarize(rows)
    print(summary)
    if not args.dry_run:
        write_level_tag_cfg(args.input)
        print(f"写入推荐版心流曲线: {args.input}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
