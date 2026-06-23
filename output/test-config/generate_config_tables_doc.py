# -*- coding: utf-8 -*-
"""从 ActivitySoccer_preview.xlsx 生成数值策划填表派生文档。"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

OUT_DIR = Path(__file__).parent
PROJECT_OUT = OUT_DIR.parent
XLSX = OUT_DIR / "ActivitySoccer_preview.xlsx"
LC_XLSX = OUT_DIR / "ActivitySoccerLanguage.xlsx"
SUMMARY = OUT_DIR / "test-config-summary.json"
OUTPUT = PROJECT_OUT / "2026世界杯主题活动-配置表结构.md"
SAMPLE_OUTPUT = PROJECT_OUT / "system-design-doc-samples" / "2026世界杯主题活动-配置表结构.md"

SHEET_META: dict[str, dict[str, str]] = {
    "ActvSoccerCharacterCfg": {"用途": "4 角色外观与展示战力", "关联": "→ 创角流程", "填表": "数值"},
    "ActvSoccerNationalityCfg": {"用途": "国籍与首签候选合同池", "关联": "ContractPool→ContractCfg；NameLcKey→Language", "填表": "数值"},
    "ActvSoccerTutorialCfg": {"用途": "试训三步与切片实例", "关联": "SliceInstanceID→SliceInstanceCfg", "填表": "关卡"},
    "ActvSoccerSlicePresetCfg": {"用途": "切片摆位预设（L2 主资产）", "关联": "SliceType；被 SliceInstanceCfg.PresetID 引用", "填表": "关卡"},
    "ActvSoccerSliceInstanceCfg": {"用途": "切片实例装配（L3）", "关联": "←Preset；→LevelCfg.SliceList、TutorialCfg", "填表": "关卡"},
    "ActvSoccerHapticCfg": {"用途": "震动事件强度/图案", "关联": "全局事件映射", "填表": "体验/数值"},
    "ActvSoccerLevelCfg": {"用途": "关卡：切片序列、胜平阈值、门票、对手", "关联": "SliceList→Instance；AiProfileID→Profile；Group→Season", "填表": "关卡+数值"},
    "ActvSoccerAiProfileCfg": {"用途": "关卡 AI 难度档", "关联": "←SliceInstanceCfg.AiProfileID", "填表": "数值"},
    "ActvSoccerEnemyAiCfg": {"用途": "门将/后卫/射手 AI 权重", "关联": "←SliceInstanceCfg（Goalkeeper/Defender/Shooter）", "填表": "数值"},
    "ActvSoccerAiModifierCfg": {"用途": "切片机制（移动门将等）", "关联": "←SliceInstanceCfg.ModifierID", "填表": "关卡"},
    "ActvSoccerSliceFlowCfg": {"用途": "各切片类型 FSM 流程参数", "关联": "按 SliceType 读取", "填表": "关卡"},
    "ActvSoccerCurrencyCfg": {"用途": "活动内货币类型说明", "关联": "被待遇/价格/奖励 vm.id 间接引用", "填表": "数值"},
    "ActvSoccerFameGrowthLevelCfg": {"用途": "知名度等级、许可、主角评分", "关联": "PlayerRating→淘汰赛主角评分；档位可拆细；ContractStarLicReward→ContractStarLicCfg", "填表": "数值"},
    "ActvSoccerLifeGrowthLevelCfg": {"用途": "生活等级、门票 buff、许可", "关联": "TicketCap/Recover 与附录常量叠加；不参与主角评分", "填表": "数值"},
    "ActvSoccerTeamCfg": {"用途": "球队展示（队名/队服/队标）", "关联": "←ContractCfg.TeamID；←LevelCfg.OpponentTeamID", "填表": "数值+美术"},
    "ActvSoccerContractStarLicCfg": {"用途": "联赛换约星级权重", "关联": "←Growth 两线 ContractStarLicReward", "填表": "数值"},
    "ActvSoccerContractCfg": {"用途": "合同待遇、赛季目标、发放场景", "关联": "TeamID→Team；首签←Nationality；换约←StarLic", "填表": "数值"},
    "ActvSoccerSeasonCfg": {"用途": "联赛轮次、NextSeason 链", "关联": "ID=LevelCfg.Group", "填表": "数值"},
    "ActvSoccerKnockoutCfg": {"用途": "淘汰赛开放、人数、分组、补位分", "关联": "OpenLeagueLevel→Level；→Phase/Simulation", "填表": "数值+活动"},
    "ActvSoccerKnockoutPhaseCfg": {"用途": "赛程阶段日期与当日内容", "关联": "←KnockoutCfg；BetOpen控制竞猜", "填表": "活动"},
    "ActvSoccerBetMultiplierCfg": {"用途": "胜率→奖励倍率对照表(=程序ChampionOddsCfg)", "关联": "同步匹配后查最近WinRatePctA；配合附录bet_*常量", "填表": "数值"},
    "ActvSoccerBetStakeTierCfg": {"用途": "投注档位(免费+5档)", "关联": "下注弹窗选项；免费档HitPayout=5", "填表": "数值"},
    "ActvSoccerAchieveCfg": {"用途": "活动成就(类型/计数/品质/图标/奖励)", "关联": "NameLcKey/DescLcKey→Language；Reward→ItemCfg/VM", "填表": "数值"},
    "ActvSoccerLanguageCfg": {"用途": "活动文案", "关联": "被所有 *LcKey 引用", "填表": "本地化"},
}

COMMON_PLATFORM_META: dict[str, str] = {
    "ActvSoccerBattlePassCfg": "BattlePassNew / ActivityBattlePass",
    "ActvSoccerExchangeShopCfg": "ExchangeShopItemCfg",
    "ActvSoccerGiftCfg": "GiftCfg / D2GiftCfg",
    "ActvSoccerRankSectionCfg": "ActivityRank / ActvRankSectionCfg",
}

SECTION_GROUPS: list[tuple[str, list[str]]] = [
    ("创角与引导", ["ActvSoccerCharacterCfg", "ActvSoccerNationalityCfg", "ActvSoccerTutorialCfg"]),
    (
        "切片与关卡",
        [
            "ActvSoccerSlicePresetCfg",
            "ActvSoccerSliceInstanceCfg",
            "ActvSoccerHapticCfg",
            "ActvSoccerLevelCfg",
            "ActvSoccerAiProfileCfg",
            "ActvSoccerEnemyAiCfg",
            "ActvSoccerAiModifierCfg",
            "ActvSoccerSliceFlowCfg",
        ],
    ),
    (
        "养成与合同",
        [
            "ActvSoccerCurrencyCfg",
            "ActvSoccerFameGrowthLevelCfg",
            "ActvSoccerLifeGrowthLevelCfg",
            "ActvSoccerTeamCfg",
            "ActvSoccerContractStarLicCfg",
            "ActvSoccerContractCfg",
        ],
    ),
    ("积分赛", ["ActvSoccerSeasonCfg"]),
    (
        "淘汰赛与赛程",
        ["ActvSoccerKnockoutCfg", "ActvSoccerKnockoutPhaseCfg"],
    ),
    (
        "竞猜",
        ["ActvSoccerBetMultiplierCfg", "ActvSoccerBetStakeTierCfg"],
    ),
    ("成就", ["ActvSoccerAchieveCfg"]),
    ("本地化", ["ActvSoccerLanguageCfg"]),
]


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def read_sheet_columns(wb_path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    cols: list[dict] = []
    max_col = ws.max_column or 0
    for col_idx in range(1, max_col + 1):
        field = _cell_str(ws.cell(3, col_idx).value)
        if not field:
            continue
        cols.append(
            {
                "field": field,
                "read": _cell_str(ws.cell(1, col_idx).value),
                "type": _cell_str(ws.cell(2, col_idx).value),
                "server": _cell_str(ws.cell(4, col_idx).value),
                "proto": _cell_str(ws.cell(5, col_idx).value),
                "desc": _cell_str(ws.cell(6, col_idx).value),
                "example": _cell_str(ws.cell(7, col_idx).value),
            }
        )
    wb.close()
    return cols


def field_table_md(columns: list[dict], *, include_read: bool = True) -> str:
    if include_read:
        header = "| 字段 | 类型 | 读取端 | 说明 | proto/示例 |\n|------|------|--------|------|------------|\n"
    else:
        header = "| 字段 | 类型 | 说明 | proto/示例 |\n|------|------|------|------------|\n"
    lines = [header]
    for col in columns:
        proto = col["proto"] or col["example"]
        if col["field"] == "Remark":
            continue
        if include_read:
            lines.append(
                f"| {col['field']} | {col['type']} | {col['read'] or '—'} | {col['desc']} | {proto or '—'} |\n"
            )
        else:
            lines.append(f"| {col['field']} | {col['type']} | {col['desc']} | {proto or '—'} |\n")
    return "".join(lines)


def appendix_in_match_md(items: list[dict]) -> str:
    lines = [
        "| ID | ItemKey | Effect | DefaultActive | FreeCount | 说明 |\n",
        "|----|---------|--------|---------------|-----------|------|\n",
    ]
    for row in items:
        lines.append(
            f"| {row['ID']} | `{row['ItemKey']}` | `{row['Effect']}` | {row['DefaultActive']} | {row['FreeCount']} | {row.get('Remark', '')} |\n"
        )
    return "".join(lines)


def appendix_const_md(const_rows: list[dict]) -> str:
    lines = [
        "| CfgID | Constant | Val | Array | 说明 |\n",
        "|-------|----------|-----|-------|------|\n",
    ]
    for row in const_rows:
        val = row.get("Val", "")
        arr = row.get("Array", "")
        comment = row.get("Comment", "")
        lines.append(
            f"| {row.get('CfgID', '—')} | `{row.get('Constant', '')}` | {val} | {arr or '—'} | {comment} |\n"
        )
    return "".join(lines)


def build_doc(summary: dict, sheet_columns: dict[str, list[dict]]) -> str:
    sheets = summary["sheets"]
    program_only = summary.get("sheets_program_only", [])
    common_platform = summary.get("sheets_common_platform", COMMON_PLATFORM_META)
    appendix_only = summary.get("appendix_only", [])
    in_match_items = summary.get("appendix_in_match_items", [])
    const_rows = summary.get("appendix_const_rows", [])
    sheet_count = len(sheets)

    lines: list[str] = [
        "# 2026世界杯主题活动 · 配置表结构（数值策划填表）\n",
        "\n",
        "> **定位**：仅描述数值/关卡/活动策划在 **`ActivitySoccer.xlsx`**（及语言表）中**需要填写**的配置表字段。  \n",
        "> **规则依据**：主策划案 `system-design-doc-samples/2026世界杯主题活动.md`。  \n",
        f"> **列级 SSOT**：以 `output/test-config/ActivitySoccer.xlsx` 页签表头为准（当前 **{sheet_count}** 个策划页签）；本文由 `generate_config_tables_doc.py` 自动生成字段表。  \n",
        f"> **生成时间**：脚本随测试配置同步更新。\n",
        "\n---\n\n",
        "## 文档边界\n\n",
        "### 本文档包含\n\n",
        "- 活动玩法、关卡、养成、淘汰赛赛程等**可配数值**。\n",
        "- 活动本地化（`ActvSoccerLanguageCfg`）。\n",
        "- **附录**：局内道具枚举、玩法常量（含竞猜 `ActvSoccer_bet_*` 系数）、合并 `dataconfig/ConstConfig.xlsx`。\n\n",
        "### 本文档不包含（由程序/服务端在开发阶段定义）\n\n",
        "| 类型 | 示例 | 说明 |\n",
        "|------|------|------|\n",
        "| 玩家存档 | `player_profile`、`current_contract_id`、`tutorial_done` | 运行时持久化，不进策划表 |\n",
        "| 局内/关卡运行时 | `slice_runtime`、`attempt_id`、`level_result` | 服务端/客户端内存态 |\n",
        "| 联赛/排行运行时 | `league_progress`、`league_rank_entry` | 服务端计算与存储 |\n",
        "| 淘汰赛运行时 | `knockout_team`、`knockout_match`、`qualifier_group_standing` | 组队/对阵/结果由服务生成 |\n",
        "| 竞猜运行时 | `bet_record`、`bet_stats`、动态 `win_rate`/`display_mult` | 下注与派彩过程数据 |\n",
        "| 程序只读枚举/类型 | `slice_type_def`（L1）、`operation_mode`、`player_ai_duty` | 程序定类型与判定；策划在 preset/instance 引用类型名 |\n",
        "| 客户端表现映射 | `character_state_config`（`ActvSoccerCharacterStateCfg`） | FSM 状态→动画 key；客户端+美术维护，读取端仅 `c` |\n",
        "| 知名度结算 | 独立 `fame_gain_rule` 表 | 由 `ContractCfg` 的 `PayFinish`/`PayGoal`/`PayAssist`/`PayFame` 按比赛结果结算 |\n",
        "| 淘汰赛演算 | `ActvSoccerMatchSimulationCfg` | 算法口径 `match_simulation_rule` 写在主策划案规则节；可调系数走 KnockoutCfg（如 `BotPlayerRating`） |\n",
        "| 竞猜币投放 | `ActvSoccerBetCoinSourceCfg` | 礼包走**通用** `GiftCfg.Reward`；每日免费见附录 `ActvSoccer_bet_daily_free_amount` |\n",
        "| 竞猜场次运行时 | `ActvSoccerBetMatchCfg` | 对阵/胜率/展示倍率由服务端按赛程生成 |\n",
        "| 接口/协议/DB | 字段名、proto、API | 技术派生文档 |\n\n",
    ]
    if program_only:
        lines.append(
            f"**不进策划 xlsx 的程序/运行时页签**（共 {len(program_only)} 个）："
            + "、".join(f"`{s}`" for s in program_only)
            + "。\n\n"
        )
    if common_platform:
        lines.append(
            "**走项目通用配套表**（不在 `ActivitySoccer.xlsx` 重复维护）："
            + "、".join(f"`{k}` → `{v}`" for k, v in common_platform.items())
            + "。\n\n"
        )
    if appendix_only:
        lines.append(
            f"**附录列举、不成活动页签**（共 {len(appendix_only)} 项）："
            + "、".join(f"`{s}`" for s in appendix_only)
            + "。\n\n"
        )
    lines += [
        "抽样/演算**算法**（如联赛换约权重、淘汰赛比分演算）写在主策划案规则节；本文只列**可调系数**所在配置表。\n\n",
        "---\n\n",
        "## 配置表概览\n\n",
        "### 配置文件清单\n\n",
        "| 文件 | 说明 | 策划维护 |\n",
        "|------|------|----------|\n",
        f"| `ActivitySoccer.xlsx` | 活动玩法主配置（{sheet_count} 个策划页签） | 数值 / 关卡 / 活动 |\n",
        "| `ActivitySoccerLanguage.xlsx` | 活动本地化 `ActvSoccerLanguageCfg` | 数值 + 本地化 |\n",
        "| `dataconfig/ConstConfig.xlsx` | 玩法全局常量（见本文附录） | 数值 |\n\n",
        "### 公共 / 外部依赖（非本活动 xlsx 内页签）\n\n",
        "| 类型 | 典型表 / 机制 | 本活动用法 | 策划动作 |\n",
        "|------|----------------|------------|----------|\n",
        "| **全局常量** | `dataconfig/ConstConfig.xlsx` → `ConstConfigCfg` | 门票默认、夹角、换约份数等 | 填附录常量表，合并全局 ConstConfig |\n",
        "| **通行证** | `BattlePassNew` / `ActivityBattlePass` | 活跃度升级、双轨奖励 | 在通用 BP 表按活动 ID 配置 |\n",
        "| **兑换商店** | `ExchangeShopItemCfg` | 竞猜币兑换外观/道具 | 在通用商店表配置消耗与奖励 |\n",
        "| **礼包** | `GiftCfg` / `D2GiftCfg` | 免费/付费礼包、竞猜币投放 | 在通用礼包表配置 Reward |\n",
        "| **排名奖励** | `ActivityRank` / `ActvRankSectionCfg` | 四榜段位与发奖 | 在通用排名表配置 RankType 与奖励 |\n",
        "| **奖励结构** | `TypIDVal_P_cspb`（proto） | BP/排名/商店/礼包/合同等 `ext[]` | 填 `typ`+`id`+`val` |\n",
        "| **全局道具** | `dataconfig/ItemCfg` | 兑换商店、BP 付费轨等 | **引用已有道具 ID** |\n",
        "| **活动内货币** | `ActvSoccerCurrencyCfg` | 金币/门票/竞猜币 | 待遇、价格、消耗在本活动表填数值 |\n",
        "| **本地化** | `ActvSoccerLanguageCfg` | 所有 `*LcKey` | 业务表只写 key，中文进语言表 |\n",
        "| **竞猜赔率** | `ActvSoccerBetMultiplierCfg` + 附录 `ActvSoccer_bet_*` | 胜率→倍率查表+公式 | 填对照表与常量 |\n",
        "| **投注档位** | `ActvSoccerBetStakeTierCfg` | 免费+50/100/150/200/300 | 填档位与免费派彩 |\n",
        "| **邮件补发** | 全局邮件模块 | 竞猜离线派彩、排名发奖 | 主案定规则 |\n",
        "| **活动入口** | 项目活动框架表 | 限时开关、活动 ID | 本活动 xlsx **不含**入口表 |\n",
        "| **每日任务→BP** | 项目 DailyTask / Quest 表 | BP 升级来源 | 通用 BP 表只配等级与奖励 |\n\n",
        "### 活动专用表一览\n\n",
        "| 页签 | 用途 | 主要关联 | 填表 |\n",
        "|------|------|----------|------|\n",
    ]
    for name in sheets:
        meta = SHEET_META.get(name, {"用途": "—", "关联": "—", "填表": "—"})
        lines.append(f"| `{name}` | {meta['用途']} | {meta['关联']} | {meta['填表']} |\n")
    lines.append("| `ActvSoccerLanguageCfg` | 活动文案 | 被所有 `*LcKey` 引用 | 本地化 |\n\n")

    lines += [
        "### 核心引用关系（简图）\n\n",
        "```mermaid\n",
        "flowchart LR\n",
        "  subgraph 创角\n",
        "    CH[CharacterCfg] --> NA[NationalityCfg]\n",
        "    NA -->|ContractPool| CO[ContractCfg]\n",
        "    TU[TutorialCfg] --> SI[SliceInstanceCfg]\n",
        "  end\n",
        "  subgraph 切片关卡\n",
        "    SP[SlicePresetCfg] --> SI\n",
        "    SI --> LV[LevelCfg]\n",
        "    SI --> AP[AiProfileCfg]\n",
        "    SI --> EA[EnemyAiCfg]\n",
        "    SI --> AM[AiModifierCfg]\n",
        "    AP[AiProfileCfg] --> LV\n",
        "    AP --> SA\n",
        "    EA[EnemyAiCfg] --> SA\n",
        "    AM[AiModifierCfg] --> SA\n",
        "    LV -->|Group| SE[SeasonCfg]\n",
        "    LV -->|OpponentTeamID| TM[TeamCfg]\n",
        "  end\n",
        "  subgraph 养成\n",
        "    FG[FameGrowthLevelCfg] --> CS[ContractStarLicCfg]\n",
        "    LG[LifeGrowthLevelCfg] --> CS\n",
        "    CS --> CO\n",
        "    CO --> TM\n",
        "  end\n",
        "  subgraph 通用配套\n",
        "    BP[BattlePassNew]\n",
        "    GF[GiftCfg]\n",
        "    EX[ExchangeShopItemCfg]\n",
        "    RK[ActivityRank]\n",
        "    RWD[TypIDVal奖励]\n",
        "    BP --> RWD\n",
        "    GF --> RWD\n",
        "    EX --> RWD\n",
        "    RK --> RWD\n",
        "  end\n",
        "  subgraph 公共\n",
        "    CC[ConstConfig]\n",
        "    LC[LanguageCfg]\n",
        "    IT[ItemCfg全局]\n",
        "    CC -.-> LV\n",
        "    LC -.-> CH\n",
        "    IT -.-> RWD\n",
        "  end\n",
        "```\n\n",
        "### 典型引用链（策划填表时按链检查）\n\n",
        "| 链路 | 路径 |\n",
        "|------|------|\n",
        "| 创角→首签 | `NationalityCfg.ContractPool` → `ContractCfg`（`GrantScene=first_sign`）→ `TeamCfg` |\n",
        "| 试训→切片 | `TutorialCfg.SliceInstanceID` → `SliceInstanceCfg` → `SlicePresetCfg` |\n",
        "| 关卡组装 | `LevelCfg.SliceList[]` → `SliceInstanceCfg`；AI 字段已并入实例表 |\n",
        "| 联赛轮次 | `SeasonCfg.ID` = `LevelCfg.Group`；`NextSeason` 链 |\n",
        "| 联赛换约 | `Growth.ContractStarLicReward` → `ContractStarLicCfg` → `ContractCfg`（`league_finish`） |\n",
        "| 待遇/知名度结算 | 比赛胜负/进球/助攻 → `ContractCfg` `PayFinish`/`PayGoal`/`PayAssist`/`PayFame` |\n",
        "| 淘汰赛→竞猜 | `KnockoutPhaseCfg.BetOpen` 排期 → 服务端 `bet_match`；演算胜率 → `BetMultiplierCfg` + `bet_*` 常量 |\n",
        "| 竞猜币来源 | 通用 `GiftCfg.Reward` + 附录 `bet_daily_free_amount`；消耗于下注与 `ExchangeShopItemCfg` |\n",
        "| 奖励发放 | 各表 `*Reward` / `TypIDVal` → `ItemCfg` 或活动 `vm` 货币 ID |\n\n",
        "---\n\n",
        "## 填表约定（K1 测试配置）\n\n",
        "| 约定 | 说明 |\n",
        "|------|------|\n",
        "| 首列 ID | 所有策划 sheet 首列 `ID`/`id` |\n",
        "| ext / ext[] | 第 5 行 proto 必填；空值 `{}` / `[]` |\n",
        "| 单参数 | 拆独立列，不包 JSON |\n",
        "| *LcKey | 业务表 string，中文进 `ActvSoccerLanguageCfg` |\n",
        "| 读取端 | 第 1 行 `c`/`s`/`cs`；`Remark` 留空 |\n",
        "| 参数叠加 | `preset → instance → ai_profile → modifier` |\n",
        f"| LcKey 格式 | `{summary.get('lc_id_format', '')}` |\n\n",
    ]

    ext_map = summary.get("ext_proto_map", {})
    if ext_map:
        lines.append("### ext proto 对照\n\n| proto | 字段 |\n|-------|------|\n")
        for proto, fields in ext_map.items():
            lines.append(f"| `{proto}` | {', '.join(fields)} |\n")
        lines.append("\n")

    id_ref = summary.get("id_cross_ref", {})
    if id_ref:
        lines.append("### ID 段交叉引用\n\n| 对象 | 段/说明 |\n|------|--------|\n")
        for key, val in id_ref.items():
            lines.append(f"| {key} | {val} |\n")
        lines.append("\n")

    lines.append("---\n\n")

    for section_name, sheet_names in SECTION_GROUPS:
        lines.append(f"## {section_name}\n\n")
        if section_name == "切片与关卡":
            lines += [
                "三层配置：**L1 slice_type（程序只读）→ L2 SlicePresetCfg → L3 SliceInstanceCfg**。  \n",
                "叠加优先级：`preset → instance.override → ai_profile → slice_ai.modifier`。  \n",
                "局内道具（哨子/回溯/瞄准）见**附录 A**，不在活动 xlsx 维护。\n\n",
            ]
        if section_name == "养成与合同":
            lines += [
                "主角评分**仅由知名度等级表** `ActvSoccerFameGrowthLevelCfg.PlayerRating` 投放；生活等级保持小量级、不参与评分。知名度 `Level` 后续可拆分为更细档位。\n\n",
            ]
        if section_name == "竞猜":
            lines += [
                "赔率公式：`mult = min(ActvSoccer_bet_constraint_return_rate / win_rate, ActvSoccer_bet_max_odds)`；胜率为 0 取 `max_odds`。  \n",
                "`ActvSoccerBetMultiplierCfg` 即程序 `ChampionOddsCfg`（胜率→倍率对照，16 行复用 X1）。  \n",
                "程序万分值与策划 Val 换算：`程序值 / 10000`（如 17500→1.75，500000→50）。\n\n",
            ]
        for sheet_name in sheet_names:
            meta = SHEET_META.get(sheet_name, {})
            lines.append(f"### {sheet_name}\n\n")
            if meta:
                lines.append(f"- **用途**：{meta.get('用途', '—')}  \n")
                lines.append(f"- **关联**：{meta.get('关联', '—')}  \n")
                lines.append(f"- **填表**：{meta.get('填表', '—')}\n\n")
            cols = sheet_columns.get(sheet_name, [])
            if cols:
                lines.append(field_table_md(cols, include_read=True))
                lines.append("\n")

    lines += [
        "---\n\n",
        "## 附录 A · 局内道具（`ActvSoccerInMatchItemCfg`）\n\n",
        "固定三项，程序按 `ItemKey`/`Effect` 读取；`FreeCount` 与生活等级 `FreeRewind` 等 buff 叠加。\n\n",
        appendix_in_match_md(in_match_items),
        "\n---\n\n",
        "## 附录 B · 玩法常量（合并 `ConstConfigCfg`）\n\n",
        f"共 **{len(const_rows)}** 项，字段对齐 `ConstConfigCfg`（`CfgID | Comment | Constant | Val | Array | Effects | requirement`）。  \n",
        f"CfgID 段：`{summary.get('const_cfgid_range', '—')}`；合并目标：`{summary.get('const_merge_target', 'dataconfig/ConstConfig.xlsx')}`。\n\n",
        appendix_const_md(const_rows),
    ]
    prog_map = summary.get("bet_program_const_map", {})
    gaps = summary.get("bet_config_gaps", [])
    if prog_map:
        lines += [
            "\n---\n\n",
            "## 附录 C · 竞猜常量程序对照（ConstCommon / controller）\n\n",
            "| 程序键 | 活动附录 Constant / 说明 |\n",
            "|--------|-------------------------|\n",
        ]
        for prog_key, actv_ref in prog_map.items():
            lines.append(f"| `{prog_key}` | {actv_ref} |\n")
    if gaps:
        lines += [
            "\n**待补项**：\n",
        ]
        for gap in gaps:
            lines.append(f"- {gap}\n")
    lines += [
        "\n---\n\n",
        "## 数值策划 TODO 清单（待填项汇总）\n\n",
        "- 门票上限/恢复/各关 TicketCost；生活等级 buff 对门票的影响列\n",
        "- 两条养成线等级曲线、ContractStarLicReward\n",
        "- Contract 待遇（PayFinish/PayGoal/PayAssist/PayFame）、SeasonGoal 阈值、Grant 门槛\n",
        "- ContractStarLic 各档 StarWeights\n",
        "- Knockout 开放关卡、人数、分组、阶段时刻、BotPlayerRating（演算公式见主案规则）\n",
        "- 通用 GiftCfg 竞猜币投放（`Reward` 含 `typ:vm` 竞猜币 ID）\n",
        "- 通用 ActivityRank 段位划分与奖励\n",
        "- 通用 BP / 商店 / 礼包各档数值\n",
        "- 切片摆位 AngleSpan* 基线、AI 难度档曲线\n",
        "- 附录常量合并进 `dataconfig/ConstConfig.xlsx`（含 `ActvSoccer_bet_*`）\n",
        "- BetMultiplier(=ChampionOddsCfg) 16 行、BetStakeTier 6 档投注\n\n",
        "---\n\n",
        "## 测试配置与正式表\n\n",
        "- 测试生成脚本：`output/test-config/generate_activity_soccer_test_config.py`\n",
        "- 派生文档脚本：`output/test-config/generate_config_tables_doc.py`\n",
        "- 汇总索引：`output/test-config/test-config-summary.json`\n",
        "- 合并正式 `dataconfig/` 前见 **config-table-editor** skill\n",
    ]
    return "".join(lines)


def main() -> None:
    summary = json.loads(SUMMARY.read_text(encoding="utf-8"))
    xlsx_path = XLSX
    if not xlsx_path.exists():
        xlsx_path = OUT_DIR / "ActivitySoccer.generated.xlsx"

    sheet_columns: dict[str, list[dict]] = {}
    for sheet in summary["sheets"]:
        sheet_columns[sheet] = read_sheet_columns(xlsx_path, sheet)
    sheet_columns["ActvSoccerLanguageCfg"] = read_sheet_columns(LC_XLSX, "ActvSoccerLanguageCfg")

    doc = build_doc(summary, sheet_columns)
    OUTPUT.write_text(doc, encoding="utf-8")
    SAMPLE_OUTPUT.write_text(doc, encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"Wrote {SAMPLE_OUTPUT}")
    print(f"Sheets documented: {len(summary['sheets'])} + LanguageCfg + appendix")


if __name__ == "__main__":
    main()
