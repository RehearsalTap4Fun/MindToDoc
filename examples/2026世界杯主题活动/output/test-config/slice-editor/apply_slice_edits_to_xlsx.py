"""把 slice-editor 网页保存的 patch json 写回 ActivitySoccer_preview.xlsx。

网页编辑器只编辑 5 个空间字段(BallPos / BallVector / BallOwner / PlayersInit / TargetPoint),
本脚本按 ID 找到对应行,直接覆盖这 5 个单元格。不读 Python specs、不做 normalize、
不补派生字段、未知 ID 跳过 + 计数告知。

用法:
    python apply_slice_edits_to_xlsx.py
    python apply_slice_edits_to_xlsx.py --dry-run
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
TEST_CONFIG_DIR = HERE.parent
XLSX_FILE = TEST_CONFIG_DIR / "ActivitySoccer_preview.xlsx"
PATCH_FILE = HERE / "slice-preset-edits.json"
SHEET_NAME = "ActvSoccerSlicePresetCfg"
FIELD_ROW = 3
DATA_START_ROW = 9
PATCH_FIELDS = ("BallPos", "BallVector", "BallOwner", "PlayersInit", "TargetPoint")


def _round_coord(v: float) -> float:
    return round(float(v), 1)


def _format_vec3(v: dict) -> str:
    return json.dumps(
        {"x": _round_coord(v["x"]), "y": _round_coord(v.get("y", 0)), "z": _round_coord(v["z"])},
        ensure_ascii=False,
    )


def _format_field(field: str, value):
    if value in (None, ""):
        return None
    if field == "BallOwner":
        return int(value)
    if field == "BallPos":
        return _format_vec3(value)
    if field in ("BallVector", "TargetPoint"):
        return json.dumps(value, ensure_ascii=False)
    if field == "PlayersInit":
        return json.dumps(value, ensure_ascii=False)
    return value


def apply(patch_path: Path = PATCH_FILE, xlsx_path: Path = XLSX_FILE, dry_run: bool = False) -> dict:
    if not patch_path.exists():
        return {"ok": False, "error": f"patch file not found: {patch_path}"}
    if not xlsx_path.exists():
        return {"ok": False, "error": f"xlsx not found: {xlsx_path}"}

    payload = json.loads(patch_path.read_text(encoding="utf-8"))
    if payload.get("schema") != "activity_soccer_slice_preset_edits.v1":
        return {"ok": False, "error": f"unknown schema: {payload.get('schema')}"}
    edits_by_id = {int(e["ID"]): e for e in payload.get("edits", [])}

    wb = load_workbook(xlsx_path)
    try:
        ws = wb[SHEET_NAME]
        field_cols = {ws.cell(FIELD_ROW, c).value: c for c in range(1, ws.max_column + 1)}
        id_col = field_cols.get("ID")
        if id_col is None:
            return {"ok": False, "error": f"{SHEET_NAME} 找不到 ID 列"}

        updated, skipped = [], []
        for r in range(DATA_START_ROW, ws.max_row + 1):
            rid = ws.cell(r, id_col).value
            if rid in (None, ""):
                continue
            edit = edits_by_id.pop(int(rid), None)
            if edit is None:
                continue
            for field in PATCH_FIELDS:
                if field not in edit:
                    continue
                col = field_cols.get(field)
                if col is None:
                    continue
                ws.cell(r, col).value = _format_field(field, edit[field])
            updated.append(int(rid))

        for pid in edits_by_id:
            skipped.append(pid)

        if not dry_run and updated:
            wb.save(xlsx_path)
    finally:
        wb.close()

    return {"ok": True, "updated": updated, "skipped_unknown_ids": skipped, "dry_run": dry_run}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--patch", type=Path, default=PATCH_FILE)
    parser.add_argument("--xlsx", type=Path, default=XLSX_FILE)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    result = apply(args.patch, args.xlsx, args.dry_run)
    if not result["ok"]:
        print(f"ERROR: {result['error']}", file=sys.stderr)
        return 1

    print(f"updated {len(result['updated'])} rows in {args.xlsx.name}{' (dry-run)' if args.dry_run else ''}")
    if result["skipped_unknown_ids"]:
        print(f"skipped {len(result['skipped_unknown_ids'])} unknown IDs (not in xlsx): {result['skipped_unknown_ids']}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
