#!/usr/bin/env python3
"""检测测试配置生成脚本与 xlsx 是否漂移。

用户经常直接手改 `output/test-config/ActivitySoccer_preview.xlsx`(改成项目真实数据),
导致脚本与 xlsx 长期偏离。等到下一次跑脚本时一覆盖,手改全丢。

本脚本:
  1. 在临时位置用 generate_activity_soccer_test_config.py 生成一份新 xlsx
  2. 与当前 ActivitySoccer_preview.xlsx 逐 sheet/字段/行做 diff
  3. 报告漂移(只在 user xlsx 里、只在 script xlsx 里、字段不一致、行内容不同)
  4. 退出码:0=完全一致 / 1=有漂移(适合 pre-commit hook 或 CI)

用法
----
直接跑(报告所有漂移,退出码反映状态):
    python scripts/check_xlsx_drift.py

只看摘要(屏蔽逐字段 diff):
    python scripts/check_xlsx_drift.py --summary

允许 script-only 的新增表(如新加的接球决策表用户还没合入):
    python scripts/check_xlsx_drift.py --allow-script-only ActvSoccerReceiveDecisionCfg ...

pre-commit 集成示例
-------------------
.git/hooks/pre-commit:
    #!/bin/sh
    python scripts/check_xlsx_drift.py --summary || {
        echo
        echo "[pre-commit] 测试配置脚本与 xlsx 已漂移。"
        echo "选项 1: 反推脚本到 xlsx → 修改 generate_activity_soccer_test_config.py"
        echo "选项 2: 跑一次脚本覆盖 xlsx → python output/test-config/generate_activity_soccer_test_config.py"
        echo "选项 3: 故意不一致(如临时占位) → git commit --no-verify"
        exit 1
    }

(此 hook 仅作示例,不自动安装。手动启用:把上面 sh 内容写到 .git/hooks/pre-commit
并 chmod +x。或在 .claude/settings.json 配置 PreCommit hook。)
"""
from __future__ import annotations

import argparse
import importlib.util
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SCRIPT = ROOT / "output" / "test-config" / "generate_activity_soccer_test_config.py"
XLSX = ROOT / "output" / "test-config" / "ActivitySoccer_preview.xlsx"


def load_workbook_dict(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        fields = [c.value for c in ws[3]]
        rows = []
        for r in range(9, ws.max_row + 1):
            d = {}
            empty = True
            for i, f in enumerate(fields, start=1):
                v = ws.cell(r, i).value
                if v not in (None, ""):
                    empty = False
                d[f] = v
            if not empty:
                rows.append(d)
        sheets[name] = {"fields": fields, "rows": rows}
    return sheets


def generate_baseline(out_path: Path) -> None:
    spec = importlib.util.spec_from_file_location("g", SCRIPT)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"无法加载 {SCRIPT}")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    lc = m.LcRegistry()
    wb = m.build_workbook(lc)
    wb.save(out_path)


def diff(user: dict, script: dict, allow_script_only: set[str], summary: bool) -> int:
    issues = 0
    sheets_user = set(user)
    sheets_script = set(script)

    only_user = sorted(sheets_user - sheets_script)
    only_script = sorted(sheets_script - sheets_user - allow_script_only)
    if only_user:
        issues += len(only_user)
        print(f"[!] 仅 user xlsx 有(脚本未生成): {only_user}")
    if only_script:
        issues += len(only_script)
        print(f"[!] 仅 script 生成(用户 xlsx 无): {only_script}")

    for name in sorted(sheets_user & sheets_script):
        u, s = user[name], script[name]
        sheet_issues = 0
        if u["fields"] != s["fields"]:
            added = [f for f in u["fields"] if f not in s["fields"]]
            removed = [f for f in s["fields"] if f not in u["fields"]]
            if added or removed:
                sheet_issues += 1
                print(f"[~] {name}: 字段不一致  user 多: {added}  script 多: {removed}")
        if len(u["rows"]) != len(s["rows"]):
            sheet_issues += 1
            print(f"[~] {name}: 行数 user={len(u['rows'])} script={len(s['rows'])}")
        cell_diffs = 0
        if u["fields"] == s["fields"] and len(u["rows"]) == len(s["rows"]):
            for i, (a, b) in enumerate(zip(u["rows"], s["rows"])):
                if a != b:
                    cell_diffs += 1
                    if not summary and cell_diffs <= 3:
                        idv = a.get("ID", a.get("CfgID", "?"))
                        ks = {k for k in (set(a) | set(b)) if a.get(k) != b.get(k)}
                        sample = "; ".join(f"{k}: {b.get(k)!r} -> {a.get(k)!r}" for k in ks)
                        print(f"    {name} ID={idv}: {sample}")
            if cell_diffs:
                sheet_issues += 1
                print(f"[~] {name}: {cell_diffs} 行有 cell 差异(显示前 3 条)")
        issues += sheet_issues

    return issues


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="检测脚本 vs xlsx 漂移")
    parser.add_argument("--summary", action="store_true", help="只报漂移摘要,不展示逐 cell diff")
    parser.add_argument("--allow-script-only", nargs="*", default=[],
                        help="允许只在 script 出现的 sheet 名(用户 xlsx 还没合入)")
    args = parser.parse_args(argv)

    if not XLSX.exists():
        print(f"找不到 {XLSX}", file=sys.stderr)
        return 2
    if not SCRIPT.exists():
        print(f"找不到 {SCRIPT}", file=sys.stderr)
        return 2

    with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
        baseline = Path(td) / "script.xlsx"
        try:
            generate_baseline(baseline)
        except Exception as e:
            print(f"生成脚本基线失败: {e}", file=sys.stderr)
            return 2

        user = load_workbook_dict(XLSX)
        script = load_workbook_dict(baseline)
        issues = diff(user, script, set(args.allow_script_only), args.summary)

    if issues:
        print(f"\n[FAIL] 共 {issues} 处漂移。")
        return 1
    print("[OK] 脚本与 xlsx 完全一致")
    return 0


if __name__ == "__main__":
    sys.exit(main())
