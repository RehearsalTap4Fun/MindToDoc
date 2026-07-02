# -*- coding: utf-8 -*-
"""生成 LevelTagCfg.xlsx 模板:LevelTags 500 行 + TagDef 词表。

策划在 LevelTags.Tags 列贴 tag,空行表示该关走 tier 默认。
TagDef 页由 level_tag_lib.TAG_REGISTRY 全量导出,作为人读对照与互斥校验依据。
首次运行后,后续运行需 force=True 才覆盖,避免吃掉已贴 tag。
"""
from __future__ import annotations

import math
import sys
from pathlib import Path

from openpyxl import Workbook

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE.parent))

import level_tag_lib  # noqa: E402

DEFAULT_OUT = HERE / "LevelTagCfg.xlsx"
ROUNDS_TOTAL = 50
LEVELS_PER_ROUND = 10


def _write_header(ws, fields: list[tuple[str, str, str]]) -> None:
    """主生成器约定:8 行表头(读取端/类型/字段名/server/4 行注释),数据从第 9 行起。"""
    for col_idx, (read, type_, field) in enumerate(fields, start=1):
        ws.cell(1, col_idx, read)
        ws.cell(2, col_idx, type_)
        ws.cell(3, col_idx, field)
        ws.cell(4, col_idx, "")
        ws.cell(5, col_idx, "")
        ws.cell(6, col_idx, "")
        ws.cell(7, col_idx, "")
        ws.cell(8, col_idx, "")


def _build_level_tags_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("LevelTags")
    _write_header(ws, [
        ("cs", "int", "ID"),
        ("c", "int", "Round"),
        ("c", "int", "LevelInRound"),
        ("c", "int", "Tier"),
        ("c", "string", "Tags"),
        ("-", "string", "Note"),
    ])
    row_idx = 9
    for r in range(1, ROUNDS_TOTAL + 1):
        for j in range(1, LEVELS_PER_ROUND + 1):
            level_id = (r - 1) * LEVELS_PER_ROUND + j
            tier = math.ceil(r / 5)
            ws.cell(row_idx, 1, level_id)
            ws.cell(row_idx, 2, r)
            ws.cell(row_idx, 3, j)
            ws.cell(row_idx, 4, tier)
            ws.cell(row_idx, 5, "")
            ws.cell(row_idx, 6, "")
            row_idx += 1


def _build_tag_def_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("TagDef")
    _write_header(ws, [
        ("c", "string", "Tag"),
        ("c", "string", "Affects"),
        ("c", "string", "MutexGroup"),
        ("c", "string", "Description"),
    ])
    row_idx = 9
    for tag in sorted(level_tag_lib.TAG_REGISTRY.keys()):
        spec = level_tag_lib.TAG_REGISTRY[tag]
        ws.cell(row_idx, 1, spec.name)
        ws.cell(row_idx, 2, ",".join(spec.affects))
        ws.cell(row_idx, 3, spec.mutex_group or "")
        ws.cell(row_idx, 4, spec.description)
        row_idx += 1


def generate(out: Path = DEFAULT_OUT, force: bool = False) -> Path:
    out = Path(out)
    if out.exists() and not force:
        raise FileExistsError(
            f"{out} 已存在;加 force=True 覆盖(会丢失已贴 tag)"
        )
    wb = Workbook()
    wb.remove(wb.active)
    _build_level_tags_sheet(wb)
    _build_tag_def_sheet(wb)
    wb.save(out)
    return out


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--force", action="store_true")
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = p.parse_args()
    path = generate(args.out, force=args.force)
    print(f"模板生成: {path}")
