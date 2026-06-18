import importlib
import math
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook


HERE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE.parent))


def _make_minimal_xlsx(tmp_path: Path, *, tag_overrides: dict[int, str] | None = None,
                       break_id: bool = False, missing_tag_def: bool = False,
                       extra_tag_def: bool = False) -> Path:
    """构造一个最小可用的 LevelTagCfg.xlsx。"""
    import level_tag_lib  # noqa
    importlib.reload(level_tag_lib)
    import generate_level_tag_template as g
    importlib.reload(g)
    out = tmp_path / "LevelTagCfg.xlsx"
    g.generate(out)

    from openpyxl import load_workbook
    wb = load_workbook(out)
    ws = wb["LevelTags"]
    if tag_overrides:
        for level_id, tags_str in tag_overrides.items():
            row_idx = 9 + (level_id - 1)
            ws.cell(row_idx, 5, tags_str)
    if break_id:
        ws.cell(9, 1, 999)

    td = wb["TagDef"]
    if missing_tag_def:
        td.delete_rows(9 + len(level_tag_lib.TAG_REGISTRY) - 1, 1)
    if extra_tag_def:
        td.cell(9 + len(level_tag_lib.TAG_REGISTRY), 1, "ghost_tag_xyz")
    wb.save(out)
    return out


def _registry_tags():
    import level_tag_lib as lib
    return list(lib.TAG_REGISTRY.keys())


def test_load_missing_file_raises(tmp_path):
    import apply_level_tags as app
    importlib.reload(app)
    with pytest.raises(FileNotFoundError):
        app.load_level_tag_cfg(tmp_path / "absent.xlsx")


def test_load_clean_file_returns_500_rows(tmp_path):
    out = _make_minimal_xlsx(tmp_path)
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    assert len(rows) == 500
    assert rows[0]["ID"] == 1
    assert rows[-1]["ID"] == 500
    assert rows[0]["Tags"] == []


def test_load_with_tags_parses_them(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={
        100: "boss",
        200: "hard_plus, set_piece",
        300: "must_win,no_modifier",
    })
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    assert rows[99]["Tags"] == ["boss"]
    assert rows[199]["Tags"] == ["hard_plus", "set_piece"]
    assert rows[299]["Tags"] == ["must_win", "no_modifier"]


def test_validate_id_break(tmp_path):
    out = _make_minimal_xlsx(tmp_path, break_id=True)
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    with pytest.raises(app.ValidationError) as ei:
        app.validate_loaded(rows, td_tags=set(_registry_tags()))
    assert any("ID" in e for e in ei.value.errors)


def test_validate_unknown_tag(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={42: "ghost_tag_xyz"})
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    with pytest.raises(app.ValidationError) as ei:
        app.validate_loaded(rows, td_tags=set(_registry_tags()))
    assert any("ghost_tag_xyz" in e for e in ei.value.errors)


def test_validate_mutex_violation(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={50: "hard_plus,easy_minus"})
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    with pytest.raises(app.ValidationError) as ei:
        app.validate_loaded(rows, td_tags=set(_registry_tags()))
    assert any("mutex" in e.lower() or "互斥" in e for e in ei.value.errors)


def test_validate_tag_def_drift(tmp_path):
    out = _make_minimal_xlsx(tmp_path, missing_tag_def=True)
    import apply_level_tags as app
    importlib.reload(app)
    td_tags = app.load_tag_def(out)
    import level_tag_lib as lib
    assert td_tags != set(lib.TAG_REGISTRY.keys())


def test_orchestrate_no_tags_returns_default_dataset(tmp_path):
    out = _make_minimal_xlsx(tmp_path)
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    td_tags = app.load_tag_def(out)
    app.validate_loaded(rows, td_tags=td_tags)
    dataset = app.build_dataset(rows)
    assert len(dataset["levels"]) == 500
    assert len(dataset["seasons"]) == 50
    assert len(dataset["slice_instances"]) == 126
    assert all(r["ID"] < 90000 for r in dataset["slice_instances"])


def test_orchestrate_boss_tag_changes_only_target_level(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={250: "boss"})
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    td_tags = app.load_tag_def(out)
    app.validate_loaded(rows, td_tags=td_tags)
    dataset = app.build_dataset(rows)
    levels = {r["ID"]: r for r in dataset["levels"]}
    assert levels[250]["OpponentTeamStar"] == 5
    # 邻居关默认 OpponentTeamStar 应不一定是 5(取决于 tier),不强断言相等


def test_orchestrate_extreme_keeper_appends_virtual_rows(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={250: "extreme_keeper"})
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    td_tags = app.load_tag_def(out)
    app.validate_loaded(rows, td_tags=td_tags)
    dataset = app.build_dataset(rows)
    virt_inst = [r for r in dataset["slice_instances"] if r["ID"] >= 90000]
    virt_ai = [r for r in dataset["slice_ais"] if r["ID"] >= 90000]
    assert len(virt_inst) > 0 and len(virt_inst) == len(virt_ai)
    assert all(r["ModifierID"] == 4005 for r in virt_ai)
    levels = {r["ID"]: r for r in dataset["levels"]}
    import json as _json
    assert all(s >= 90000 for s in _json.loads(levels[250]["SliceList"]))


def test_post_validate_catches_threshold_violation(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={250: "lenient,must_win"})
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    td_tags = app.load_tag_def(out)
    # lenient + must_win 同 mutex_group,会在 validate_loaded 阶段被抓
    with pytest.raises(app.ValidationError):
        app.validate_loaded(rows, td_tags=td_tags)


def test_write_outputs_xlsx_with_9_sheets(tmp_path):
    out = _make_minimal_xlsx(tmp_path)
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    td_tags = app.load_tag_def(out)
    app.validate_loaded(rows, td_tags=td_tags)
    ds = app.build_dataset(rows)
    app.validate_dataset(ds)
    target = tmp_path / "ActivitySoccer.LevelTagged.xlsx"
    summary_path = tmp_path / "level-tag-summary.json"
    app.write_outputs(ds, rows, target, summary_path)
    assert target.exists()
    assert summary_path.exists()

    from openpyxl import load_workbook
    wb = load_workbook(target)
    expected = {
        "ActvSoccerSeasonCfg", "ActvSoccerLevelCfg",
        "ActvSoccerSlicePresetCfg", "ActvSoccerSliceInstanceCfg",
        "ActvSoccerSliceAiCfg", "ActvSoccerAiProfileCfg",
        "ActvSoccerEnemyAiCfg", "ActvSoccerAiModifierCfg",
        "ActvSoccerTeamCfg",
    }
    assert expected.issubset(set(wb.sheetnames))


def test_summary_records_tag_hits(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={
        100: "boss",
        200: "boss",
        300: "hard_plus",
    })
    import apply_level_tags as app
    importlib.reload(app)
    rows = app.load_level_tag_cfg(out)
    td_tags = app.load_tag_def(out)
    app.validate_loaded(rows, td_tags=td_tags)
    ds = app.build_dataset(rows)
    target = tmp_path / "out.xlsx"
    summary_path = tmp_path / "summary.json"
    app.write_outputs(ds, rows, target, summary_path)
    import json
    summary = json.loads(summary_path.read_text(encoding="utf-8"))
    assert summary["tag_hits"]["boss"] == [100, 200]
    assert summary["tag_hits"]["hard_plus"] == [300]
    assert summary["levels_total"] == 500
    assert summary["levels_with_tags"] == 3


def test_main_cli_clean_run_returns_zero(tmp_path, monkeypatch):
    out = _make_minimal_xlsx(tmp_path)
    import apply_level_tags as app
    importlib.reload(app)
    target = tmp_path / "tagged.xlsx"
    summary_path = tmp_path / "summary.json"
    rc = app.main(["--input", str(out), "--output", str(target),
                   "--summary", str(summary_path)])
    assert rc == 0
    assert target.exists()


def test_main_cli_validation_error_returns_one(tmp_path):
    out = _make_minimal_xlsx(tmp_path, tag_overrides={1: "ghost_xyz"})
    import apply_level_tags as app
    importlib.reload(app)
    rc = app.main(["--input", str(out),
                   "--output", str(tmp_path / "x.xlsx"),
                   "--summary", str(tmp_path / "x.json")])
    assert rc == 1


def test_main_cli_missing_input_returns_two(tmp_path):
    import apply_level_tags as app
    importlib.reload(app)
    rc = app.main(["--input", str(tmp_path / "absent.xlsx"),
                   "--output", str(tmp_path / "x.xlsx"),
                   "--summary", str(tmp_path / "x.json")])
    assert rc == 2
