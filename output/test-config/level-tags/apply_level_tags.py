# -*- coding: utf-8 -*-
"""关卡 tag 配置工具入口:读取 LevelTagCfg.xlsx,生成关卡 tag 产物。"""
from __future__ import annotations

import math
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE.parent))

import level_tag_lib  # noqa: E402

DEFAULT_INPUT = HERE / "LevelTagCfg.xlsx"
ROUNDS_TOTAL = 50
LEVELS_PER_ROUND = 10


@dataclass
class ValidationError(Exception):
    errors: list[str] = field(default_factory=list)

    def __str__(self) -> str:
        return "\n".join(self.errors) or "(no errors)"


def _parse_tags(cell_value) -> list[str]:
    if cell_value is None:
        return []
    s = str(cell_value).strip()
    if not s:
        return []
    return [t.strip() for t in s.replace(",", " ").split() if t.strip()]


def load_level_tag_cfg(path: Path) -> list[dict]:
    """读取 LevelTags 页,返回 500 行 dict 列表。Tags 字段已解析为 list[str]。"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    wb = load_workbook(p, read_only=True)
    if "LevelTags" not in wb.sheetnames:
        raise ValueError(f"{p} 缺少 LevelTags 页")
    ws = wb["LevelTags"]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[0] is None:
            continue
        rows.append({
            "ID": int(row[0]),
            "Round": int(row[1]) if row[1] is not None else 0,
            "LevelInRound": int(row[2]) if row[2] is not None else 0,
            "Tier": int(row[3]) if row[3] is not None else 0,
            "Tags": _parse_tags(row[4]),
            "Note": row[5] or "",
        })
    return rows


def load_tag_def(path: Path) -> set[str]:
    p = Path(path)
    wb = load_workbook(p, read_only=True)
    if "TagDef" not in wb.sheetnames:
        raise ValueError(f"{p} 缺少 TagDef 页")
    ws = wb["TagDef"]
    tags: set[str] = set()
    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[0]:
            tags.add(str(row[0]))
    return tags


def validate_loaded(rows: list[dict], td_tags: set[str]) -> None:
    """加载阶段全量校验,错误统一收集后一次性抛出。"""
    errors: list[str] = []

    seen = set()
    for r in rows:
        if r["ID"] in seen:
            errors.append(f"ID 重复: {r['ID']}")
        seen.add(r["ID"])
    expected = set(range(1, ROUNDS_TOTAL * LEVELS_PER_ROUND + 1))
    missing = expected - seen
    extra = seen - expected
    if missing:
        errors.append(f"ID 缺失: {sorted(missing)[:10]}{'...' if len(missing) > 10 else ''}")
    if extra:
        errors.append(f"ID 越界: {sorted(extra)}")

    for r in rows:
        lid = r["ID"]
        if not (1 <= lid <= 500):
            continue
        expected_round = (lid - 1) // LEVELS_PER_ROUND + 1
        expected_lir = (lid - 1) % LEVELS_PER_ROUND + 1
        expected_tier = math.ceil(expected_round / 5)
        if r["Round"] != expected_round:
            errors.append(f"level {lid} Round 不一致: {r['Round']} vs {expected_round}")
        if r["LevelInRound"] != expected_lir:
            errors.append(f"level {lid} LevelInRound 不一致: {r['LevelInRound']} vs {expected_lir}")
        if r["Tier"] != expected_tier:
            errors.append(f"level {lid} Tier 不一致: {r['Tier']} vs {expected_tier}")

    lib_tags = set(level_tag_lib.TAG_REGISTRY.keys())
    if td_tags != lib_tags:
        only_def = td_tags - lib_tags
        only_lib = lib_tags - td_tags
        if only_def:
            errors.append(f"TagDef 多余 tag(lib 未注册): {sorted(only_def)}")
        if only_lib:
            errors.append(f"lib 注册但 TagDef 缺少: {sorted(only_lib)}")

    for r in rows:
        for tag in r["Tags"]:
            if tag not in level_tag_lib.TAG_REGISTRY:
                errors.append(f"level {r['ID']} 未知 tag: {tag}")
        groups: dict[str, list[str]] = {}
        for tag in r["Tags"]:
            spec = level_tag_lib.TAG_REGISTRY.get(tag)
            if spec and spec.mutex_group:
                groups.setdefault(spec.mutex_group, []).append(tag)
        for group, conflict in groups.items():
            if len(conflict) > 1:
                errors.append(f"level {r['ID']} mutex 组 {group} 冲突: {conflict}")

    if errors:
        raise ValidationError(errors)


import json as _json


def _import_main_generator():
    """惰性导入主生成器,避免顶层 import 顺序污染。"""
    sys.path.insert(0, str(HERE.parent))
    import generate_activity_soccer_test_config as g  # noqa
    return g


def _build_default_dataset() -> dict:
    """复用主生成器构造默认数据集(独立 LcRegistry,语言行不写出)。"""
    g = _import_main_generator()
    lc = g.LcRegistry()
    presets = g._build_presets(lc)
    instances = g._build_instance_library()
    levels = g._build_levels(lc)
    seasons = g._build_seasons(lc)
    teams = g._build_theme_teams(lc)
    return {
        "presets": presets,
        "slice_instances": instances,
        "ai_profiles": g._build_ai_profiles(),
        "enemy_ais": g._build_enemy_ai(),
        "ai_modifiers": g._build_ai_modifiers(),
        "teams": teams,
        "seasons": seasons,
        "levels": levels,
    }


def _slot_alloc(level_id: int):
    counter = {"v": 90000 + level_id * 10}
    def alloc():
        counter["v"] += 1
        return counter["v"]
    return alloc


def build_dataset(tag_rows: list[dict]) -> dict:
    """加载主生成器默认数据集,逐关 patch,返回带 9xxx 虚拟行的完整产物。"""
    ds = _build_default_dataset()
    levels_by_id = {r["ID"]: r for r in ds["levels"]}
    insts_by_id = {r["ID"]: r for r in ds["slice_instances"]}
    library_snapshot = {"insts": insts_by_id}

    for trow in tag_rows:
        if not trow["Tags"]:
            continue
        lid = trow["ID"]
        level_row = levels_by_id[lid]
        ctx = level_tag_lib.PatchContext(
            level_row=level_row,
            slice_instance_rows=ds["slice_instances"],
            new_id_alloc=_slot_alloc(lid),
            level_in_round=trow["LevelInRound"],
            tier=trow["Tier"],
            library=library_snapshot,
        )
        for tag in trow["Tags"]:
            level_tag_lib.TAG_REGISTRY[tag].patch(ctx)

    return ds


def validate_dataset(ds: dict) -> None:
    """生成阶段校验(spec §7)。"""
    errors: list[str] = []
    inst_ids = {r["ID"] for r in ds["slice_instances"]}

    for r in ds["levels"]:
        sl = _json.loads(r["SliceList"])
        n = len(sl)
        if not (0 < r["DrawThreshold"] < r["WinThreshold"] <= n):
            errors.append(
                f"level {r['ID']} 阈值非法: lose<draw<win<=n 不成立 "
                f"(draw={r['DrawThreshold']}, win={r['WinThreshold']}, n={n})"
            )
        for s in sl:
            if s not in inst_ids:
                errors.append(f"level {r['ID']} SliceList 含未注册 SliceInstance: {s}")
        if not (1001 <= r["AiProfileID"] <= 1010):
            errors.append(f"level {r['ID']} AiProfileID 越界: {r['AiProfileID']}")

    if errors:
        raise ValidationError(errors)


from collections import defaultdict


_SHEET_SCHEMA: dict[str, list[tuple[str, str, str]]] = {
    "ActvSoccerSeasonCfg": [
        ("cs", "int", "ID"), ("c", "string", "LeagueNameLcKey"),
        ("cs", "int", "NextSeason"), ("cs", "int", "ContractOfferCount"),
        ("-", "string", "Remark"),
    ],
    "ActvSoccerLevelCfg": [
        ("cs", "int", "ID"), ("cs", "bool", "IsTutorial"),
        ("cs", "int[]", "SliceList"), ("cs", "int", "AiProfileID"),
        ("cs", "int", "WinThreshold"), ("cs", "int", "DrawThreshold"),
        ("cs", "int", "TicketCost"), ("cs", "int", "OpponentTeamID"),
        ("cs", "int", "OpponentTeamStar"), ("cs", "int", "SeasonID"),
        ("-", "string", "Remark"),
    ],
    "ActvSoccerSliceInstanceCfg": [
        ("cs", "int", "ID"), ("cs", "string", "SliceType"),
        ("cs", "int", "PresetID"), ("c", "float", "OverrideOperableAngle"),
        ("cs", "string", "ObjectiveType"), ("cs", "ext[]", "ExtraObjectives"),
        ("cs", "ext[]", "Modifiers"),
        ("cs", "int", "AiProfileID"), ("cs", "int", "GoalkeeperAiID"),
        ("cs", "int", "DefenderAiID"), ("cs", "int", "ShooterAiID"),
        ("cs", "int", "ModifierID"), ("cs", "bool", "IsGuideAi"),
        ("cs", "bool", "RewindRandom"), ("cs", "int", "OverrideReactionTimeMs"),
        ("-", "string", "Remark"),
    ],
    "ActvSoccerAiProfileCfg": [
        ("cs", "int", "ID"), ("cs", "string", "Difficulty"),
        ("cs", "int", "GoalkeeperSaveRate"), ("cs", "int", "DefenderSuccessRate"),
        ("cs", "int", "ShooterSuccessRate"), ("cs", "int", "DeadCornerCanSave"),
        ("cs", "int", "ReactionTimeMs"), ("-", "string", "Remark"),
    ],
    "ActvSoccerEnemyAiCfg": [
        ("cs", "int", "ID"), ("cs", "int", "Duty"),
        ("cs", "int", "SaveWeight"), ("cs", "int", "LeftWeight"),
        ("cs", "int", "RightWeight"), ("cs", "int", "UpWeight"),
        ("cs", "int", "InterceptWeight"), ("cs", "int", "ClearanceWeight"),
        ("cs", "int", "KeeperCatchFail"), ("cs", "int", "OutOfBoundsFail"),
        ("c", "string", "AnimationKey"), ("-", "string", "Remark"),
    ],
    "ActvSoccerAiModifierCfg": [
        ("cs", "int", "ID"), ("cs", "string", "ModifierType"),
        ("cs", "string", "Param1Key"), ("cs", "string", "Param1Value"),
        ("cs", "string", "Param2Key"), ("cs", "string", "Param2Value"),
        ("cs", "string", "Param3Key"), ("cs", "string", "Param3Value"),
        ("-", "string", "Remark"),
    ],
    "ActvSoccerTeamCfg": [
        ("cs", "int", "ID"), ("c", "string", "NameLcKey"),
        ("cs", "string", "Region"), ("c", "string", "KitKey"),
        ("c", "string", "BadgeKey"), ("-", "string", "Remark"),
    ],
}


def _write_sheet(wb, name: str, schema: list[tuple[str, str, str]], rows: list[dict]) -> None:
    ws = wb.create_sheet(name)
    for col_idx, (read, type_, field) in enumerate(schema, start=1):
        ws.cell(1, col_idx, read)
        ws.cell(2, col_idx, type_)
        ws.cell(3, col_idx, field)
    for r_idx, row in enumerate(rows, start=9):
        for c_idx, (_, type_, field) in enumerate(schema, start=1):
            val = row.get(field)
            if val is None:
                if type_ == "ext[]":
                    val = "[]"
                elif type_ == "ext":
                    val = "{}"
            if val is not None:
                ws.cell(r_idx, c_idx, val)


def write_outputs(ds: dict, tag_rows: list[dict], target: Path, summary_path: Path) -> Path:
    """写 3 表 xlsx + summary.json。文件被占用时回退 *.generated.xlsx。

    tag 工具只输出关卡 tag 会直接修改或追加虚拟行的表。基础模板、AI 定义、
    modifier 定义、球队与赛季表均由主配置生成器维护，避免改 tag 时重写。
    """
    from openpyxl import Workbook
    target = Path(target); summary_path = Path(summary_path)

    wb = Workbook(); wb.remove(wb.active)
    _write_sheet(wb, "ActvSoccerLevelCfg", _SHEET_SCHEMA["ActvSoccerLevelCfg"], ds["levels"])
    _write_sheet(wb, "ActvSoccerSliceInstanceCfg", _SHEET_SCHEMA["ActvSoccerSliceInstanceCfg"], ds["slice_instances"])

    actual = target
    try:
        wb.save(actual)
    except PermissionError:
        actual = target.with_suffix(".generated.xlsx")
        wb.save(actual)
        print(f"[warn] {target} 被占用,回退写入 {actual}")

    tag_hits: dict[str, list[int]] = defaultdict(list)
    levels_with_tags = 0
    for trow in tag_rows:
        if trow["Tags"]:
            levels_with_tags += 1
            for tag in trow["Tags"]:
                tag_hits[tag].append(trow["ID"])
    summary = {
        "input": str(target),
        "levels_total": len(ds["levels"]),
        "levels_with_tags": levels_with_tags,
        "virtual_slice_instance_count": sum(1 for r in ds["slice_instances"] if r["ID"] >= 90000),
        "tag_hits": {k: sorted(v) for k, v in sorted(tag_hits.items())},
    }
    summary_path.write_text(_json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return actual


def main(argv: list[str] | None = None) -> int:
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=HERE / "ActivitySoccer.LevelTagged.xlsx")
    parser.add_argument("--summary", type=Path, default=HERE / "level-tag-summary.json")
    args = parser.parse_args(argv)

    try:
        rows = load_level_tag_cfg(args.input)
        td_tags = load_tag_def(args.input)
    except FileNotFoundError as e:
        print(f"[error] 输入文件不存在: {e}", file=sys.stderr)
        return 2
    except Exception as e:
        print(f"[error] 输入加载失败: {e}", file=sys.stderr)
        return 2

    try:
        validate_loaded(rows, td_tags=td_tags)
    except ValidationError as e:
        print("[error] 加载阶段校验失败:", file=sys.stderr)
        for err in e.errors:
            print(f"  - {err}", file=sys.stderr)
        return 1

    ds = build_dataset(rows)
    try:
        validate_dataset(ds)
    except ValidationError as e:
        print("[error] 生成阶段校验失败:", file=sys.stderr)
        for err in e.errors:
            print(f"  - {err}", file=sys.stderr)
        return 1

    actual = write_outputs(ds, rows, args.output, args.summary)
    tagged = sum(1 for r in rows if r["Tags"])
    print(f"[ok] 关卡 tag 产物写入: {actual} (贴 tag 关数 {tagged}/500)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
