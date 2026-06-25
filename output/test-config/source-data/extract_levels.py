"""从 ActivitySoccer_preview.xlsx 提取 LevelCfg 数据，写入 level.json。

工作流：
1. 策划/程序在 ActivitySoccer_preview.xlsx 中调整 ActvSoccerLevelCfg 表
2. 运行 `python source-data/extract_levels.py` 将改动同步回 level.json
3. 再次运行主生成器 `python generate_activity_soccer_test_config.py` 时，
   _build_levels() 会从 level.json 加载，输出与 xlsx 一致

同理可扩展为 extract_slice_preset.py / extract_slice_instance.py，但目前
preset / instance 仍由 Python 代码生成（编号规则稳定），无需走 JSON 同步。
"""
from __future__ import annotations

import json
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
XLSX = ROOT / "ActivitySoccer_preview.xlsx"
OUT = HERE / "level.json"


def dump_sheet(wb, name: str) -> dict:
    ws = wb[name]
    fields = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(9, ws.max_row + 1):
        rec = OrderedDict()
        for c, f in enumerate(fields, start=1):
            rec[f] = ws.cell(r, c).value
        rows.append(rec)
    return {"fields": fields, "rows": rows}


def main() -> None:
    if not XLSX.exists():
        raise FileNotFoundError(f"xlsx not found: {XLSX}")
    wb = load_workbook(XLSX, data_only=True)
    try:
        data = dump_sheet(wb, "ActvSoccerLevelCfg")
    finally:
        wb.close()
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} ({len(data['rows'])} rows)")


if __name__ == "__main__":
    main()
