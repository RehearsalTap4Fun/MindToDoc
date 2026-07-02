import importlib
import sys
from pathlib import Path

import openpyxl


def _run_template(tmp_path: Path):
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))  # output/test-config/
    import generate_level_tag_template as g
    importlib.reload(g)
    out = tmp_path / "LevelTagCfg.xlsx"
    g.generate(out)
    return out


def test_template_has_500_level_rows(tmp_path):
    out = _run_template(tmp_path)
    wb = openpyxl.load_workbook(out)
    assert "LevelTags" in wb.sheetnames
    ws = wb["LevelTags"]
    data_rows = [r for r in ws.iter_rows(min_row=9, values_only=True) if r[0] is not None]
    assert len(data_rows) == 500
    first = data_rows[0]
    assert first[0] == 1 and first[1] == 1 and first[2] == 1 and first[3] == 1
    last = data_rows[-1]
    assert last[0] == 500 and last[1] == 50 and last[2] == 10 and last[3] == 10


def test_template_tag_def_matches_registry(tmp_path):
    out = _run_template(tmp_path)
    wb = openpyxl.load_workbook(out)
    assert "TagDef" in wb.sheetnames
    ws = wb["TagDef"]
    rows = [r for r in ws.iter_rows(min_row=9, values_only=True) if r[0]]
    tags_in_sheet = {r[0] for r in rows}
    import level_tag_lib as lib
    assert tags_in_sheet == set(lib.TAG_REGISTRY.keys())


def test_template_refuses_overwrite(tmp_path):
    out = _run_template(tmp_path)
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    import generate_level_tag_template as g
    try:
        g.generate(out)
    except FileExistsError as e:
        assert "force" in str(e).lower()
    else:
        raise AssertionError("expected FileExistsError on second call without force")


def test_template_force_overwrites(tmp_path):
    out = _run_template(tmp_path)
    import generate_level_tag_template as g
    g.generate(out, force=True)
