# -*- coding: utf-8 -*-
"""Generate 2026 World Cup mini-soccer test config (K1 xlsx header format)."""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

OUT_DIR = Path(__file__).parent


def make_sheet(
    wb: Workbook,
    name: str,
    columns: list[dict],
    rows: list[dict],
) -> None:
    """Create sheet with K1 8-row header + data from row 9."""
    ws: Worksheet = wb.create_sheet(name)
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(1, col_idx, col.get("cs", "cs"))
        ws.cell(2, col_idx, col["type"])
        ws.cell(3, col_idx, col["field"])
        ws.cell(4, col_idx, col.get("server", ""))
        ws.cell(5, col_idx, col.get("comment1", ""))
        ws.cell(6, col_idx, col.get("comment2", ""))
        ws.cell(7, col_idx, col.get("comment3", ""))
        ws.cell(8, col_idx, col.get("comment4", ""))

    for row_idx, row in enumerate(rows, start=9):
        for col_idx, col in enumerate(columns, start=1):
            val = row.get(col["field"])
            if val is None:
                continue
            ws.cell(row_idx, col_idx, val)


def cols(*specs: tuple[str, str, str]) -> list[dict]:
    return [{"field": f, "type": t, "comment2": c} for f, t, c in specs]


def build_minisoccer_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    make_sheet(
        wb,
        "CharacterCfg",
        cols(
            ("CfgID", "int", "角色ID"),
            ("AppearanceKey", "string", "外观资源键"),
            ("DisplayPower", "int", "展示战力"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "AppearanceKey": "WC_Char_01", "DisplayPower": 1200, "Remark": "角色A"},
            {"CfgID": 2, "AppearanceKey": "WC_Char_02", "DisplayPower": 1200, "Remark": "角色B"},
            {"CfgID": 3, "AppearanceKey": "WC_Char_03", "DisplayPower": 1200, "Remark": "角色C"},
            {"CfgID": 4, "AppearanceKey": "WC_Char_04", "DisplayPower": 1200, "Remark": "角色D"},
        ],
    )

    make_sheet(
        wb,
        "NationalityCfg",
        cols(
            ("CfgID", "int", "国籍ID"),
            ("Name", "string", "国籍名"),
            ("Region", "string", "地区"),
            ("TeamPool", "int[]", "首签球队池"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "Name": "中国", "Region": "asia", "TeamPool": "[101,201,202]", "Remark": "测试"},
            {"CfgID": 2, "Name": "德国", "Region": "europe", "TeamPool": "[102,203,204]", "Remark": "测试"},
            {"CfgID": 3, "Name": "巴西", "Region": "south_america", "TeamPool": "[103,205,206]", "Remark": "测试"},
            {"CfgID": 4, "Name": "阿根廷", "Region": "south_america", "TeamPool": "[104,205,207]", "Remark": "测试"},
            {"CfgID": 5, "Name": "法国", "Region": "europe", "TeamPool": "[105,203,208]", "Remark": "测试"},
            {"CfgID": 6, "Name": "西班牙", "Region": "europe", "TeamPool": "[106,203,209]", "Remark": "测试"},
            {"CfgID": 7, "Name": "葡萄牙", "Region": "europe", "TeamPool": "[107,203,210]", "Remark": "测试"},
            {"CfgID": 8, "Name": "比利时", "Region": "europe", "TeamPool": "[108,203,211]", "Remark": "测试"},
        ],
    )

    make_sheet(
        wb,
        "TutorialCfg",
        cols(
            ("StepIndex", "int", "步骤序号"),
            ("SliceType", "string", "切片类型"),
            ("SliceInstanceID", "int", "切片实例"),
            ("ForcedOrder", "bool", "强制顺序"),
            ("Remark", "string", "备注"),
        ),
        [
            {"StepIndex": 1, "SliceType": "attack", "SliceInstanceID": 101, "ForcedOrder": 1, "Remark": "进攻教学"},
            {"StepIndex": 2, "SliceType": "free_kick", "SliceInstanceID": 102, "ForcedOrder": 1, "Remark": "任意球教学"},
            {"StepIndex": 3, "SliceType": "penalty", "SliceInstanceID": 103, "ForcedOrder": 1, "Remark": "点球教学"},
        ],
    )

    make_sheet(
        wb,
        "SlicePresetCfg",
        cols(
            ("CfgID", "int", "预设ID"),
            ("SliceType", "string", "切片类型"),
            ("Name", "string", "预设名"),
            ("Tags", "string[]", "标签"),
            ("BallPos", "string", "足球位置"),
            ("BallVector", "string", "足球向量"),
            ("BallOwner", "int", "控球球员"),
            ("PlayersInit", "string", "球员初始"),
            ("Camera", "string", "相机"),
            ("TargetPoint", "string", "目标点"),
            ("OperableAngle", "float", "可操作夹角"),
            ("TypePayload", "string", "类型参数"),
            ("RecommendedModes", "string[]", "建议模式"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgID": 1,
                "SliceType": "attack",
                "Name": "右路单刀",
                "Tags": '["side","distance","easy"]',
                "BallPos": '{"x":12,"y":0,"z":35}',
                "BallVector": '{"x":0,"y":0,"z":1}',
                "BallOwner": 0,
                "PlayersInit": '[{"team":"home","idx":0,"duty":3,"pos":{"x":12,"y":0,"z":35}},{"team":"away","idx":0,"duty":1,"pos":{"x":12,"y":0,"z":55}},{"team":"away","idx":1,"duty":2,"pos":{"x":8,"y":0,"z":48}}]',
                "Camera": '{"fov":45,"yaw":0}',
                "TargetPoint": '{"x":12,"y":0,"z":58}',
                "OperableAngle": 35.0,
                "TypePayload": '{"keeper_weight":5000,"angle":35}',
                "RecommendedModes": '["draw_line","slingshot"]',
                "Remark": "试训/引导复用",
            },
            {
                "CfgID": 2,
                "SliceType": "free_kick",
                "Name": "中路任意球",
                "Tags": '["center","medium"]',
                "BallPos": '{"x":0,"y":0,"z":42}',
                "BallVector": '{"x":0,"y":0,"z":1}',
                "BallOwner": 0,
                "PlayersInit": '[{"team":"home","idx":0,"duty":3,"pos":{"x":0,"y":0,"z":42}},{"team":"away","idx":0,"duty":1,"pos":{"x":0,"y":0,"z":58}},{"team":"away","idx":1,"duty":2,"pos":{"x":-2,"y":0,"z":50}},{"team":"away","idx":2,"duty":2,"pos":{"x":2,"y":0,"z":50}}]',
                "Camera": '{"fov":42,"yaw":0}',
                "TargetPoint": '{"x":0,"y":1.8,"z":58}',
                "OperableAngle": 28.0,
                "TypePayload": '{"wall_count":4,"keeper_weight":4500,"angle":28}',
                "RecommendedModes": '["draw_line","slingshot"]',
                "Remark": "试训任意球",
            },
            {
                "CfgID": 3,
                "SliceType": "penalty",
                "Name": "标准点球",
                "Tags": '["penalty","easy"]',
                "BallPos": '{"x":0,"y":0,"z":50}',
                "BallVector": '{"x":0,"y":0,"z":1}',
                "BallOwner": 0,
                "PlayersInit": '[{"team":"home","idx":0,"duty":3,"pos":{"x":0,"y":0,"z":50}},{"team":"away","idx":0,"duty":1,"pos":{"x":0,"y":0,"z":58}}]',
                "Camera": '{"fov":40,"yaw":0}',
                "TargetPoint": '{"x":0,"y":0.5,"z":58}',
                "OperableAngle": 0.0,
                "TypePayload": '{"keeper_dirs":[2500,2500,2500,2500],"reaction_ms":0}',
                "RecommendedModes": '["draw_line","slingshot"]',
                "Remark": "试训点球",
            },
            {
                "CfgID": 4,
                "SliceType": "goalkeep",
                "Name": "基础守门",
                "Tags": '["gk","easy"]',
                "BallPos": '{"x":0,"y":0,"z":56}',
                "BallVector": '{"x":0,"y":0,"z":-1}',
                "BallOwner": None,
                "PlayersInit": '[{"team":"home","idx":0,"duty":3,"pos":{"x":0,"y":0,"z":58}},{"team":"away","idx":0,"duty":2,"pos":{"x":0,"y":0,"z":56}}]',
                "Camera": '{"fov":50,"yaw":180}',
                "TargetPoint": None,
                "OperableAngle": 0.0,
                "TypePayload": '{"shot_dirs":[3000,3000,4000],"reaction_ms":2500}',
                "RecommendedModes": '["draw_line"]',
                "Remark": "第一关守门切片",
            },
        ],
    )

    make_sheet(
        wb,
        "SliceInstanceCfg",
        cols(
            ("CfgID", "int", "实例ID"),
            ("SliceType", "string", "切片类型"),
            ("PresetID", "int", "预设ID"),
            ("Overrides", "string", "覆盖"),
            ("TypePayload", "string", "类型参数覆盖"),
            ("Objectives", "string", "目标"),
            ("Modifiers", "string", "机制"),
            ("AiOverride", "string", "AI覆盖"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 101, "SliceType": "attack", "PresetID": 1, "Objectives": '[{"type":"score"}]', "Remark": "试训进攻"},
            {"CfgID": 102, "SliceType": "free_kick", "PresetID": 2, "Objectives": '[{"type":"score"}]', "Remark": "试训任意球"},
            {"CfgID": 103, "SliceType": "penalty", "PresetID": 3, "Objectives": '[{"type":"score"}]', "Remark": "试训点球"},
            {
                "CfgID": 201,
                "SliceType": "attack",
                "PresetID": 1,
                "Overrides": '{"operable_angle":30}',
                "Objectives": '[{"type":"score"}]',
                "Remark": "引导关切片1",
            },
            {
                "CfgID": 202,
                "SliceType": "attack",
                "PresetID": 1,
                "Overrides": '{"operable_angle":28}',
                "Objectives": '[{"type":"pass_to","params":{"target":1}},{"type":"score"}]',
                "Remark": "引导关切片2(助攻)",
            },
            {"CfgID": 203, "SliceType": "goalkeep", "PresetID": 4, "Objectives": '[{"type":"survive"}]', "Remark": "引导关守门"},
            {
                "CfgID": 301,
                "SliceType": "attack",
                "PresetID": 1,
                "Objectives": '[{"type":"score"}]',
                "Modifiers": '[{"id":"moving_keeper","params":{"speed":1.2,"range":2}}]',
                "Remark": "正式关2-切片1",
            },
            {
                "CfgID": 302,
                "SliceType": "penalty",
                "PresetID": 3,
                "Objectives": '[{"type":"score"}]',
                "Remark": "正式关2-切片2",
            },
        ],
    )

    make_sheet(
        wb,
        "LevelCfg",
        cols(
            ("CfgID", "int", "关卡ID"),
            ("IsTutorial", "bool", "引导关"),
            ("SliceList", "int[]", "切片序列"),
            ("AiProfileID", "int", "AI档位"),
            ("WinThreshold", "int", "胜利阈值"),
            ("DrawThreshold", "int", "平局阈值"),
            ("TicketCost", "int", "门票消耗"),
            ("OpponentTeamID", "int", "对手球队(队名/队服/队标)"),
            ("OpponentTeamStar", "int", "对手球队星级(球员属性计算依据)"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgID": 1,
                "IsTutorial": 1,
                "SliceList": "[201,202,203]",
                "AiProfileID": 1,
                "WinThreshold": 2,
                "DrawThreshold": 1,
                "TicketCost": 1,
                "OpponentTeamID": 201,
                "OpponentTeamStar": 1,
                "Remark": "第一关引导",
            },
            {
                "CfgID": 2,
                "IsTutorial": 0,
                "SliceList": "[301,302]",
                "AiProfileID": 1,
                "WinThreshold": 1,
                "DrawThreshold": 0,
                "TicketCost": 1,
                "OpponentTeamID": 202,
                "OpponentTeamStar": 1,
                "Remark": "CHN-L2 第1轮 第1场",
            },
            {
                "CfgID": 3,
                "IsTutorial": 0,
                "SliceList": "[301,301,302]",
                "AiProfileID": 2,
                "WinThreshold": 2,
                "DrawThreshold": 1,
                "TicketCost": 1,
                "OpponentTeamID": 203,
                "OpponentTeamStar": 2,
                "Remark": "CHN-L2 第1轮 第2场",
            },
        ],
    )

    make_sheet(
        wb,
        "AiProfileCfg",
        cols(
            ("CfgID", "int", "AI档位"),
            ("Difficulty", "string", "难度"),
            ("ParamOverrides", "string", "参数覆盖"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgID": 1,
                "Difficulty": "easy",
                "ParamOverrides": '{"keeper_weight_mul":0.9}',
                "Remark": "新手",
            },
            {
                "CfgID": 2,
                "Difficulty": "normal",
                "ParamOverrides": '{"keeper_weight_mul":1.0}',
                "Remark": "普通",
            },
        ],
    )

    make_sheet(
        wb,
        "TeamCfg",
        cols(
            ("CfgID", "int", "球队ID"),
            ("Name", "string", "球队名"),
            ("Star", "int", "星级"),
            ("Region", "string", "地区"),
            ("BadgeKey", "string", "队徽"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 101, "Name": "红星联", "Star": 1, "Region": "asia", "BadgeKey": "WC_Badge_Default", "Remark": "默认1星"},
            {"CfgID": 102, "Name": "莱茵青年", "Star": 1, "Region": "europe", "BadgeKey": "WC_Badge_DE", "Remark": "德国池"},
            {"CfgID": 103, "Name": "桑巴之星", "Star": 1, "Region": "south_america", "BadgeKey": "WC_Badge_BR", "Remark": "巴西池"},
            {"CfgID": 104, "Name": "蓝白雄鹰", "Star": 1, "Region": "south_america", "BadgeKey": "WC_Badge_AR", "Remark": "阿根廷池"},
            {"CfgID": 201, "Name": "海港FC", "Star": 1, "Region": "asia", "BadgeKey": "WC_Badge_CN_A", "Remark": "中国动态池A"},
            {"CfgID": 202, "Name": "东方之鹰", "Star": 1, "Region": "asia", "BadgeKey": "WC_Badge_CN_B", "Remark": "中国动态池B"},
            {"CfgID": 203, "Name": "北欧狼", "Star": 2, "Region": "europe", "BadgeKey": "WC_Badge_EU", "Remark": "联赛对手"},
            {"CfgID": 204, "Name": "阿尔卑斯", "Star": 2, "Region": "europe", "BadgeKey": "WC_Badge_EU2", "Remark": "联赛对手"},
        ],
    )

    make_sheet(
        wb,
        "SeasonCfg",
        cols(
            ("CfgID", "int", "赛季/大关卡ID"),
            ("LeagueCode", "string", "联赛代码"),
            ("SubLevelIDs", "int[]", "小关卡序列"),
            ("ContractOnFinish", "bool", "完成发合同"),
            ("UnlockPrevSeason", "int", "前置赛季"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgID": 1,
                "LeagueCode": "CHN-L2",
                "SubLevelIDs": "[1,2,3]",
                "ContractOnFinish": 1,
                "UnlockPrevSeason": 0,
                "Remark": "测试赛季第1轮",
            },
        ],
    )

    make_sheet(
        wb,
        "ContractCfg",
        cols(
            ("CfgID", "int", "合同ID"),
            ("TeamID", "int", "球队"),
            ("TeamStar", "int", "星级"),
            ("PayFinish", "int", "完赛待遇"),
            ("PayGoal", "int", "进球待遇"),
            ("PayAssist", "int", "助攻待遇"),
            ("SeasonGoal", "string", "赛季目标"),
            ("SeasonReward", "string", "目标奖励"),
            ("GrantLevelReq", "string", "发放门槛"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgID": 1,
                "TeamID": 101,
                "TeamStar": 1,
                "PayFinish": 100,
                "PayGoal": 50,
                "PayAssist": 30,
                "SeasonGoal": '[{"type":"goal","threshold":3,"settle_at":"season_end"}]',
                "SeasonReward": '[{"typ":"vm","id":11151001,"val":50}]',
                "GrantLevelReq": '{"fame":1,"life":1}',
                "Remark": "首签1星合同",
            },
            {
                "CfgID": 2,
                "TeamID": 203,
                "TeamStar": 2,
                "PayFinish": 200,
                "PayGoal": 80,
                "PayAssist": 50,
                "SeasonGoal": '[{"type":"slice_win","threshold":5,"settle_at":"season_end"}]',
                "SeasonReward": '[{"typ":"vm","id":11151001,"val":100}]',
                "GrantLevelReq": '{"fame":2,"life":2}',
                "Remark": "完成第1轮联赛后",
            },
        ],
    )

    make_sheet(
        wb,
        "GrowthLevelCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("GrowthType", "string", "养成线"),
            ("Level", "int", "等级"),
            ("ExpRequired", "int", "升级经验"),
            ("LevelUpReward", "string", "升级奖励"),
            ("LevelUpEffect", "string", "升级效果"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 101, "GrowthType": "fame", "Level": 1, "ExpRequired": 0, "LevelUpEffect": '{"team_star_unlock":1}', "Remark": "知名度Lv1"},
            {"CfgID": 102, "GrowthType": "fame", "Level": 2, "ExpRequired": 100, "LevelUpEffect": '{"team_star_unlock":2}', "Remark": "知名度Lv2"},
            {"CfgID": 201, "GrowthType": "life", "Level": 1, "ExpRequired": 0, "LevelUpEffect": '{"ticket_cap":10,"ticket_recover_min":30}', "Remark": "生活Lv1"},
            {
                "CfgID": 202,
                "GrowthType": "life",
                "Level": 2,
                "ExpRequired": 200,
                "LevelUpEffect": '{"ticket_cap":12,"ticket_recover_min":28,"free_rewind":1}',
                "Remark": "生活Lv2",
            },
        ],
    )

    make_sheet(
        wb,
        "InMatchItemCfg",
        cols(
            ("CfgID", "int", "道具"),
            ("ItemKey", "string", "道具键"),
            ("Effect", "string", "效果"),
            ("DefaultActive", "bool", "默认生效"),
            ("FreeCount", "int", "免费次数"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "ItemKey": "whistle", "Effect": "add_non_gk_slice", "DefaultActive": 1, "FreeCount": 0, "Remark": "哨子"},
            {"CfgID": 2, "ItemKey": "rewind", "Effect": "reset_slice", "DefaultActive": 0, "FreeCount": 1, "Remark": "回溯"},
            {"CfgID": 3, "ItemKey": "aim", "Effect": "aim_line", "DefaultActive": 1, "FreeCount": 0, "Remark": "瞄准"},
        ],
    )

    make_sheet(
        wb,
        "HapticCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("EventID", "string", "事件"),
            ("Category", "string", "分类"),
            ("Intensity", "string", "强度"),
            ("Pattern", "string", "图案"),
            ("DurationMs", "int", "时长ms"),
            ("MinIntervalMs", "int", "节流ms"),
            ("EnabledDefault", "bool", "默认开启"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "EventID": "charge_locked", "Category": "core_operation", "Intensity": "light", "Pattern": "tick", "DurationMs": 30, "MinIntervalMs": 200, "EnabledDefault": 1},
            {"CfgID": 2, "EventID": "ball_released", "Category": "core_operation", "Intensity": "medium", "Pattern": "tick", "DurationMs": 40, "MinIntervalMs": 200, "EnabledDefault": 1},
            {"CfgID": 3, "EventID": "slice_success", "Category": "key_result", "Intensity": "heavy", "Pattern": "double", "DurationMs": 80, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 4, "EventID": "saved", "Category": "key_result", "Intensity": "medium", "Pattern": "tick", "DurationMs": 50, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 5, "EventID": "out_of_bounds", "Category": "key_result", "Intensity": "medium", "Pattern": "tick", "DurationMs": 50, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 6, "EventID": "gk_timeout", "Category": "key_result", "Intensity": "heavy", "Pattern": "tick", "DurationMs": 60, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 7, "EventID": "gk_wrong_judge", "Category": "key_result", "Intensity": "medium", "Pattern": "double", "DurationMs": 70, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 8, "EventID": "gk_save", "Category": "key_result", "Intensity": "heavy", "Pattern": "double", "DurationMs": 80, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 9, "EventID": "hit_post", "Category": "key_result", "Intensity": "medium", "Pattern": "double", "DurationMs": 60, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"CfgID": 10, "EventID": "rewind_reset", "Category": "key_result", "Intensity": "light", "Pattern": "continuous", "DurationMs": 120, "MinIntervalMs": 500, "EnabledDefault": 1},
        ],
    )

    make_sheet(
        wb,
        "KnockoutCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("Phase", "string", "阶段"),
            ("StartTime", "string", "开始UTC"),
            ("EndTime", "string", "结束UTC"),
            ("GroupCount", "int", "海选组数G"),
            ("QualifyPerGroup", "int", "每组晋级K"),
            ("OpenLevel", "int", "开放关卡L"),
            ("TeamSizeMax", "int", "队伍人数M"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "Phase": "team_build", "StartTime": "2026-07-10 00:00:00", "EndTime": "2026-07-11 23:59:59", "OpenLevel": 3, "TeamSizeMax": 4, "Remark": "组队期"},
            {"CfgID": 2, "Phase": "qualifier", "StartTime": "2026-07-12 00:00:00", "EndTime": "2026-07-12 23:59:59", "GroupCount": 16, "QualifyPerGroup": 4, "Remark": "海选64强"},
            {"CfgID": 3, "Phase": "round_64", "StartTime": "2026-07-13 00:00:00", "EndTime": "2026-07-13 23:59:59", "Remark": "64强"},
            {"CfgID": 4, "Phase": "round_32", "StartTime": "2026-07-14 00:00:00", "EndTime": "2026-07-14 23:59:59", "Remark": "32强"},
            {"CfgID": 5, "Phase": "round_16", "StartTime": "2026-07-15 00:00:00", "EndTime": "2026-07-15 23:59:59", "Remark": "16强"},
            {"CfgID": 6, "Phase": "round_8", "StartTime": "2026-07-16 00:00:00", "EndTime": "2026-07-16 23:59:59", "Remark": "8强"},
            {"CfgID": 7, "Phase": "round_4", "StartTime": "2026-07-17 00:00:00", "EndTime": "2026-07-17 23:59:59", "Remark": "半决赛"},
            {"CfgID": 8, "Phase": "final", "StartTime": "2026-07-18 00:00:00", "EndTime": "2026-07-18 23:59:59", "Remark": "决赛"},
            {"CfgID": 9, "Phase": "showcase", "StartTime": "2026-07-19 00:00:00", "EndTime": "2026-07-19 23:59:59", "Remark": "展示期"},
        ],
    )

    make_sheet(
        wb,
        "BetCoinSourceCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("DailyFree", "int", "每日免费竞猜币"),
            ("GiftGrant", "int", "礼包投放"),
            ("LoseRecycleRate", "int", "失败回收万分比"),
            ("Remark", "string", "备注"),
        ),
        [{"CfgID": 1, "DailyFree": 100, "GiftGrant": 0, "LoseRecycleRate": 8000, "Remark": "测试默认"}],
    )

    make_sheet(
        wb,
        "GlobalConstCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("Constant", "string", "常量名"),
            ("Value", "string", "值"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "Constant": "TicketCapDefault", "Value": "10", "Remark": "门票上限"},
            {"CfgID": 2, "Constant": "TicketRecoverMinutes", "Value": "30", "Remark": "恢复间隔分钟"},
            {"CfgID": 3, "Constant": "KnockoutOpenLevel", "Value": "3", "Remark": "淘汰赛开放关卡"},
            {"CfgID": 4, "Constant": "BetStakeOptions", "Value": "[100,200,500,1000,5000]", "Remark": "竞猜快捷档位"},
            {"CfgID": 5, "Constant": "MoveSpeedIdle", "Value": "0", "Remark": "待机移动速度(固定值 TODO)"},
            {"CfgID": 6, "Constant": "MoveSpeedWalk", "Value": "2.5", "Remark": "慢走移动速度(固定值 TODO)"},
            {"CfgID": 7, "Constant": "MoveSpeedRunMin", "Value": "4", "Remark": "跑动速度下限"},
            {"CfgID": 8, "Constant": "MoveSpeedRunMax", "Value": "8", "Remark": "跑动速度上限"},
            {"CfgID": 9, "Constant": "MoveSpeedRatioIdle", "Value": "0", "Remark": "待机;0=读MoveSpeedIdle"},
            {"CfgID": 10, "Constant": "MoveSpeedRatioWalk", "Value": "0", "Remark": "慢走;0=读MoveSpeedWalk"},
            {"CfgID": 11, "Constant": "MoveSpeedRatioRun", "Value": "0", "Remark": "正常跑;0=能力值映射run速度"},
            {"CfgID": 12, "Constant": "MoveSpeedRatioJog", "Value": "0.75", "Remark": "慢跑;相对run倍率"},
            {"CfgID": 13, "Constant": "MoveSpeedRatioSprint", "Value": "1.25", "Remark": "冲刺;相对run倍率"},
            {"CfgID": 14, "Constant": "MoveSpeedRatioDribble", "Value": "0.85", "Remark": "带球推进;相对run倍率"},
            {"CfgID": 15, "Constant": "MoveSpeedRatioPress", "Value": "1.1", "Remark": "逼抢;相对run倍率"},
            {"CfgID": 16, "Constant": "MoveSpeedRatioKeeperLateral", "Value": "0.9", "Remark": "门将横移;相对run倍率"},
            {"CfgID": 17, "Constant": "KickForceMin", "Value": "10", "Remark": "出球力量下限"},
            {"CfgID": 18, "Constant": "KickForceMax", "Value": "25", "Remark": "出球力量上限"},
            {"CfgID": 19, "Constant": "BallControlDistance", "Value": "1.2", "Remark": "停球/控球距离;距离<此值时获得足球控制权"},
            {"CfgID": 20, "Constant": "OperableAngleSpanMin", "Value": "20", "Remark": "可操作夹角宽度下限(°)"},
            {"CfgID": 21, "Constant": "OperableAngleSpanMax", "Value": "70", "Remark": "可操作夹角宽度上限(°)"},
        ],
    )

    return wb


def build_k1_patch_workbook() -> Workbook:
    """K1 配套活动测试补丁（ActivityOnline / ActivityRank / Gift / BattlePassNew）。"""
    wb = Workbook()
    wb.remove(wb.active)

    # 活动 Type 为占位值，待程序注册 ActivityType 后替换
    ACT_MAIN = 190
    ACT_BP = 191
    ACT_RANK = 192
    ACT_GIFT = 193
    ACT_SHOP = 194
    CONTENT = 992026

    make_sheet(
        wb,
        "ActvOnlineCfg",
        cols(
            ("ID", "int", "活动ID"),
            ("ContentID", "int", "内容ID"),
            ("DropEventID", "int", "掉落ID"),
            ("SeasonId", "int[]", "赛季"),
            ("Available", "bool", "开启"),
            ("Type", "int", "活动类型"),
            ("Calendar", "bool", "日历"),
            ("Calendar2", "int", "入口"),
            ("ServerGroups", "string[]", "服务器组"),
            ("UnServerGroup", "string[]", "排除服"),
            ("Remark", "string", "备注"),
            ("HideForcast", "int", "隐藏预告"),
            ("CrossServerType", "int", "跨服类型"),
            ("IsCrossDay", "int", "跨天"),
            ("TriggerType", "int", "触发类型"),
            ("TriggerVal", "string", "触发值"),
            ("WeekTime", "string", "周时间"),
            ("Reopen", "bool", "循环"),
            ("TriggerRept", "string", "重复触发"),
            ("ForecastBeforeStart", "string", "预告"),
            ("Duration", "string", "持续"),
            ("CloseAfterEnd", "string", "结束后关闭"),
            ("LevelCondition", "int", "城堡等级条件"),
            ("LevelVal", "int", "城堡等级"),
            ("LevelShowup", "int", "出现等级"),
            ("ClientVersion", "string", "最低版本"),
        ),
        [
            {
                "ID": CONTENT,
                "ContentID": CONTENT,
                "DropEventID": 0,
                "SeasonId": "[1,2,3]",
                "Available": 1,
                "Type": ACT_MAIN,
                "Calendar": 0,
                "Calendar2": 2,
                "ServerGroups": '["test"]',
                "UnServerGroup": "[]",
                "Remark": "【测试】2026世界杯主题活动-主玩法",
                "HideForcast": 0,
                "CrossServerType": 0,
                "IsCrossDay": 0,
                "TriggerType": 1,
                "TriggerVal": "2026-06-08 00:00:00",
                "WeekTime": "[]",
                "Reopen": 0,
                "TriggerRept": "0d",
                "ForecastBeforeStart": "1d",
                "Duration": "30d",
                "CloseAfterEnd": "3d",
                "LevelCondition": 0,
                "LevelVal": 0,
                "LevelShowup": 1,
                "ClientVersion": "0.0.0",
            },
            {
                "ID": CONTENT + 1,
                "ContentID": CONTENT,
                "DropEventID": 0,
                "SeasonId": "[1,2,3]",
                "Available": 1,
                "Type": ACT_BP,
                "Calendar": 0,
                "Calendar2": 2,
                "ServerGroups": '["test"]',
                "UnServerGroup": "[]",
                "Remark": "【测试】世界杯BP通行证",
                "HideForcast": 0,
                "CrossServerType": 0,
                "IsCrossDay": 0,
                "TriggerType": 1,
                "TriggerVal": "2026-06-08 00:00:00",
                "WeekTime": "[]",
                "Reopen": 0,
                "TriggerRept": "0d",
                "ForecastBeforeStart": "0h",
                "Duration": "30d",
                "CloseAfterEnd": "3d",
                "LevelCondition": 0,
                "LevelVal": 0,
                "LevelShowup": 1,
                "ClientVersion": "0.0.0",
            },
            {
                "ID": CONTENT + 2,
                "ContentID": CONTENT,
                "DropEventID": 0,
                "SeasonId": "[1,2,3]",
                "Available": 1,
                "Type": ACT_GIFT,
                "Calendar": 0,
                "Calendar2": 2,
                "ServerGroups": '["test"]',
                "UnServerGroup": "[]",
                "Remark": "【测试】世界杯礼包",
                "HideForcast": 0,
                "CrossServerType": 0,
                "IsCrossDay": 0,
                "TriggerType": 1,
                "TriggerVal": "2026-06-08 00:00:00",
                "WeekTime": "[]",
                "Reopen": 0,
                "TriggerRept": "0d",
                "ForecastBeforeStart": "0h",
                "Duration": "30d",
                "CloseAfterEnd": "3d",
                "LevelCondition": 0,
                "LevelVal": 0,
                "LevelShowup": 1,
                "ClientVersion": "0.0.0",
            },
            {
                "ID": CONTENT + 3,
                "ContentID": CONTENT,
                "DropEventID": 0,
                "SeasonId": "[1,2,3]",
                "Available": 1,
                "Type": ACT_SHOP,
                "Calendar": 0,
                "Calendar2": 2,
                "ServerGroups": '["test"]',
                "UnServerGroup": "[]",
                "Remark": "【测试】世界杯竞猜兑换商店",
                "HideForcast": 0,
                "CrossServerType": 0,
                "IsCrossDay": 0,
                "TriggerType": 1,
                "TriggerVal": "2026-06-08 00:00:00",
                "WeekTime": "[]",
                "Reopen": 0,
                "TriggerRept": "0d",
                "ForecastBeforeStart": "0h",
                "Duration": "30d",
                "CloseAfterEnd": "3d",
                "LevelCondition": 0,
                "LevelVal": 0,
                "LevelShowup": 1,
                "ClientVersion": "0.0.0",
            },
        ],
    )

    make_sheet(
        wb,
        "ActvRankCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("ContentID", "int", "内容ID"),
            ("RankType", "string", "榜单类型"),
            ("RankName", "string", "榜单名"),
            ("CrossServer", "bool", "跨服"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "ContentID": CONTENT, "RankType": "slice_win", "RankName": "积分榜", "CrossServer": 1, "Remark": "切片胜利数"},
            {"CfgID": 2, "ContentID": CONTENT, "RankType": "goal", "RankName": "进球榜", "CrossServer": 1, "Remark": "积分赛进球"},
            {"CfgID": 3, "ContentID": CONTENT, "RankType": "bet_hit", "RankName": "竞猜命中榜", "CrossServer": 1, "Remark": "命中次数"},
            {"CfgID": 4, "ContentID": CONTENT, "RankType": "bet_profit", "RankName": "竞猜收益榜", "CrossServer": 1, "Remark": "总收益"},
        ],
    )

    make_sheet(
        wb,
        "ActvRankSectionCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("RankCfgID", "int", "榜单"),
            ("RankMin", "int", "名次下限"),
            ("RankMax", "int", "名次上限"),
            ("Reward", "string", "奖励"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "RankCfgID": 1, "RankMin": 1, "RankMax": 1, "Reward": '[{"typ":"item","id":5012109,"val":1}]', "Remark": "冠军段-测试"},
            {"CfgID": 2, "RankCfgID": 1, "RankMin": 2, "RankMax": 10, "Reward": '[{"typ":"vm","id":11151001,"val":200}]', "Remark": "前10-测试"},
            {"CfgID": 3, "RankCfgID": 1, "RankMin": 11, "RankMax": 100, "Reward": '[{"typ":"vm","id":11151001,"val":50}]', "Remark": "前100-测试"},
        ],
    )

    make_sheet(
        wb,
        "D2GiftCfg",
        cols(
            ("CfgId", "int", "礼包ID"),
            ("ContentID", "int", "活动ContentID"),
            ("GiftType", "int", "礼包类型"),
            ("Price", "int", "价格档位"),
            ("Reward", "string", "奖励"),
            ("BuyLimit", "int", "限购"),
            ("ParamList", "int", "折扣万分比"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgId": 99202601,
                "ContentID": CONTENT,
                "GiftType": 0,
                "Price": 0,
                "Reward": '[{"typ":"vm","id":11151001,"val":100}]',
                "BuyLimit": 1,
                "ParamList": 10000,
                "Remark": "免费每日礼包-测试",
            },
            {
                "CfgId": 99202602,
                "ContentID": CONTENT,
                "GiftType": 1,
                "Price": 1,
                "Reward": '[{"typ":"item","id":5012109,"val":1},{"typ":"vm","id":11151001,"val":300}]',
                "BuyLimit": 3,
                "ParamList": 10000,
                "Remark": "付费礼包-测试",
            },
        ],
    )

    make_sheet(
        wb,
        "BPShopCfg",
        cols(
            ("CfgID", "int", "编号"),
            ("ContentID", "int", "内容ID"),
            ("Level", "int", "BP等级"),
            ("ExpRequired", "int", "所需活跃度"),
            ("FreeReward", "string", "免费轨"),
            ("PaidReward", "string", "付费轨"),
            ("Remark", "string", "备注"),
        ),
        [
            {"CfgID": 1, "ContentID": CONTENT, "Level": 1, "ExpRequired": 100, "FreeReward": '[{"typ":"vm","id":11151001,"val":20}]', "PaidReward": '[{"typ":"vm","id":11151001,"val":50}]'},
            {"CfgID": 2, "ContentID": CONTENT, "Level": 2, "ExpRequired": 200, "FreeReward": '[{"typ":"vm","id":11151001,"val":30}]', "PaidReward": '[{"typ":"vm","id":11151001,"val":80}]'},
            {"CfgID": 3, "ContentID": CONTENT, "Level": 3, "ExpRequired": 300, "FreeReward": '[{"typ":"item","id":5012109,"val":1}]', "PaidReward": '[{"typ":"item","id":5012109,"val":2}]'},
        ],
    )

    make_sheet(
        wb,
        "ExchangeShopItemCfg",
        cols(
            ("CfgID", "int", "商品ID"),
            ("ContentID", "int", "内容ID"),
            ("CostItemID", "int", "消耗道具(竞猜币)"),
            ("CostVal", "int", "消耗数量"),
            ("Reward", "string", "兑换奖励"),
            ("BuyLimit", "int", "限购"),
            ("RefreshCycle", "string", "刷新周期"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "CfgID": 1,
                "ContentID": CONTENT,
                "CostItemID": 0,
                "CostVal": 500,
                "Reward": '[{"typ":"item","id":5012109,"val":1}]',
                "BuyLimit": 1,
                "RefreshCycle": "7d",
                "Remark": "世界杯表情-测试",
            },
            {
                "CfgID": 2,
                "ContentID": CONTENT,
                "CostItemID": 0,
                "CostVal": 1000,
                "Reward": '[{"typ":"item","id":5012109,"val":1}]',
                "BuyLimit": 1,
                "RefreshCycle": "7d",
                "Remark": "回溯道具包-测试",
            },
        ],
    )

    return wb


def export_json_summary(minisoccer_wb: Workbook, patch_wb: Workbook) -> None:
    summary = {
        "generated_for": "2026世界杯主题活动",
        "source_doc": "output/2026世界杯主题活动-开发文档.md",
        "files": [
            "ActivityMiniSoccer.xlsx",
            "WorldCup2026_K1ActivityPatch.xlsx",
        ],
        "test_flow": [
            "创角(4角色) -> 选国籍 -> 试训3切片(101-103)",
            "首签球队(101/201/202) -> 引导关Level1(201-203)",
            "积分赛Season1: Level1->2->3",
            "淘汰赛7.10-7.19时间表见KnockoutCfg",
        ],
        "placeholder_activity_types": {
            "main": 190,
            "bp": 191,
            "rank": 192,
            "gift": 193,
            "shop": 194,
            "content_id": 992026,
        },
        "notes": [
            "ActivityType 为占位值，程序注册后需替换",
            "奖励 item/vm ID 沿用 K1 测试道具，正式服需替换",
            "数值均为测试量级，标注 TODO 的字段已填合理占位",
        ],
    }
    (OUT_DIR / "test-config-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def main() -> None:
    mini_path = OUT_DIR / "ActivityMiniSoccer.xlsx"
    patch_path = OUT_DIR / "WorldCup2026_K1ActivityPatch.xlsx"

    mini_wb = build_minisoccer_workbook()
    mini_wb.save(mini_path)

    patch_wb = build_k1_patch_workbook()
    patch_wb.save(patch_path)

    export_json_summary(mini_wb, patch_wb)
    print(f"Wrote {mini_path}")
    print(f"Wrote {patch_path}")
    print(f"Wrote {OUT_DIR / 'test-config-summary.json'}")


if __name__ == "__main__":
    main()
