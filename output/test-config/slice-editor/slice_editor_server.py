from __future__ import annotations

import argparse
import json
import sys
import webbrowser
from datetime import datetime, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

HERE = Path(__file__).resolve().parent
TEST_CONFIG_DIR = HERE.parent
PROJECT_ROOT = TEST_CONFIG_DIR.parent.parent
PATCH_FILE = HERE / "slice-preset-edits.json"
XLSX_FILE = TEST_CONFIG_DIR / "ActivitySoccer_preview.xlsx"
DEFAULT_EDITOR_ANGLE = 120.0

sys.path.insert(0, str(TEST_CONFIG_DIR))

ALLOWED_EDIT_FIELDS = {
    "ID",
    "BallPos",
    "BallVector",
    "BallOwner",
    "PlayersInit",
    "TargetPoint",
}


def _json_or_none(value: Any) -> Any:
    if value in (None, ""):
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        return json.loads(value)
    return value


def _round_float(value: Any) -> float:
    return round(float(value), 3)


def _normalize_vec3(value: Any) -> dict[str, float]:
    parsed = _json_or_none(value)
    if not isinstance(parsed, dict):
        raise ValueError("vec3 field must be an object")
    return {
        "x": _round_float(parsed.get("x", 0)),
        "y": _round_float(parsed.get("y", 0)),
        "z": _round_float(parsed.get("z", 0)),
    }


def _normalize_players(value: Any) -> list[dict[str, Any]]:
    parsed = _json_or_none(value)
    if not isinstance(parsed, list):
        raise ValueError("PlayersInit must be a list")
    players: list[dict[str, Any]] = []
    for player in parsed:
        if not isinstance(player, dict):
            raise ValueError("each player must be an object")
        pos = _normalize_vec3(player.get("pos", {}))
        players.append({
            "team": str(player["team"]),
            "idx": int(player["idx"]),
            "duty": int(player["duty"]),
            "pos": pos,
            "facing": _round_float(player.get("facing", 0)),
        })
    return players


def _normalize_edit(edit: dict[str, Any]) -> dict[str, Any]:
    unknown = set(edit) - ALLOWED_EDIT_FIELDS
    if unknown:
        raise ValueError(f"unknown edit fields: {', '.join(sorted(unknown))}")
    if "ID" not in edit:
        raise ValueError("edit requires ID")

    normalized: dict[str, Any] = {"ID": int(edit["ID"])}
    for field, value in edit.items():
        if field == "ID":
            continue
        if field in {"BallPos", "BallVector", "TargetPoint"}:
            normalized[field] = None if value in (None, "") else _normalize_vec3(value)
        elif field == "PlayersInit":
            normalized[field] = _normalize_players(value)
        elif field == "BallOwner":
            normalized[field] = int(value)
        else:
            normalized[field] = _round_float(value)
    return normalized


def _sheet_rows(sheet_name: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_FILE, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        fields = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
        rows: list[dict[str, Any]] = []
        for row_idx in range(9, ws.max_row + 1):
            row = {
                field: ws.cell(row_idx, col_idx).value
                for col_idx, field in enumerate(fields, start=1)
                if field
            }
            if row.get("ID") is None:
                continue
            rows.append(row)
        return rows
    finally:
        wb.close()


def _normalize_patch_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "ID": int(row["ID"]),
        "BallPos": _normalize_vec3(row["BallPos"]),
        "BallVector": _normalize_vec3(row["BallVector"]),
        "BallOwner": int(row["BallOwner"]),
        "PlayersInit": _normalize_players(row["PlayersInit"]),
        "TargetPoint": None if row.get("TargetPoint") in (None, "") else _normalize_vec3(row["TargetPoint"]),
    }


def build_patch_from_xlsx() -> dict[str, Any]:
    return {
        "schema": "activity_soccer_slice_preset_edits.v1",
        "savedAt": datetime.now(timezone.utc).isoformat(),
        "source": XLSX_FILE.name,
        "edits": [_normalize_patch_row(row) for row in _sheet_rows("ActvSoccerSlicePresetCfg")],
    }


def overwrite_patch_from_xlsx(patch_path: Path = PATCH_FILE) -> dict[str, Any]:
    patch = build_patch_from_xlsx()
    patch_path.parent.mkdir(parents=True, exist_ok=True)
    patch_path.write_text(json.dumps(patch, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "saved": len(patch["edits"]), "path": str(patch_path)}


def build_editor_payload() -> dict[str, Any]:
    import generate_activity_soccer_test_config as g

    presets = _sheet_rows("ActvSoccerSlicePresetCfg")
    for row in presets:
        row["ID"] = int(row["ID"])
        row["BallOwner"] = int(row["BallOwner"])
        row["BallPosParsed"] = _normalize_vec3(row["BallPos"])
        row["BallVectorParsed"] = _normalize_vec3(row["BallVector"])
        row["TargetPointParsed"] = _json_or_none(row.get("TargetPoint"))
        row["PlayersInitParsed"] = _normalize_players(row["PlayersInit"])
        row["TagsParsed"] = _json_or_none(row.get("Tags")) or []
        row["RecommendedModesParsed"] = _json_or_none(row.get("RecommendedModes")) or []

    return {
        "schema": "activity_soccer_slice_editor_payload.v1",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "default_angle": DEFAULT_EDITOR_ANGLE,
        "coordinate_protocol": {
            "field_x_half": g.FIELD_X_HALF,
            "field_z_near": g.FIELD_Z_NEAR,
            "field_z_mid": g.FIELD_Z_MID,
            "field_z_far": g.FIELD_Z_FAR,
            "goal_width": g.GOAL_WIDTH,
            "goal_height": g.GOAL_HEIGHT,
            "penalty_area_z_far": g.PENALTY_AREA_Z_FAR,
            "penalty_area_x_half": g.PENALTY_AREA_X_HALF,
            "goal_area_z_far": g.GOAL_AREA_Z_FAR,
            "ball_control_distance": g.BALL_CONTROL_DISTANCE,
        },
        "slice_type_order": ["attack", "free_kick", "corner", "throw_in", "penalty", "goalkeep"],
        "presets": presets,
    }


def save_edit_patch(payload: dict[str, Any], patch_path: Path = PATCH_FILE) -> dict[str, Any]:
    edits = payload.get("edits")
    if not isinstance(edits, list):
        raise ValueError("payload requires edits list")
    normalized_edits = [_normalize_edit(edit) for edit in edits]
    patch = {
        "schema": "activity_soccer_slice_preset_edits.v1",
        "savedAt": datetime.now(timezone.utc).isoformat(),
        "source": "slice-editor",
        "edits": normalized_edits,
    }
    patch_path.parent.mkdir(parents=True, exist_ok=True)
    patch_path.write_text(json.dumps(patch, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "saved": len(normalized_edits), "path": str(patch_path)}


class SliceEditorHandler(BaseHTTPRequestHandler):
    def _send_json(self, status: int, payload: dict[str, Any]) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, path: Path) -> None:
        if not path.exists() or not path.is_file():
            self._send_json(404, {"ok": False, "error": "file not found"})
            return
        content_type = {
            ".html": "text/html; charset=utf-8",
            ".css": "text/css; charset=utf-8",
            ".js": "text/javascript; charset=utf-8",
        }.get(path.suffix, "application/octet-stream")
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:  # noqa: N802
        request_path = urlparse(self.path).path
        if request_path == "/api/presets":
            try:
                self._send_json(200, build_editor_payload())
            except Exception as exc:  # pragma: no cover - surfaced in browser
                self._send_json(500, {"ok": False, "error": str(exc)})
            return
        if request_path == "/":
            request_path = "/index.html"
        safe_name = request_path.strip("/")
        if "/" in safe_name or "\\" in safe_name:
            self._send_json(404, {"ok": False, "error": "file not found"})
            return
        self._send_file(HERE / safe_name)

    def do_POST(self) -> None:  # noqa: N802
        request_path = urlparse(self.path).path
        if request_path != "/api/save-edits":
            self._send_json(404, {"ok": False, "error": "unknown endpoint"})
            return
        try:
            size = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(size).decode("utf-8"))
            self._send_json(200, save_edit_patch(payload))
        except Exception as exc:
            self._send_json(400, {"ok": False, "error": str(exc)})


def main() -> None:
    parser = argparse.ArgumentParser(description="Serve the Activity Soccer slice editor.")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--open", action="store_true")
    parser.add_argument("--sync-from-xlsx", action="store_true", help="Overwrite slice-preset-edits.json from ActivitySoccer_preview.xlsx and exit.")
    args = parser.parse_args()

    if args.sync_from_xlsx:
        result = overwrite_patch_from_xlsx()
        print(json.dumps(result, ensure_ascii=False))
        return

    server = ThreadingHTTPServer(("127.0.0.1", args.port), SliceEditorHandler)
    url = f"http://127.0.0.1:{args.port}/"
    print(f"Slice editor running at {url}")
    print(f"Serving script: {Path(__file__).resolve()}")
    print(f"Handler: {SliceEditorHandler.__name__}")
    if args.open:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping slice editor")


if __name__ == "__main__":
    main()
