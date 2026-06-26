# -*- coding: utf-8 -*-
"""Generate ActivitySoccer_preview.xlsx test config from 2026 World Cup DingTalk doc.

新增一张 sheet 时,以下位置必须同步更新(否则 summary/读取端/分组会漏):
  1. build_workbook(): 在合适的 # --- 分节注释下加 make_sheet(...)
  2. SHEET_DEFAULT_READ / READ_OVERRIDES (本文件顶部): 设读取端,默认 cs
  3. SHEETS_PROGRAM_ONLY (顶部): 仅程序/服务端表才登记,数值表不要登记
  4. export_summary().sheet_groups: 加进对应分组(或新建分组)
  5. export_summary().id_cross_ref: 登记本表 ID 段(供下次新增/排错对照)
新增引用列时,记得在被引用表确实存在对应行,以及在 README 的 ID 段交叉引用表回贴一行。
"""
from __future__ import annotations

import json
import math
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

OUT_DIR = Path(__file__).parent
PROJECT_ROOT = OUT_DIR.parent.parent
OUTPUT_FILE = OUT_DIR / "ActivitySoccer_preview.xlsx"
OUTPUT_LC_FILE = OUT_DIR / "ActivitySoccerLanguage.xlsx"
REFERENCE_PRESETS_FILE = OUT_DIR / "slice-reference-presets.json"
LATEST_ACTIVITY_SOCCER_XLSX = Path("C:/Project/K1Dataconfig/dataconfig/ActivitySoccer.xlsx")
GUIDE_STEP_SOURCE = PROJECT_ROOT / "input" / "世界杯引导.xlsx"
# 追加到 dataconfig/ConstConfig.xlsx 的 CfgID 段（当前线上最大约 10135564）
ACTV_SOCCER_CONST_CFGID_BASE = 10135601
SOURCE_DOCS = [
    "钉钉文档 2026世界杯主题活动-开发文档 (amweZ92PV6vDZdmDCKwo2Ev4VxEKBD6p)",
    "K1Client docs/plans/2026-06-09-worldcup-fsm-bt-enemy-ai-design.md",
]
# 程序/运行时定义，不进数值策划填表 xlsx（见 2026世界杯主题活动-配置表结构.md §文档边界）
SHEETS_PROGRAM_ONLY = (
    "ActvSoccerSliceTypeDefCfg",      # L1 slice_type 程序只读枚举
    "ActvSoccerPlayerAiDutyEnumCfg",  # player_ai_duty 程序只读枚举
    "ActvSoccerBetMatchCfg",          # 正式对阵/赔率由服务端生成
    "ActvSoccerCharacterStateCfg",    # FSM状态→动画映射，客户端+美术维护
)
# 走项目通用配套表，不在本活动 xlsx 重复维护
SHEETS_COMMON_PLATFORM = (
    "ActvSoccerBattlePassCfg",        # → BattlePassNew / ActivityBattlePass
    "ActvSoccerExchangeShopCfg",      # → ExchangeShopItemCfg
    "ActvSoccerGiftCfg",              # → GiftCfg / D2GiftCfg
    "ActvSoccerRankSectionCfg",       # → ActivityRank / ActvRankSectionCfg
)
# 附录列举，不成活动 xlsx 页签
APPENDIX_ONLY = (
    "ActvSoccerInMatchItemCfg",       # 局内道具，见配置表结构文档附录
    "ActvSoccerGlobalConstCfg",       # 玩法常量，合并 dataconfig/ConstConfig.xlsx
)

IN_MATCH_ITEMS: list[dict] = [
    {"ID": 1, "ItemKey": "whistle", "Effect": "add_non_gk_slice", "DefaultActive": 1, "FreeCount": 0, "Remark": "哨子"},
    {"ID": 2, "ItemKey": "rewind", "Effect": "reset_slice", "DefaultActive": 0, "FreeCount": 1, "Remark": "回溯"},
    {"ID": 3, "ItemKey": "aim", "Effect": "aim_line", "DefaultActive": 1, "FreeCount": 0, "Remark": "瞄准"},
]

# 球员 AI 职责枚举（程序侧 PlayerAiDuty）
# 约定：守门员；对方除门将外均为后卫；我方所有球员(含玩家)均为前锋
PLAYER_AI_DUTY_ENUM: dict[str, int] = {
    "Goalkeeper": 1,
    "Defender": 2,
    "Forward": 3,
}
PLAYER_AI_DUTY_ENUM_COMMENT = "1=Goalkeeper,2=Defender,3=Forward"


# 第1行读取端: c=仅客户端 s=仅服务端 cs=双端; Remark 留空(备注列)
SHEET_DEFAULT_READ: dict[str, str] = {
    "ActvSoccerTutorialCfg": "c",
    "ActvSoccerHapticCfg": "c",
    "ActvSoccerFirstTouchAnimCfg": "c",
    "ActvSoccerSliceFlowCfg": "c",
    "ActvSoccerContractCfg": "cs",
    "ActvSoccerContractStarLicCfg": "s",
    "ActvSoccerKnockoutCfg": "s",
    "ActvSoccerKnockoutPhaseCfg": "s",
    "ActvSoccerBetMultiplierCfg": "cs",
    "ActvSoccerBetStakeTierCfg": "cs",
}
READ_OVERRIDES: dict[str, dict[str, str]] = {
    "ActvSoccerCharacterCfg": {"AppearanceKey": "c", "DisplayPower": "c"},
    "ActvSoccerNationalityCfg": {"NameLcKey": "c", "ContractPool": "s"},
    "ActvSoccerTutorialCfg": {"SliceInstanceID": "cs", "DescLcKey": "c"},
    "ActvSoccerGuideStepCfg": {
        "DialogueLcKey": "c", "TextStyle": "c",
        "FocusTarget": "c", "MaskType": "c", "GestureDesc": "c",
        "WaitType": "c", "WaitTarget": "c", "SkipPolicy": "c",
    },
    "ActvSoccerSlicePresetCfg": {
        "NameLcKey": "c", "Tags": "c", "BallPos": "c", "BallVector": "c", "BallOwner": "c",
        "PlayersInit": "c", "CameraFov": "c", "TargetPoint": "c", "RecommendedModes": "c",
        "AngleSpanMin": "c", "AngleSpanMax": "c", "AngleMaxCenterShift": "c", "AngleMargin": "c",
    },
    "ActvSoccerSliceInstanceCfg": {"OverrideOperableAngle": "c"},
    "ActvSoccerReceiveDecisionCfg": {"HighBallHeight": "c"},
    "ActvSoccerEnemyAiCfg": {"AnimationKey": "c"},
    "ActvSoccerTeamCfg": {"NameLcKey": "c", "KitKey": "c", "BadgeKey": "c"},
    "ActvSoccerFameGrowthLevelCfg": {"TitleLcKey": "c"},
    "ActvSoccerLifeGrowthLevelCfg": {"TitleLcKey": "c"},
    "ActvSoccerSeasonCfg": {"LeagueNameLcKey": "c"},
    "ActvSoccerKnockoutPhaseCfg": {
        "PhaseLcKey": "cs", "PhaseKey": "cs", "DayContentLcKey": "c", "BetOpen": "cs",
    },
}


def _resolve_read(sheet_name: str, field: str) -> str:
    if field == "Remark" or (sheet_name == "ActvSoccerLanguageCfg" and field == "Source"):
        return ""
    overrides = READ_OVERRIDES.get(sheet_name, {})
    if field in overrides:
        return overrides[field]
    return SHEET_DEFAULT_READ.get(sheet_name, "cs")


def make_sheet(wb: Workbook, name: str, columns: list[dict], rows: list[dict]) -> None:
    ws: Worksheet = wb.create_sheet(name)
    for col_idx, col in enumerate(columns, start=1):
        read_side = col.get("read", _resolve_read(name, col["field"]))
        if read_side:
            ws.cell(1, col_idx, read_side)
        ws.cell(2, col_idx, col["type"])
        ws.cell(3, col_idx, col["field"])
        ws.cell(4, col_idx, col.get("server", ""))
        ws.cell(5, col_idx, col.get("comment1", ""))
        ws.cell(6, col_idx, col.get("comment2", ""))
        ws.cell(7, col_idx, col.get("comment3", ""))
        ws.cell(8, col_idx, col.get("comment4", ""))
    for row_idx, row in enumerate(rows, start=9):
        for col_idx, col in enumerate(columns, start=1):
            val = row.get(col["field"])
            if val is None or (isinstance(val, str) and val.strip() == ""):
                if col["type"] == "ext[]":
                    val = col.get("default", "[]")
                elif col["type"] == "ext":
                    val = col.get("default", "{}")
            if val is not None:
                ws.cell(row_idx, col_idx, val)


def _field_index(ws: Worksheet) -> dict[str, int]:
    return {
        str(ws.cell(3, col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
        if ws.cell(3, col_idx).value is not None and str(ws.cell(3, col_idx).value).strip()
    }


def align_to_latest_activity_soccer_schema(wb: Workbook) -> Workbook:
    """Align generated workbook sheet/column headers to dataconfig ActivitySoccer.xlsx."""
    if not LATEST_ACTIVITY_SOCCER_XLSX.exists():
        return wb

    latest_wb = load_workbook(LATEST_ACTIVITY_SOCCER_XLSX, data_only=True)
    aligned = Workbook()
    aligned.remove(aligned.active)
    try:
        for sheet_name in latest_wb.sheetnames:
            latest_ws = latest_wb[sheet_name]
            old_ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            old_fields = _field_index(old_ws) if old_ws is not None else {}
            latest_fields = [
                latest_ws.cell(3, col_idx).value
                for col_idx in range(1, latest_ws.max_column + 1)
            ]
            old_rows_by_field: dict[int, dict[str, object]] = {}
            if old_ws is not None:
                for row_idx in range(9, old_ws.max_row + 1):
                    old_rows_by_field[row_idx] = {
                        field: old_ws.cell(row_idx, col_idx).value
                        for field, col_idx in old_fields.items()
                    }
            latest_rows = {
                row_idx: [
                    latest_ws.cell(row_idx, col_idx).value
                    for col_idx in range(1, latest_ws.max_column + 1)
                ]
                for row_idx in range(1, latest_ws.max_row + 1)
            }

            ws = aligned.create_sheet(sheet_name)
            for row_idx in range(1, min(latest_ws.max_row, 8) + 1):
                for col_idx, value in enumerate(latest_rows[row_idx], start=1):
                    if value is not None:
                        ws.cell(row_idx, col_idx, value)

            max_row = old_ws.max_row if old_ws is not None else latest_ws.max_row
            for row_idx in range(9, max_row + 1):
                latest_row = latest_rows.get(row_idx, [])
                old_row = old_rows_by_field.get(row_idx, {})
                for col_idx, field in enumerate(latest_fields, start=1):
                    value = None
                    if field and str(field) in old_row:
                        value = old_row[str(field)]
                    if value is None and col_idx <= len(latest_row):
                        value = latest_row[col_idx - 1]
                    if value is not None:
                        ws.cell(row_idx, col_idx, value)
        return aligned
    finally:
        latest_wb.close()


def copy_sheet_from_xlsx(wb: Workbook, name: str, source_path: Path, source_sheet: str = "Sheet1") -> None:
    """Copy a source worksheet's value matrix into the generated workbook."""
    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    try:
        source_ws = source_wb[source_sheet]
        ws = wb.create_sheet(name)
        for row_idx in range(1, source_ws.max_row + 1):
            for col_idx in range(1, source_ws.max_column + 1):
                ws.cell(row_idx, col_idx, source_ws.cell(row_idx, col_idx).value)
    finally:
        source_wb.close()


# dataconfig/ConstConfig.xlsx → ConstConfigCfg 表头（首列 CfgID，非活动表 ID 约定）
CONST_CONFIG_COLUMNS: list[dict] = [
    {"field": "CfgID", "type": "int", "server": "id", "read": "cs", "comment2": "编号"},
    {"field": "Comment", "type": "string", "server": "comment", "read": "s", "comment2": "备注"},
    {"field": "Constant", "type": "string", "server": "constant", "read": "cs", "comment2": "字符串"},
    {"field": "Val", "type": "double", "server": "val", "read": "cs", "comment2": "值"},
    {"field": "Array", "type": "int[]", "server": "array", "read": "cs", "comment2": "数组", "default": "[]"},
    {
        "field": "Effects", "type": "ext[]", "server": "quintuple", "read": "cs",
        "comment1": "Effect_P", "comment2": "", "default": "[]",
    },
    {"field": "", "type": "map", "server": "requirement", "read": "c", "comment2": "", "default": "{}"},
]


def make_const_format_sheet(wb: Workbook, sheet_name: str, rows: list[dict]) -> None:
    """Write a sheet whose columns match dataconfig/ConstConfig.xlsx → ConstConfigCfg."""
    ws: Worksheet = wb.create_sheet(sheet_name)
    cols = CONST_CONFIG_COLUMNS
    for col_idx, col in enumerate(cols, start=1):
        if col.get("read"):
            ws.cell(1, col_idx, col["read"])
        ws.cell(2, col_idx, col["type"])
        ws.cell(3, col_idx, col["field"] if col["field"] else "")
        ws.cell(4, col_idx, col.get("server", ""))
        ws.cell(5, col_idx, col.get("comment1", ""))
        ws.cell(6, col_idx, col.get("comment2", ""))
    req_field = cols[-1]["server"]
    for row_idx, row in enumerate(rows, start=9):
        for col_idx, col in enumerate(cols, start=1):
            key = col["field"] if col["field"] else req_field
            val = row.get(key)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                if col["type"] == "ext[]":
                    val = col.get("default", "[]")
                elif col["type"] == "map":
                    val = col.get("default", "{}")
            if val is not None:
                ws.cell(row_idx, col_idx, val)


def _const(
    offset: int,
    comment: str,
    constant: str,
    val: str | int | float = "0",
    array: str = "[]",
    effects: str = "[]",
    requirement: str = "{}",
) -> dict:
    return {
        "CfgID": ACTV_SOCCER_CONST_CFGID_BASE + offset,
        "Comment": comment,
        "Constant": constant,
        "Val": val,
        "Array": array,
        "Effects": effects,
        "requirement": requirement,
    }


def _first_sign_contract_rows() -> list[dict]:
    """首签 1 星合同：签约/存档关联 contract_id，展示球队通过 TeamID 反查。"""
    season_goal = '[{"type":"rank","threshold":12,"settle_at":"season_end"}]'
    season_reward = '[{"typ":"vm","id":11151001,"val":84}]'
    pairs = [
        (1, 101, "全局默认"),
        (11, 102, "欧洲默认"),
        (12, 103, "南美默认"),
        (13, 104, "南美默认2"),
        (101, 201, "中国动态池"),
        (102, 202, "中国动态池"),
        (103, 203, "欧洲动态池"),
        (104, 204, "欧洲/南美动态池"),
        (105, 205, "南美动态池"),
        (106, 206, "南美动态池"),
        (107, 207, "南美动态池"),
        (108, 208, "欧洲动态池"),
    ]
    return [
        {
            "ID": cid, "TeamID": tid, "TeamStar": 1,
            "PayFinish": 10, "PayGoal": 5, "PayAssist": 5, "PayFame": 20,
            "SeasonGoal": season_goal,
            "SeasonReward": season_reward,
            "SignReward": "[]",
            "GrantFameLevel": 1, "GrantLifeLevel": 1,
            "GrantScene": "first_sign",
            "Remark": f"首签1星-{note}",
        }
        for cid, tid, note in pairs
    ]


def _contract_star_license_rows() -> list[dict]:
    """合同星级许可表：ID=2~5，allowed_team_stars 与 star_weights 等长按下标对应。"""
    return [
        {"ID": 2, "AllowedTeamStars": "[1,2]", "StarWeights": "[10,30]", "Remark": "2星许可"},
        {"ID": 3, "AllowedTeamStars": "[1,2,3]", "StarWeights": "[10,30,60]", "Remark": "3星许可-文档示例"},
        {"ID": 4, "AllowedTeamStars": "[1,2,3,4]", "StarWeights": "[10,20,40,60]", "Remark": "4星许可"},
        {"ID": 5, "AllowedTeamStars": "[1,2,3,4,5]", "StarWeights": "[10,15,25,35,60]", "Remark": "5星许可"},
    ]


def _league_finish_contract_rows() -> list[dict]:
    """联赛完成换约池：星级由 ContractStarLicenseCfg 权重决定，同星级内均匀抽合同。"""
    season_goal_rank = '[{"type":"rank","threshold":12,"settle_at":"season_end"}]'
    season_goal_slice = '[{"type":"slice_win","threshold":5,"settle_at":"season_end"}]'
    reward_84 = '[{"typ":"vm","id":11151001,"val":84}]'
    reward_120 = '[{"typ":"vm","id":11151001,"val":120}]'
    reward_200 = '[{"typ":"vm","id":11151001,"val":200}]'
    reward_300 = '[{"typ":"vm","id":11151001,"val":300}]'
    reward_500 = '[{"typ":"vm","id":11151001,"val":500}]'
    rows = [
        (2, 204, 2, 20, 8, 6, 30, season_goal_slice, reward_120, 2, 2, "2星-联赛换约"),
        (3, 203, 2, 18, 7, 5, 28, season_goal_rank, reward_84, 1, 2, "2星-联赛换约(低门槛)"),
        (4, 205, 2, 22, 9, 7, 32, season_goal_slice, reward_120, 2, 3, "2星-联赛换约(高门槛)"),
        (5, 208, 3, 35, 12, 9, 45, season_goal_rank, reward_200, 3, 3, "3星-联赛换约"),
        (6, 207, 3, 32, 11, 8, 42, season_goal_slice, reward_200, 3, 2, "3星-联赛换约(生活门槛低)"),
        (7, 206, 1, 12, 4, 4, 18, season_goal_rank, reward_84, 1, 1, "1星-联赛换约(保底)"),
        (8, 201, 4, 45, 15, 10, 55, season_goal_rank, reward_300, 4, 4, "4星-联赛换约"),
        (9, 202, 5, 60, 20, 12, 70, season_goal_slice, reward_500, 5, 5, "5星-联赛换约"),
    ]
    return [
        {
            "ID": cid, "TeamID": tid, "TeamStar": star,
            "PayFinish": pf, "PayGoal": pg, "PayAssist": pa, "PayFame": pfm,
            "SeasonGoal": sg, "SeasonReward": sr, "SignReward": "[]",
            "GrantFameLevel": gfl, "GrantLifeLevel": gll,
            "GrantScene": "league_finish",
            "Remark": note,
        }
        for cid, tid, star, pf, pg, pa, pfm, sg, sr, gfl, gll, note in rows
    ]


def actv_soccer_const_rows() -> list[dict]:
    """ActivitySoccer 玩法常量行（活动表与 ConstConfig 补丁共用同一份数据）。"""
    return [
        _const(0, "主界面门票上限(测试)", "ActvSoccer_ticket_cap_default", "250"),
        _const(1, "门票恢复间隔(分钟)", "ActvSoccer_ticket_recover_minutes", "30"),
        _const(2, "淘汰赛开放关卡(round15起=level141)", "ActvSoccer_knockout_open_level", str(KNOCKOUT_OPEN_LEVEL)),
        _const(3, "可选角色数", "ActvSoccer_default_character_count", "4"),
        _const(4, "首签固定合同星级", "ActvSoccer_first_sign_contract_star", "1"),
        _const(
            5,
            "SeededRng种子组成(attempt_id+slice_id+ai_profile_id+rewind_count)",
            "ActvSoccer_ai_random_seed_formula",
            "0",
        ),
        _const(6, "死角不可扑(策划确认)", "ActvSoccer_dead_corner_can_save", "0"),
        _const(
            7,
            "参数叠加顺序 preset->instance->ai_profile->modifier",
            "ActvSoccer_config_overlay_order",
            "0",
        ),
        _const(8, "待机移动速度(固定值,单位m/s)", "ActvSoccer_move_speed_idle", "0"),
        _const(9, "慢走移动速度(固定值,单位m/s TODO)", "ActvSoccer_move_speed_walk", "2.5"),
        _const(10, "跑动速度下限;run=lerp(能力值,min,max)", "ActvSoccer_move_speed_run_min", "4"),
        _const(11, "跑动速度上限;run=lerp(能力值,min,max)", "ActvSoccer_move_speed_run_max", "8"),
        _const(12, "待机倍率;0=读MoveSpeedIdle,不乘run", "ActvSoccer_move_speed_ratio_idle", "0"),
        _const(13, "慢走倍率;0=读MoveSpeedWalk,不乘run", "ActvSoccer_move_speed_ratio_walk", "0"),
        _const(14, "正常跑倍率;0=读能力值映射run速度", "ActvSoccer_move_speed_ratio_run", "0"),
        _const(15, "慢跑跑位-队友/对手;相对run倍率", "ActvSoccer_move_speed_ratio_jog", "0.75"),
        _const(16, "冲刺-主角/对手;相对run倍率", "ActvSoccer_move_speed_ratio_sprint", "1.25"),
        _const(17, "带球推进-控球者;相对run倍率", "ActvSoccer_move_speed_ratio_dribble", "0.85"),
        _const(18, "逼抢-对手;相对run倍率", "ActvSoccer_move_speed_ratio_press", "1.1"),
        _const(19, "门将横移;相对run倍率,moving_keeper再乘params.speed", "ActvSoccer_move_speed_ratio_keeper_lateral", "0.9"),
        _const(20, "出球力量下限", "ActvSoccer_kick_force_min", "10"),
        _const(21, "出球力量上限", "ActvSoccer_kick_force_max", "25"),
        _const(22, "停球/控球距离(m)", "ActvSoccer_ball_control_distance", str(BALL_CONTROL_DISTANCE)),
        _const(23, "可操作夹角宽度下限(°)", "ActvSoccer_operable_angle_span_min", "20"),
        _const(24, "可操作夹角宽度上限(°)", "ActvSoccer_operable_angle_span_max", "70"),
        _const(25, "联赛换约候选合同份数", "ActvSoccer_contract_offer_count", "3"),
        # --- 竞猜赔率与经济（对齐程序 ConstCommon + controller；Val 为策划可读小数，Comment 标注程序键/万分值）---
        _const(26, "约束返奖率=ChampionMinOdds(程序12000)", "ActvSoccer_bet_constraint_return_rate", "1.2"),
        _const(27, "最大赔率=ChampionMaxOdds(程序500000)", "ActvSoccer_bet_max_odds", "50"),
        _const(28, "默认/初始展示倍率=ChampionDefaultOdds(程序17500)", "ActvSoccer_bet_initial_display_odds", "1.75"),
        _const(29, "赔率下限(程序ChampionMinOdds同值1.2)", "ActvSoccer_bet_min_odds", "1.2"),
        _const(30, "战力比上限=ChampionPowerRatioMax(程序20000=200%)", "ActvSoccer_bet_power_ratio_max", "2"),
        _const(31, "战力比下限=ChampionPowerRatioMin(程序3000=30%)", "ActvSoccer_bet_power_ratio_min", "0.3"),
        _const(32, "单场演算批次=oneMatchPreFightTimes", "ActvSoccer_bet_sim_count", "10"),
        _const(33, "每tick最多开战场次=maxPreFightBattlesPerTick", "ActvSoccer_bet_max_prefight_battles_per_tick", "50"),
        _const(34, "演算批次间隔秒=roundFightInterval", "ActvSoccer_bet_round_fight_interval_sec", "1"),
        _const(35, "单场演算超时秒=preFightTimeout", "ActvSoccer_bet_prefight_timeout_sec", "30"),
        _const(36, "同步匹配阶段时长秒(活动排期,近preFightTimeout)", "ActvSoccer_bet_sync_phase_duration_sec", "30"),
        _const(37, "单次购买竞猜币上限(TODO)", "ActvSoccer_bet_coin_purchase_limit", "5000"),
        _const(38, "免费档命中固定派彩", "ActvSoccer_bet_free_hit_payout", "5"),
        _const(39, "每日免费领取竞猜币(TODO)", "ActvSoccer_bet_daily_free_amount", "100"),
        _const(40, "截止前锁定最终倍率(秒)", "ActvSoccer_bet_odds_lock_before_settle_sec", "59"),
        _const(41, "展示倍率逐分钟收敛间隔(秒)", "ActvSoccer_bet_display_odds_tick_sec", "60"),
    ]


# X1 奖励倍率对照表（A 胜率% → A/B 奖励倍率）；查表取 WinRatePctA 最接近行
# 二维赔率表(=程序ChampionOddsCfg):winRate × powerRate → oddsLeft/oddsRight 万分值
# winRate_int: 0/1000/.../10000 共11档(单位 /10000)
# powerRate_int: 3000..20000 步长 500 共35档(战力比 /10000)
# 11×35=385 行,行序按 (powerRate, winRate) 升序填充
BET_MULTIPLIER_GRID: list[tuple[int, int, int, int]] = [
    (0, 3000, 500000, 12000),
    (1000, 3000, 500000, 12000),
    (2000, 3000, 500000, 12000),
    (3000, 3000, 500000, 12000),
    (4000, 3000, 500000, 12000),
    (5000, 3000, 500000, 12000),
    (6000, 3000, 475000, 12000),
    (7000, 3000, 451300, 12000),
    (8000, 3000, 428700, 12000),
    (9000, 3000, 407300, 12000),
    (10000, 3000, 386900, 12000),
    (0, 3500, 500000, 12000),
    (1000, 3500, 486300, 12000),
    (2000, 3500, 463100, 12000),
    (3000, 3500, 441000, 12000),
    (4000, 3500, 420000, 12000),
    (5000, 3500, 400000, 12000),
    (6000, 3500, 380000, 12000),
    (7000, 3500, 361000, 12000),
    (8000, 3500, 343000, 12000),
    (9000, 3500, 325900, 12000),
    (10000, 3500, 309600, 12000),
    (0, 4000, 382900, 12000),
    (1000, 4000, 364700, 12000),
    (2000, 4000, 347300, 12000),
    (3000, 4000, 330800, 12000),
    (4000, 4000, 315000, 12000),
    (5000, 4000, 300000, 12000),
    (6000, 4000, 285000, 12000),
    (7000, 4000, 270800, 12000),
    (8000, 4000, 257300, 12000),
    (9000, 4000, 244400, 12000),
    (10000, 4000, 232200, 12000),
    (0, 4500, 255300, 12000),
    (1000, 4500, 243100, 12000),
    (2000, 4500, 231500, 12000),
    (3000, 4500, 220500, 12000),
    (4000, 4500, 210000, 12000),
    (5000, 4500, 200000, 12000),
    (6000, 4500, 190000, 12000),
    (7000, 4500, 180500, 12000),
    (8000, 4500, 171500, 12000),
    (9000, 4500, 162900, 12000),
    (10000, 4500, 154800, 12000),
    (0, 5000, 191500, 12000),
    (1000, 5000, 182400, 12000),
    (2000, 5000, 173700, 12000),
    (3000, 5000, 165400, 12000),
    (4000, 5000, 157500, 12000),
    (5000, 5000, 150000, 12000),
    (6000, 5000, 142500, 12000),
    (7000, 5000, 135400, 12000),
    (8000, 5000, 128600, 12000),
    (9000, 5000, 122200, 12000),
    (10000, 5000, 116100, 12000),
    (0, 5500, 153100, 12000),
    (1000, 5500, 145800, 12000),
    (2000, 5500, 138900, 12000),
    (3000, 5500, 132300, 12000),
    (4000, 5500, 126000, 12000),
    (5000, 5500, 120000, 12000),
    (6000, 5500, 114000, 12000),
    (7000, 5500, 108300, 12000),
    (8000, 5500, 102900, 12000),
    (9000, 5500, 97800, 12000),
    (10000, 5500, 92900, 12000),
    (0, 6000, 102100, 12000),
    (1000, 6000, 97200, 12000),
    (2000, 6000, 92600, 12000),
    (3000, 6000, 88200, 12000),
    (4000, 6000, 84000, 12000),
    (5000, 6000, 80000, 12000),
    (6000, 6000, 76000, 12000),
    (7000, 6000, 72200, 12000),
    (8000, 6000, 68600, 12000),
    (9000, 6000, 65200, 12000),
    (10000, 6000, 61900, 12000),
    (0, 6500, 76700, 12000),
    (1000, 6500, 73000, 12000),
    (2000, 6500, 69500, 12000),
    (3000, 6500, 66200, 12000),
    (4000, 6500, 63000, 12000),
    (5000, 6500, 60000, 12000),
    (6000, 6500, 57000, 12000),
    (7000, 6500, 54200, 12000),
    (8000, 6500, 51500, 12000),
    (9000, 6500, 48900, 12000),
    (10000, 6500, 46500, 13200),
    (0, 7000, 63800, 12000),
    (1000, 7000, 60800, 12000),
    (2000, 7000, 57900, 12000),
    (3000, 7000, 55100, 12000),
    (4000, 7000, 52500, 12000),
    (5000, 7000, 50000, 12000),
    (6000, 7000, 47500, 13200),
    (7000, 7000, 45100, 13200),
    (8000, 7000, 42800, 14400),
    (9000, 7000, 40700, 14400),
    (10000, 7000, 38700, 15600),
    (0, 7500, 41000, 14400),
    (1000, 7500, 39000, 15600),
    (2000, 7500, 37100, 15600),
    (3000, 7500, 35300, 16800),
    (4000, 7500, 33600, 18000),
    (5000, 7500, 32000, 18000),
    (6000, 7500, 30400, 19200),
    (7000, 7500, 28900, 20400),
    (8000, 7500, 27500, 21600),
    (9000, 7500, 26100, 22800),
    (10000, 7500, 24800, 24000),
    (0, 8000, 38300, 15600),
    (1000, 8000, 36500, 16800),
    (2000, 8000, 34800, 16800),
    (3000, 8000, 33100, 18000),
    (4000, 8000, 31500, 19200),
    (5000, 8000, 30000, 19200),
    (6000, 8000, 28500, 20400),
    (7000, 8000, 27100, 21600),
    (8000, 8000, 25700, 22800),
    (9000, 8000, 24400, 24000),
    (10000, 8000, 23200, 25200),
    (0, 8500, 36000, 16800),
    (1000, 8500, 34300, 16800),
    (2000, 8500, 32700, 18000),
    (3000, 8500, 31100, 19200),
    (4000, 8500, 29600, 20400),
    (5000, 8500, 28200, 20400),
    (6000, 8500, 26800, 21600),
    (7000, 8500, 25500, 22800),
    (8000, 8500, 24200, 24000),
    (9000, 8500, 23000, 25200),
    (10000, 8500, 21900, 26400),
    (0, 9000, 34000, 18000),
    (1000, 9000, 32400, 18000),
    (2000, 9000, 30900, 19200),
    (3000, 9000, 29400, 20400),
    (4000, 9000, 28000, 21600),
    (5000, 9000, 26700, 21600),
    (6000, 9000, 25400, 22800),
    (7000, 9000, 24100, 24000),
    (8000, 9000, 22900, 25200),
    (9000, 9000, 21800, 26400),
    (10000, 9000, 20700, 28800),
    (0, 9500, 32300, 18000),
    (1000, 9500, 30800, 19200),
    (2000, 9500, 29300, 20400),
    (3000, 9500, 27900, 21600),
    (4000, 9500, 26600, 22800),
    (5000, 9500, 25300, 22800),
    (6000, 9500, 24000, 24000),
    (7000, 9500, 22800, 26400),
    (8000, 9500, 21700, 27600),
    (9000, 9500, 20600, 28800),
    (10000, 9500, 19600, 30000),
    (0, 10000, 30700, 19200),
    (1000, 10000, 29200, 20400),
    (2000, 10000, 27800, 21600),
    (3000, 10000, 26500, 22800),
    (4000, 10000, 25200, 24000),
    (5000, 10000, 24000, 24000),
    (6000, 10000, 22800, 26400),
    (7000, 10000, 21700, 27600),
    (8000, 10000, 20600, 28800),
    (9000, 10000, 19600, 30000),
    (10000, 10000, 18600, 31200),
    (0, 10500, 29200, 20400),
    (1000, 10500, 27800, 21600),
    (2000, 10500, 26500, 22800),
    (3000, 10500, 25200, 24000),
    (4000, 10500, 24000, 24000),
    (5000, 10500, 22900, 25200),
    (6000, 10500, 21800, 26400),
    (7000, 10500, 20700, 28800),
    (8000, 10500, 19700, 30000),
    (9000, 10500, 18700, 31200),
    (10000, 10500, 17800, 32400),
    (0, 11000, 27800, 21600),
    (1000, 11000, 26500, 22800),
    (2000, 11000, 25200, 24000),
    (3000, 11000, 24000, 24000),
    (4000, 11000, 22900, 25200),
    (5000, 11000, 21800, 26400),
    (6000, 11000, 20700, 28800),
    (7000, 11000, 19700, 30000),
    (8000, 11000, 18700, 31200),
    (9000, 11000, 17800, 32400),
    (10000, 11000, 16900, 34800),
    (0, 11500, 26700, 21600),
    (1000, 11500, 25400, 22800),
    (2000, 11500, 24200, 24000),
    (3000, 11500, 23000, 25200),
    (4000, 11500, 21900, 26400),
    (5000, 11500, 20900, 27600),
    (6000, 11500, 19900, 30000),
    (7000, 11500, 18900, 31200),
    (8000, 11500, 18000, 32400),
    (9000, 11500, 17100, 33600),
    (10000, 11500, 16200, 36000),
    (0, 12000, 25600, 22800),
    (1000, 12000, 24400, 24000),
    (2000, 12000, 23200, 25200),
    (3000, 12000, 22100, 26400),
    (4000, 12000, 21000, 27600),
    (5000, 12000, 20000, 28800),
    (6000, 12000, 19000, 31200),
    (7000, 12000, 18100, 32400),
    (8000, 12000, 17200, 33600),
    (9000, 12000, 16300, 36000),
    (10000, 12000, 15500, 37200),
    (0, 12500, 24600, 24000),
    (1000, 12500, 23400, 25200),
    (2000, 12500, 22300, 26400),
    (3000, 12500, 21200, 27600),
    (4000, 12500, 20200, 28800),
    (5000, 12500, 19200, 30000),
    (6000, 12500, 18200, 32400),
    (7000, 12500, 17300, 33600),
    (8000, 12500, 16400, 36000),
    (9000, 12500, 15600, 37200),
    (10000, 12500, 14800, 39600),
    (0, 13000, 23600, 25200),
    (1000, 13000, 22500, 26400),
    (2000, 13000, 21400, 27600),
    (3000, 13000, 20400, 28800),
    (4000, 13000, 19400, 30000),
    (5000, 13000, 18500, 31200),
    (6000, 13000, 17600, 33600),
    (7000, 13000, 16700, 34800),
    (8000, 13000, 15900, 37200),
    (9000, 13000, 15100, 38400),
    (10000, 13000, 14300, 40800),
    (0, 13500, 22700, 26400),
    (1000, 13500, 21600, 27600),
    (2000, 13500, 20600, 28800),
    (3000, 13500, 19600, 30000),
    (4000, 13500, 18700, 31200),
    (5000, 13500, 17800, 32400),
    (6000, 13500, 16900, 34800),
    (7000, 13500, 16100, 36000),
    (8000, 13500, 15300, 38400),
    (9000, 13500, 14500, 39600),
    (10000, 13500, 13800, 42000),
    (0, 14000, 21800, 26400),
    (1000, 14000, 20800, 28800),
    (2000, 14000, 19800, 30000),
    (3000, 14000, 18900, 31200),
    (4000, 14000, 18000, 32400),
    (5000, 14000, 17100, 33600),
    (6000, 14000, 16200, 36000),
    (7000, 14000, 15400, 38400),
    (8000, 14000, 14600, 39600),
    (9000, 14000, 13900, 42000),
    (10000, 14000, 13200, 44400),
    (0, 14500, 21200, 27600),
    (1000, 14500, 20200, 28800),
    (2000, 14500, 19200, 30000),
    (3000, 14500, 18300, 32400),
    (4000, 14500, 17400, 33600),
    (5000, 14500, 16600, 34800),
    (6000, 14500, 15800, 37200),
    (7000, 14500, 15000, 38400),
    (8000, 14500, 14300, 40800),
    (9000, 14500, 13600, 43200),
    (10000, 14500, 12900, 45600),
    (0, 15000, 20400, 28800),
    (1000, 15000, 19400, 30000),
    (2000, 15000, 18500, 31200),
    (3000, 15000, 17600, 33600),
    (4000, 15000, 16800, 34800),
    (5000, 15000, 16000, 36000),
    (6000, 15000, 15200, 38400),
    (7000, 15000, 14400, 40800),
    (8000, 15000, 13700, 42000),
    (9000, 15000, 13000, 44400),
    (10000, 15000, 12400, 46800),
    (0, 15500, 19800, 30000),
    (1000, 15500, 18900, 31200),
    (2000, 15500, 18000, 32400),
    (3000, 15500, 17100, 33600),
    (4000, 15500, 16300, 36000),
    (5000, 15500, 15500, 37200),
    (6000, 15500, 14700, 39600),
    (7000, 15500, 14000, 42000),
    (8000, 15500, 13300, 43200),
    (9000, 15500, 12600, 45600),
    (10000, 15500, 12000, 48000),
    (0, 16000, 19200, 30000),
    (1000, 16000, 18300, 32400),
    (2000, 16000, 17400, 33600),
    (3000, 16000, 16600, 34800),
    (4000, 16000, 15800, 37200),
    (5000, 16000, 15000, 38400),
    (6000, 16000, 14300, 40800),
    (7000, 16000, 13600, 43200),
    (8000, 16000, 12900, 45600),
    (9000, 16000, 12300, 46800),
    (10000, 16000, 12000, 49200),
    (0, 16500, 18500, 31200),
    (1000, 16500, 17600, 33600),
    (2000, 16500, 16800, 34800),
    (3000, 16500, 16000, 36000),
    (4000, 16500, 15200, 38400),
    (5000, 16500, 14500, 39600),
    (6000, 16500, 13800, 42000),
    (7000, 16500, 13100, 44400),
    (8000, 16500, 12400, 46800),
    (9000, 16500, 12000, 49200),
    (10000, 16500, 12000, 51600),
    (0, 17000, 18000, 32400),
    (1000, 17000, 17100, 33600),
    (2000, 17000, 16300, 36000),
    (3000, 17000, 15500, 37200),
    (4000, 17000, 14800, 39600),
    (5000, 17000, 14100, 40800),
    (6000, 17000, 13400, 43200),
    (7000, 17000, 12700, 45600),
    (8000, 17000, 12100, 48000),
    (9000, 17000, 12000, 50400),
    (10000, 17000, 12000, 52800),
    (0, 17500, 17500, 33600),
    (1000, 17500, 16700, 34800),
    (2000, 17500, 15900, 37200),
    (3000, 17500, 15100, 38400),
    (4000, 17500, 14400, 40800),
    (5000, 17500, 13700, 42000),
    (6000, 17500, 13000, 44400),
    (7000, 17500, 12400, 46800),
    (8000, 17500, 12000, 49200),
    (9000, 17500, 12000, 51600),
    (10000, 17500, 12000, 55200),
    (0, 18000, 17000, 34800),
    (1000, 18000, 16200, 36000),
    (2000, 18000, 15400, 38400),
    (3000, 18000, 14700, 39600),
    (4000, 18000, 14000, 42000),
    (5000, 18000, 13300, 43200),
    (6000, 18000, 12600, 45600),
    (7000, 18000, 12000, 48000),
    (8000, 18000, 12000, 50400),
    (9000, 18000, 12000, 54000),
    (10000, 18000, 12000, 56400),
    (0, 18500, 16700, 34800),
    (1000, 18500, 15900, 37200),
    (2000, 18500, 15100, 38400),
    (3000, 18500, 14400, 40800),
    (4000, 18500, 13700, 42000),
    (5000, 18500, 13000, 44400),
    (6000, 18500, 12400, 46800),
    (7000, 18500, 12000, 49200),
    (8000, 18500, 12000, 51600),
    (9000, 18500, 12000, 55200),
    (10000, 18500, 12000, 57600),
    (0, 19000, 16100, 36000),
    (1000, 19000, 15300, 38400),
    (2000, 19000, 14600, 39600),
    (3000, 19000, 13900, 42000),
    (4000, 19000, 13200, 44400),
    (5000, 19000, 12600, 45600),
    (6000, 19000, 12000, 48000),
    (7000, 19000, 12000, 50400),
    (8000, 19000, 12000, 54000),
    (9000, 19000, 12000, 56400),
    (10000, 19000, 12000, 58800),
    (0, 19500, 15600, 37200),
    (1000, 19500, 14900, 39600),
    (2000, 19500, 14200, 40800),
    (3000, 19500, 13500, 43200),
    (4000, 19500, 12900, 45600),
    (5000, 19500, 12300, 46800),
    (6000, 19500, 12000, 49200),
    (7000, 19500, 12000, 52800),
    (8000, 19500, 12000, 55200),
    (9000, 19500, 12000, 57600),
    (10000, 19500, 12000, 61200),
    (0, 20000, 15300, 38400),
    (1000, 20000, 14600, 39600),
    (2000, 20000, 13900, 42000),
    (3000, 20000, 13200, 44400),
    (4000, 20000, 12600, 45600),
    (5000, 20000, 12000, 48000),
    (6000, 20000, 12000, 50400),
    (7000, 20000, 12000, 54000),
    (8000, 20000, 12000, 56400),
    (9000, 20000, 12000, 58800),
    (10000, 20000, 12000, 62400),
]

BET_STAKE_TIER_ROWS: list[dict] = [
    {"ID": 1, "TierKey": "free", "Stake": 0, "HitPayout": 5, "Sort": 1, "Remark": "免费档;命中固定5竞猜币"},
    {"ID": 2, "TierKey": "stake_50", "Stake": 50, "HitPayout": 0, "Sort": 2, "Remark": "命中=stake×locked_mult"},
    {"ID": 3, "TierKey": "stake_100", "Stake": 100, "HitPayout": 0, "Sort": 3, "Remark": ""},
    {"ID": 4, "TierKey": "stake_150", "Stake": 150, "HitPayout": 0, "Sort": 4, "Remark": ""},
    {"ID": 5, "TierKey": "stake_200", "Stake": 200, "HitPayout": 0, "Sort": 5, "Remark": ""},
    {"ID": 6, "TierKey": "stake_300", "Stake": 300, "HitPayout": 0, "Sort": 6, "Remark": "最高档"},
]


def build_const_config_workbook(rows: list[dict]) -> Workbook:
    """ConstConfig 补丁 xlsx，便于合并进 dataconfig/ConstConfig.xlsx。"""
    wb = Workbook()
    wb.remove(wb.active)
    make_const_format_sheet(wb, "ConstConfigCfg", rows)
    return wb


def id_col(col_type: str = "int", desc: str = "编号") -> dict:
    """首列固定 ID / id（K1 测试配置约定）。"""
    return {"field": "ID", "type": col_type, "server": "id", "comment2": desc}


def c(*specs: tuple | dict) -> list[dict]:
    """Column spec: (field, type, desc) or (field, type, desc, proto_row5) or + json_example row7; or id_col dict."""
    cols: list[dict] = []
    for spec in specs:
        if isinstance(spec, dict):
            cols.append(spec)
            continue
        field, col_type, desc = spec[0], spec[1], spec[2]
        col: dict = {"field": field, "type": col_type, "comment2": desc}
        if len(spec) > 3 and spec[3]:
            col["comment1"] = spec[3]
        if len(spec) > 4 and spec[4]:
            col["comment3"] = spec[4]
            if col_type in ("ext", "ext[]"):
                col["default"] = spec[4]
        cols.append(col)
    return cols


# K1 ext 第5行结构标记（见 config-table-editor skill）
P_TYIDVAL = "TypIDVal_P_cspb"
P_VEC3 = "PositionTuple_P"
V_TYPE_PAYLOAD = "SoccerTypePayload_V"
V_MODIFIER = "SoccerModifier_V"
V_PLAYER_INIT = "SoccerPlayerInit_V"
V_SEASON_GOAL = "SoccerSeasonGoal_V"

# SoccerPlayerInit_V: team, idx, duty(→PlayerAiDuty), pos, facing(°)

# ============================================================
# 坐标系协议 v1 (2026-06-18 确认)
# 真相源:references/soccer-coordinate-protocol.md
# 修改本段必须同步更新协议;两侧由 scripts/check_protocol_drift.py lint。
# ============================================================

# §1 基础:m / 右手系 / +z 进攻 / facing ∈ [-180, 180]
COORD_UNIT = "m"

# §2 场地边界
FIELD_X_HALF = 18.0          # x ∈ [-18, 18]
FIELD_Z_NEAR = 0.0           # 对方球门线
FIELD_Z_MID = -30.0          # 中线
FIELD_Z_FAR = -60.0           # 本方球门线

# §3.1 球门
AWAY_GOAL_CENTER = (0.0, 0.0, 0.0)
HOME_GOAL_CENTER = (0.0, 0.0, -60.0)
GOAL_WIDTH = 8.5
GOAL_HEIGHT = 3.0
DEAD_CORNER_THICKNESS = 0.2

# 对方门将站位 z（贴近球门线；游戏比例下门将不大幅出迎）
AWAY_KEEPER_Z_DEFAULT = -0.5
AWAY_KEEPER_Z_PENALTY = 0.0
AWAY_KEEPER_Z_LONG_ATTACK = -1.0
LONG_ATTACK_BALL_Z = -16.0   # 球点 z ≤ 此值视为远距进攻(吊射/远射,门将略向前出)


def away_keeper_z(slice_type: str, ball_z: float | None = None) -> float:
    if slice_type == "penalty":
        return AWAY_KEEPER_Z_PENALTY
    if slice_type == "attack" and ball_z is not None and ball_z <= LONG_ATTACK_BALL_Z:
        return AWAY_KEEPER_Z_LONG_ATTACK
    return AWAY_KEEPER_Z_DEFAULT

# §3.2 大禁区
PENALTY_AREA_Z_FAR = -10.0
PENALTY_AREA_X_HALF = 11.5

# §3.3 小禁区
GOAL_AREA_Z_FAR = -3.5

# §3.4 点球
PENALTY_SPOT = (0.0, 0.0, -6.5)
PENALTY_FREE_RADIUS = 5.0    # 弧外球员距球 ≥ 5m

# §3.5 角球
CORNER_LEFT_BALL = (-17.0, 0.0, -1.0)
CORNER_RIGHT_BALL = (17.0, 0.0, -1.0)
CORNER_FLAG_X = 18.0
CORNER_FREE_RADIUS = 5.0

# §3.6 中圈
CENTER_CIRCLE_CENTER = (0.0, 0.0, -30.0)
CENTER_CIRCLE_RADIUS = 4.5

# §4 物理
PLAYER_RADIUS = 0.5
BALL_RADIUS = 0.2
BALL_CONTROL_DISTANCE = 0.5
WALL_PLAYER_GAP_MIN = 2 * PLAYER_RADIUS + 2 * BALL_RADIUS  # 1.4m,人墙间距下限

# §5 兼容:旧代码用 GOAL_CENTER_X/Z,保留指向 AWAY_GOAL_CENTER
GOAL_CENTER_X, GOAL_CENTER_Z = AWAY_GOAL_CENTER[0], AWAY_GOAL_CENTER[2]


def away_goal_target(y: float = 0.0) -> str:
    """对方球门 + 指定 y(高度)的 TargetPoint JSON。"""
    return json.dumps({"x": 0.0, "y": y, "z": AWAY_GOAL_CENTER[2]})


def penalty_ball_pos() -> str:
    """点球 BallPos JSON(协议 §3.4)。"""
    return json.dumps({"x": PENALTY_SPOT[0], "y": 0.0, "z": PENALTY_SPOT[2]})


def corner_ball_pos(side: str) -> str:
    """角球 BallPos JSON,side='left'/'right'(协议 §3.5)。"""
    pos = CORNER_LEFT_BALL if side == "left" else CORNER_RIGHT_BALL
    return json.dumps({"x": pos[0], "y": 0.0, "z": pos[2]})

# === 协议常量段结束 ===
PLAYER_INIT_DEFAULT = (
    '[{"team":"home","idx":0,"duty":3,'
    '"pos":{"x":0,"y":0,"z":0},"facing":0}]'
)


def _yaw_deg(dx: float, dz: float) -> float:
    return round(math.degrees(math.atan2(dx, dz)), 1)


def _face_toward(fx: float, fz: float, x: float, z: float) -> float:
    return _yaw_deg(fx - x, fz - z)


def player_init(
    team: str, idx: int, duty: int, x: float, y: float, z: float, facing: float
) -> dict:
    return {
        "team": team,
        "idx": idx,
        "duty": duty,
        "pos": {"x": x, "y": y, "z": z},
        "facing": facing,
    }


def players_init_json(players: list[dict]) -> str:
    return json.dumps(players, ensure_ascii=False)


def pos_json(x: float, y: float, z: float) -> str:
    return json.dumps({"x": x, "y": y, "z": z})


def _forward_from_facing(facing: float) -> tuple[float, float]:
    yaw = math.radians(facing)
    return math.sin(yaw), math.cos(yaw)


def _round_coord(value: float) -> float:
    return round(value, 1)


LC_PREFIX = "ActvSoccer"


def lc_key(*parts: str) -> str:
    """生成本地化唯一 ID，格式: ActvSoccer_{语义}_{序号}。"""
    return f"{LC_PREFIX}_" + "_".join(parts)


class LcRegistry:
    """收集测试配置引用的本地化条目，ID 为英文字符串且全局唯一。"""

    def __init__(self) -> None:
        self._rows: list[dict] = []
        self._keys: dict[str, str] = {}

    def add(self, lc_id: str, cn: str, source: str = "") -> str:
        if lc_id in self._keys:
            return self._keys[lc_id]
        self._keys[lc_id] = lc_id
        self._rows.append({"ID": lc_id, "Cn": cn, "Source": source})
        return lc_id

    @property
    def rows(self) -> list[dict]:
        return self._rows


def build_language_workbook(registry: LcRegistry) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    make_sheet(
        wb,
        "ActvSoccerLanguageCfg",
        c(
            id_col("string", "本地化唯一ID(英文+数字)"),
            ("Cn", "string", "简体中文"),
            ("Source", "string", "引用来源(测试追溯)"),
        ),
        registry.rows,
    )
    return wb


def build_guide_step_rows(lc: LcRegistry) -> list[dict]:
    """合并试训、签约后主界面强引导、第一关引导关的步骤配置。"""

    def text(step_id: int, kind: str, cn: str) -> str:
        return lc.add(lc_key("guide", kind, str(step_id)), cn, f"GuideStepCfg/{step_id}")

    def row(
        step_id: int,
        guide_id: str,
        step_index: int,
        trigger: str,
        level_id: int,
        slice_id: int,
        dialogue: str,
        text_style: str,
        focus_target: str,
        mask_type: str,
        gesture: str,
        wait_type: str,
        wait_target: str,
        pass_condition: str,
        fail_action: str,
        save_scope: str,
        save_key: str,
        save_value: str,
        save_timing: str,
        next_id: int,
        skip_policy: str,
        remark: str = "",
    ) -> dict:
        return {
            "ID": step_id,
            "GuideID": guide_id,
            "StepIndex": step_index,
            "Trigger": trigger,
            "LevelID": level_id,
            "SliceID": slice_id,
            "DialogueLcKey": text(step_id, "dialogue", dialogue) if dialogue else "",
            "TextStyle": text_style,
            "FocusTarget": focus_target,
            "MaskType": mask_type,
            "GestureDesc": gesture,
            "WaitType": wait_type,
            "WaitTarget": wait_target,
            "PassCondition": pass_condition,
            "FailAction": fail_action,
            "SaveScope": save_scope,
            "SaveKey": save_key,
            "SaveValue": save_value,
            "SaveTiming": save_timing,
            "NextID": next_id,
            "SkipPolicy": skip_policy,
            "Remark": remark,
        }

    source_rows = [
        row(1001, "trial", 1, "on_confirm_character_nationality", 0, 0, "先完成一次试训，熟悉比赛中的射门操作。", "bottom_bar", "screen_center", "none", "", "auto_delay", "1.0s", "delay_end", "none", "profile", "trial_state", "started", "step_complete", 1010, "cannot_skip", "进入试训前说明"),
        row(1010, "trial", 2, "on_trial_slice_enter", 0, 101, "按住球员脚下区域，画出你想踢出的路线。", "bottom_bar", "player_operate_area", "spotlight", "白色手指从球员脚下按住，沿推荐轨迹拖向白色目标框，轨迹线循环播放。", "drag", "player_operate_area", "player_drag_start", "retry_step", "profile", "trial_step", "1", "step_complete", 1011, "cannot_skip", "试训1:进攻划线"),
        row(1011, "trial", 3, "on_player_drag_start", 0, 101, "松手后，球会沿着轨迹飞出。", "bottom_bar", "predicted_trajectory", "spotlight", "手指沿轨迹到达目标框后松开，展示释放射门。", "release", "target_goal_frame", "player_release_ball", "retry_step", "", "", "", "", 1012, "cannot_skip", "试训1:松手射门"),
        row(1012, "trial", 4, "on_slice_101_success", 0, 101, "做得好！继续下一项训练。", "toast", "target_goal_frame", "none", "", "auto_delay", "0.8s", "delay_end", "none", "profile", "trial_step_1_done", "true", "step_complete", 1020, "cannot_skip", "试训1成功"),
        row(1013, "trial", 5, "on_slice_101_fail", 0, 101, "再试一次，把路线画向白色目标框。", "bottom_bar", "target_goal_frame", "spotlight", "重复播放从球员脚下划向目标框的轨迹手势。", "drag", "player_operate_area", "player_release_ball", "retry_step", "", "", "", "", 1011, "cannot_skip", "试训1失败重试"),
        row(1020, "trial", 6, "on_trial_slice_enter", 0, 102, "任意球需要避开人墙。拖动来选择方向和力量。", "bottom_bar", "free_kick_target_area", "spotlight", "手指从球向后拖拽，出现弹弓方向线和力度百分比。", "drag", "power_angle_area", "player_drag_power", "retry_step", "profile", "trial_step", "2", "step_complete", 1021, "cannot_skip", "试训2:任意球方向和力度"),
        row(1021, "trial", 7, "on_player_drag_power", 0, 102, "力量控制在65%左右，更容易命中目标。", "bottom_bar", "power_percent_ui", "spotlight", "手指拖动到推荐力度点，力度数值从49%动到65%附近。", "release", "power_percent_ui", "power_in_range_or_release", "retry_step", "", "", "", "", 1022, "cannot_skip", "试训2:力度教学"),
        row(1022, "trial", 8, "on_slice_102_success", 0, 102, "很好，任意球训练完成。", "toast", "goal_target", "none", "", "auto_delay", "0.8s", "delay_end", "none", "profile", "trial_step_2_done", "true", "step_complete", 1030, "cannot_skip", "试训2成功"),
        row(1023, "trial", 9, "on_slice_102_fail_power_low", 0, 102, "力量太小了，再向后拖动一些。", "bottom_bar", "power_percent_ui", "spotlight", "重复播放向后拖拽至65%附近的力度手势。", "drag", "power_angle_area", "player_release_ball", "retry_step", "", "", "", "", 1021, "cannot_skip", "任意球力量太小"),
        row(1024, "trial", 10, "on_slice_102_fail_power_high", 0, 102, "力量太大了，稍微收一点力。", "bottom_bar", "power_percent_ui", "spotlight", "重复播放力度回收到65%附近的手势。", "drag", "power_angle_area", "player_release_ball", "retry_step", "", "", "", "", 1021, "cannot_skip", "任意球力量太大"),
        row(1025, "trial", 11, "on_slice_102_fail_angle", 0, 102, "角度偏离目标，再调整射门方向。", "bottom_bar", "angle_sector", "spotlight", "重复播放拖动方向线进入黄色允许角度范围的手势。", "drag", "power_angle_area", "player_release_ball", "retry_step", "", "", "", "", 1021, "cannot_skip", "任意球角度偏离"),
        row(1030, "trial", 12, "on_trial_slice_enter", 0, 103, "试训中可以自由切换操作方式，选择你更顺手的模式。", "bottom_bar", "mode_toggle_button", "spotlight", "点击手指点按模式切换按钮，按钮在划线/弹弓之间切换。", "tap", "mode_toggle_button", "mode_toggle_clicked_or_delay", "none", "profile", "trial_step", "3", "step_complete", 1031, "auto_after_delay", "试训3:模式切换"),
        row(1031, "trial", 13, "on_mode_toggle_done", 0, 103, "现在把球踢向高亮区域。", "bottom_bar", "penalty_goal_target", "spotlight", "按玩家当前模式播放对应手势：划线轨迹或弹弓拖拽。", "release", "penalty_goal_target", "player_release_ball", "retry_step", "", "", "", "", 1032, "cannot_skip", "试训3:点球射门"),
        row(1032, "trial", 14, "on_slice_103_success", 0, 103, "点球训练完成。", "toast", "penalty_goal_target", "none", "", "auto_delay", "0.8s", "delay_end", "none", "profile", "trial_step_3_done", "true", "step_complete", 1040, "cannot_skip", "试训3成功"),
        row(1033, "trial", 15, "on_slice_103_fail", 0, 103, "再试一次，瞄准高亮区域射门。", "bottom_bar", "penalty_goal_target", "spotlight", "按当前操作模式重复播放点球射门手势。", "release", "penalty_goal_target", "player_release_ball", "retry_step", "", "", "", "", 1031, "cannot_skip", "点球失败重试"),
        row(1040, "trial", 16, "on_trial_all_slices_done", 0, 0, "试训完成。球队给你发来了邀请，挑选一支作为职业生涯起点。", "popup", "trial_complete_confirm", "full", "点击手指循环点按确认按钮。", "tap", "trial_complete_confirm", "button_clicked", "none", "profile", "trial_done", "true", "button_clicked", 0, "cannot_skip", "试训完成进入选队"),
        row(2001, "main_to_level1", 1, "on_first_contract_signed", 0, 0, "第一场比赛已经准备好了，先从基础比赛开始。", "bubble", "main_next_match_button", "spotlight", "点击手指点按主界面下一场比赛按钮。", "tap", "main_next_match_button", "button_clicked", "none", "profile", "main_guide_step", "enter_level1", "step_complete", 2010, "cannot_skip", "签约后强引导点下一场"),
        row(2010, "main_to_level1", 2, "on_level1_prebattle_enter", 1, 0, "这里可以看到本场比赛目标。", "bubble", "level_goal_bar", "spotlight", "", "auto_delay", "1.2s", "delay_end", "none", "profile", "main_guide_step", "prebattle_goal_seen", "step_complete", 2011, "cannot_skip", "说明战前页目标条"),
        row(2011, "main_to_level1", 3, "after_goal_bar_tip", 1, 0, "点击开球，开始你的第一场比赛。", "bubble", "kickoff_button", "spotlight", "点击手指点按黄色开球按钮。", "tap", "kickoff_button", "button_clicked", "none", "runtime", "level1_attempt_started", "true", "button_clicked", 0, "cannot_skip", "点击开球创建attempt"),
        row(3001, "level1_tutorial", 1, "on_level1_attempt_started", 1, 0, "比赛即将开始。", "bottom_bar", "stadium_entry", "none", "", "none", "auto", "entry_camera_end", "none", "runtime", "guide_level_step", "entry_done", "step_complete", 3010, "cannot_skip", "入场/开球/黑幕转切片"),
        row(3010, "level1_tutorial", 2, "on_slice_201_enter", 1, 201, "黄色光圈表示当前由你操作的球员。", "bubble", "controlled_player_ring", "spotlight", "", "tap", "guide_confirm_button", "button_clicked", "none", "runtime", "guide_level_step", "201_player_focus_done", "step_complete", 3011, "cannot_skip", "当前操作球员"),
        row(3011, "level1_tutorial", 3, "after_player_ring_tip", 1, 201, "这个标志表示这个方向上有你的队友。", "bubble", "teammate_indicator", "spotlight", "点击手指点按明白了按钮。", "tap", "guide_confirm_button", "button_clicked", "none", "runtime", "guide_level_step", "201_teammate_indicator_done", "step_complete", 3012, "cannot_skip", "队友方向浮标"),
        row(3012, "level1_tutorial", 4, "after_teammate_indicator_tip", 1, 201, "在屏幕上半部分左右滑动，可以移动相机观察场上情况。", "bottom_bar", "camera_swipe_area", "spotlight", "手指在上半屏左右滑动一次，循环播放。", "swipe", "camera_swipe_area", "camera_swiped_or_delay", "none", "runtime", "guide_level_step", "201_camera_done", "step_complete", 3013, "auto_after_delay", "视角滑动教学"),
        row(3013, "level1_tutorial", 5, "after_camera_tip", 1, 201, "现在把球踢向白色目标区域。", "bottom_bar", "slice_201_target", "spotlight", "按当前锁定前模式播放射门手势：划线或弹弓。", "release", "slice_201_target", "player_release_ball", "show_rewind", "runtime", "guide_level_step", "201_shoot_done", "step_complete", 3014, "cannot_skip", "基础射门;首个非守门切片锁模式"),
        row(3014, "level1_tutorial", 6, "on_slice_201_success", 1, 201, "进攻完成，点击继续。", "toast", "replay_continue_button", "none", "点击手指点按继续按钮。", "click_continue", "replay_continue_button", "button_clicked", "none", "attempt", "slice_201_done", "true", "click_continue", 3020, "cannot_skip", "成功回放后继续"),
        row(3015, "level1_tutorial", 7, "on_slice_201_fail", 1, 201, "错过这次机会可能会影响比赛结果，可以使用回溯重试。", "bubble", "rewind_popup", "spotlight", "点击手指点按免费回溯按钮。", "tap", "rewind_free_button", "button_clicked", "continue_as_failed", "profile", "rewind_tutorial_seen", "true", "button_clicked", 3013, "cannot_skip", "第一次失败教回溯"),
        row(3020, "level1_tutorial", 8, "on_slice_202_enter", 1, 202, "有些机会需要先传给队友。把球传到队友脚下。", "bottom_bar", "teammate_target", "spotlight", "手指从主角脚下拖向队友目标区域。", "drag", "teammate_target", "pass_released_to_teammate_area", "show_rewind", "runtime", "guide_level_step", "202_pass_prompt_done", "step_complete", 3021, "cannot_skip", "助攻教学"),
        row(3021, "level1_tutorial", 9, "on_pass_to_teammate_success", 1, 202, "传球成功后，队友会完成射门，这会计为你的助攻。", "bottom_bar", "teammate_shoot_target", "spotlight", "", "none", "auto", "teammate_shot_end", "none", "runtime", "guide_level_step", "202_teammate_shot_seen", "step_complete", 3022, "cannot_skip", "队友自动射门表现"),
        row(3022, "level1_tutorial", 10, "on_slice_202_success", 1, 202, "助攻完成，点击继续。", "toast", "replay_continue_button", "none", "点击手指点按继续按钮。", "click_continue", "replay_continue_button", "button_clicked", "none", "attempt", "slice_202_done", "true", "click_continue", 3030, "cannot_skip", "助攻成功继续"),
        row(3023, "level1_tutorial", 11, "on_slice_202_fail", 1, 202, "传球没有到位，可以回溯后再试一次。", "bubble", "rewind_popup", "spotlight", "点击手指点按回溯按钮；若玩家放弃则按失败继续。", "tap", "rewind_free_or_confirm_button", "button_clicked_or_give_up", "continue_as_failed", "", "", "", "", 3020, "allow_skip_if_seen", "助攻失败可回溯"),
        row(3030, "level1_tutorial", 12, "on_slice_203_enter", 1, 203, "现在轮到你防守。看准射门方向，划线完成扑救。", "bottom_bar", "goalkeeper", "spotlight", "手指从门将位置滑向来球方向。", "swipe", "save_direction_area", "goalkeep_swipe_finished", "show_rewind", "runtime", "guide_level_step", "203_goalkeep_prompt_done", "step_complete", 3031, "cannot_skip", "守门强制划线"),
        row(3031, "level1_tutorial", 13, "on_slice_203_success", 1, 203, "扑救完成，点击继续。", "toast", "replay_continue_button", "none", "点击手指点按继续按钮。", "click_continue", "replay_continue_button", "button_clicked", "none", "attempt", "slice_203_done", "true", "click_continue", 3040, "cannot_skip", "守门成功继续"),
        row(3032, "level1_tutorial", 14, "on_slice_203_fail", 1, 203, "再试一次，沿着来球方向滑动。", "bubble", "save_direction_area", "spotlight", "重复播放门将滑向来球方向的手势；若玩家放弃则按失败继续。", "swipe", "save_direction_area", "goalkeep_swipe_finished_or_give_up", "continue_as_failed", "", "", "", "", 3030, "allow_skip_if_seen", "守门失败提示"),
        row(3040, "level1_tutorial", 15, "on_level1_all_slices_done", 1, 0, "比赛结束，查看本场奖励。", "popup", "level_settlement_confirm", "full", "点击手指点按结算确认按钮。", "tap", "level_settlement_confirm", "button_clicked", "none", "attempt", "level1_settlement_confirmed", "true", "button_clicked", 3041, "cannot_skip", "关卡结算确认"),
        row(3041, "level1_tutorial", 16, "on_rank_settlement_show", 1, 0, "这里会展示本轮后的排名变化。", "popup", "rank_settlement_confirm", "full", "点击手指点按排名结算确认按钮。", "tap", "rank_settlement_confirm", "button_clicked", "none", "profile", "first_level_tutorial_done", "true", "button_clicked", 0, "cannot_skip", "排名确认，引导完成"),
    ]

    def first_expanded_id(source_id: int) -> int:
        return source_id * 10 + 1

    def remap_next(source_next_id: int) -> int:
        return first_expanded_id(source_next_id) if source_next_id else 0

    expanded_rows: list[dict] = []
    for src in source_rows:
        source_id = int(src["ID"])
        has_dialogue = bool(src.get("DialogueLcKey"))
        has_gesture = bool(src.get("GestureDesc"))
        if has_dialogue and has_gesture:
            dialogue_row = src.copy()
            dialogue_row["ID"] = first_expanded_id(source_id)
            dialogue_row["GestureDesc"] = ""
            dialogue_row["WaitType"] = "tap"
            dialogue_row["WaitTarget"] = "guide_confirm_button"
            dialogue_row["PassCondition"] = "button_clicked"
            dialogue_row["FailAction"] = "none"
            dialogue_row["SaveScope"] = ""
            dialogue_row["SaveKey"] = ""
            dialogue_row["SaveValue"] = ""
            dialogue_row["SaveTiming"] = ""
            dialogue_row["NextID"] = source_id * 10 + 2
            dialogue_row["Remark"] = f"{src.get('Remark', '')};对话步骤".strip(";")
            expanded_rows.append(dialogue_row)

            gesture_row = src.copy()
            gesture_row["ID"] = source_id * 10 + 2
            gesture_row["DialogueLcKey"] = ""
            gesture_row["TextStyle"] = ""
            gesture_row["NextID"] = remap_next(int(src["NextID"]))
            gesture_row["Remark"] = f"{src.get('Remark', '')};手势/等待操作步骤".strip(";")
            expanded_rows.append(gesture_row)
        else:
            expanded_row = src.copy()
            expanded_row["ID"] = first_expanded_id(source_id)
            expanded_row["NextID"] = remap_next(int(src["NextID"]))
            expanded_rows.append(expanded_row)

    step_index_by_guide: dict[str, int] = {}
    for expanded_row in expanded_rows:
        guide_id = expanded_row["GuideID"]
        step_index_by_guide[guide_id] = step_index_by_guide.get(guide_id, 0) + 1
        expanded_row["StepIndex"] = step_index_by_guide[guide_id]

    return expanded_rows


# =====================================================================
# 关卡设计 · 参数化生成（50 轮 × 10 关 = 500 关；分层可复用切片实例库）
# 详见 output/关卡设计方案.md。tier = ceil(round/5)，每档 5 轮。
# =====================================================================

ROUNDS_TOTAL = 50
LEVELS_PER_ROUND = 10
TIERS_TOTAL = 10

# slice_type 编号（程序侧 L1 枚举）：1 attack / 2 corner / 3 free_kick / 4 goalkeep / 5 penalty / 6 throw_in
# 该数字编号同时编入 SlicePresetCfg.ID 千位、SliceInstanceCfg.ID 百位
SLICE_TYPE_NAME: dict[int, str] = {
    1: "attack", 2: "corner", 3: "free_kick",
    4: "goalkeep", 5: "penalty", 6: "throw_in",
}
SLICE_TYPE_ORDER = [1, 2, 3, 4, 5, 6]

# 手工编排 preset 旧 ID → 新 ID（按 SliceType 千位 + 类型内连续编号 重新分段）
# NameLcKey 保持 ActvSoccer_preset_name_{old} 不变，仅 ID 改变。
MANUAL_PRESET_ID_REMAP: dict[int, int] = {
    # attack 1xxx
    1: 1001, 5: 1002, 6: 1003, 16: 1004,
    19: 1005, 20: 1006, 21: 1007, 22: 1008, 23: 1009, 24: 1010,
    25: 1011, 26: 1012, 27: 1013, 28: 1014,
    # corner 2xxx
    10: 2001, 11: 2002, 18: 2003,
    37: 2004, 38: 2005, 39: 2006, 40: 2007, 41: 2008, 42: 2009, 43: 2010,
    # free_kick 3xxx
    2: 3001, 7: 3002, 8: 3003, 17: 3004,
    29: 3005, 30: 3006, 31: 3007, 32: 3008, 33: 3009,
    # goalkeep 4xxx
    4: 4001, 14: 4002, 15: 4003, 50: 4004,
    # penalty 5xxx
    3: 5001, 9: 5002, 34: 5003, 35: 5004, 36: 5005,
    # throw_in 6xxx
    12: 6001, 13: 6002,
    44: 6003, 45: 6004, 46: 6005, 47: 6006, 48: 6007, 49: 6008,
}
# 参考截图 preset 旧 9xxx → 新 ID（在对应类型段末尾续编）
REFERENCE_PRESET_ID_REMAP: dict[int, int] = {
    # attack 续 1015-1033
    9001: 1015, 9002: 1016, 9003: 1017, 9004: 1018, 9005: 1019,
    9006: 1020, 9007: 1021, 9008: 1022, 9009: 1023, 9010: 1024,
    9011: 1025, 9012: 1026, 9013: 1027, 9014: 1028, 9015: 1029,
    9016: 1030, 9017: 1031, 9018: 1032, 9019: 1033,
    # corner 续 2011-2012
    9026: 2011, 9027: 2012,
    # free_kick 续 3010-3014
    9020: 3010, 9021: 3011, 9022: 3012, 9023: 3013, 9024: 3014,
    # goalkeep 续 4005
    9029: 4005,
    # penalty 续 5006
    9025: 5006,
    # throw_in 续 6009
    9028: 6009,
}

# 每个 slice_type 的可用 preset 池（按 tier 轮换取用，丰富画面）；ID 千位=类型编号
PRESET_POOL: dict[int, list[int]] = {
    1: [1001, 1002, 1003, 1004, 1005, 1006, 1007, 1008, 1009, 1010,
        1011, 1012, 1013, 1014,
        1015, 1016, 1017, 1018, 1019, 1020, 1021, 1022, 1023, 1024,
        1025, 1026, 1027, 1028, 1029, 1030, 1031, 1032, 1033],  # attack
    2: [2001, 2002, 2003, 2004, 2005, 2006, 2007, 2008, 2009, 2010,
        2011, 2012],                                            # corner
    3: [3001, 3002, 3003, 3004, 3005, 3006, 3007, 3008, 3009,
        3010, 3011, 3012, 3013, 3014],                           # free_kick
    4: [4001, 4002, 4003, 4004, 4005],                          # goalkeep
    5: [5001, 5002, 5003, 5004, 5005, 5006],                    # penalty
    6: [6001, 6002, 6003, 6004, 6005, 6006, 6007, 6008, 6009],  # throw_in
}

# preset 池起始偏移：attack 池前 3 个(1001-1003)预留给试训/引导关，常规库实例从 1004 起。
PRESET_POOL_OFFSET: dict[int, int] = {1: 3}
SLICE_TYPE_VARIANT_COUNT: dict[int, int] = {
    1: 5,  # attack
    2: 3,  # corner
    3: 4,  # free_kick
    4: 2,  # goalkeep
    5: 2,  # penalty
    6: 2,  # throw_in
}

# 关卡应用权重：进攻最高、任意球次之；角球/界外球/点球低频，守门作为特殊调剂。
WEIGHTED_SLICE_TYPE_ORDER = [
    1, 1, 3, 1, 2,
    1, 3, 1, 6, 3,
    1, 5, 1, 3, 4,
    1, 2, 3, 1, 6,
]

# 实例层首脚可操作夹角曲线，tier1 → tier10 单调收紧。
SLICE_TYPE_ANGLE_RANGE: dict[int, tuple[float, float]] = {
    1: (55.0, 38.0),  # attack
    2: (42.0, 30.0),  # corner
    3: (36.0, 24.0),  # free_kick
    4: (0.0, 0.0),    # goalkeep: 专属判定
    5: (0.0, 0.0),    # penalty: 专属判定
    6: (42.0, 30.0),  # throw_in
}

# tier 主题（联赛名 + 主题对手球队池）
TIER_THEME: dict[int, str] = {
    1: "社区杯", 2: "城市联赛", 3: "省级联赛", 4: "大区联赛", 5: "全国联赛",
    6: "洲际资格赛", 7: "洲际联赛", 8: "国际邀请赛", 9: "世界精英赛", 10: "世界杯总决赛",
}

# 淘汰赛开放：round 15（tier3 末）→ level_id = (15-1)*10 + 1 = 141
KNOCKOUT_OPEN_ROUND = 15
KNOCKOUT_OPEN_LEVEL = (KNOCKOUT_OPEN_ROUND - 1) * LEVELS_PER_ROUND + 1


def _tier_specs() -> dict[int, dict]:
    """10 档单一真源：AiProfile / Level / SliceInstance AI字段 / 角度 override 共用。"""
    gk = [25, 30, 36, 42, 48, 54, 60, 66, 72, 78]
    dfn = [12, 18, 24, 30, 36, 42, 48, 54, 60, 66]
    sht = [30, 36, 42, 48, 54, 60, 66, 72, 78, 84]
    react = [1300, 1200, 1100, 1000, 950, 880, 820, 760, 700, 640]
    span_min = [40, 38, 36, 34, 32, 30, 28, 26, 24, 22]
    span_max = [70, 66, 62, 58, 54, 50, 46, 42, 36, 30]
    center_shift = [0, 0, 3, 3, 5, 5, 7, 7, 9, 10]
    margin = [8, 8, 7, 7, 6, 6, 5, 5, 4, 4]
    specs: dict[int, dict] = {}
    for t in range(1, TIERS_TOTAL + 1):
        i = t - 1
        difficulty = "easy" if t <= 2 else "normal" if t <= 5 else "hard"
        specs[t] = {
            "ai_profile_id": 1000 + t,
            "difficulty": difficulty,
            "gk_save": gk[i],
            "def_success": dfn[i],
            "shooter_success": sht[i],
            "react_ms": react[i],
            "opponent_star": math.ceil(t / 2),
            "span_min": span_min[i],
            "span_max": span_max[i],
            "center_shift": center_shift[i],
            "margin": margin[i],
        }
    return specs


TIER = _tier_specs()


def _tier_primary_modifier(tier: int, slice_type: int | None = None) -> int:
    """SliceInstanceCfg.ModifierID 单一主机制（0=无）。"""
    if tier <= 2:
        return 0
    if tier <= 4:
        return 4001
    if tier <= 6:
        return 4002
    if tier <= 8:
        if slice_type in (4, 5):  # goalkeep / penalty 不收角
            return 0
        return 4006
    return 4005


def _instance_operable_angle(tier: int, slice_type: int) -> float:
    """点球/守门无扇形(0)；其余按类型曲线随 tier 单调收窄。"""
    angle_max, angle_min = SLICE_TYPE_ANGLE_RANGE[slice_type]
    if angle_max == 0.0:
        return 0.0
    val = angle_max - (angle_max - angle_min) * (tier - 1) / 9.0
    return round(min(70.0, max(20.0, val)), 1)


def _instance_modifiers_json(tier: int, slice_type: int, variant: int) -> str:
    """实例 ext[] Modifiers（SoccerModifier_V: {id, params}）。"""
    mods: list[dict] = []
    if 3 <= tier <= 4:
        mods.append({"id": "moving_keeper", "params": {"speed": 1.0}})
    elif 5 <= tier <= 6:
        mods.append({"id": "moving_keeper", "params": {"speed": 1.5}})
    elif 7 <= tier <= 8 and slice_type not in (4, 5):  # goalkeep/penalty 不收角
        mods.append({"id": "narrow_angle", "params": {"shrink": 0.7}})
    elif tier >= 9:
        mods.append({"id": "moving_keeper", "params": {"speed": 2.0}})
    if slice_type == 3 and tier >= 6:  # 任意球固定人墙
        mods.append({"id": "fixed_wall", "params": {}})
    if variant == 2 and slice_type in (3, 5):  # 任意球/点球加压：关辅助线
        mods.append({"id": "no_aim_line", "params": {}})
    if variant == 2 and slice_type == 4:  # 守门加压：门将移动
        if not any(m["id"] == "moving_keeper" for m in mods):
            mods.append({"id": "moving_keeper", "params": {"speed": 1.5}})
    if tier == 10:
        mods.append({"id": "random_dive", "params": {"randomness": 0.6}})
    return json.dumps(mods, ensure_ascii=False) if mods else "[]"


def _instance_ai_cols(iid: int, slice_type: int, *, is_guide: int = 0) -> dict:
    """返回并入 SliceInstanceCfg 的 AI 字段。

    新 ID 编码：iid = 1{type_digit}{seq:02d}，其中
      type_digit 1=attack 2=corner 3=free_kick 4=goalkeep 5=penalty 6=throw_in
      seq = (tier-1)*variant_count + variant，从 1 开始
    tier 取 ceil(seq / variant_count)。
    """
    type_digit = (iid // 100) - 10
    seq = iid % 100
    variant_count = SLICE_TYPE_VARIANT_COUNT[type_digit]
    tier = (seq - 1) // variant_count + 1
    s = TIER.get(tier, TIER[1])
    if slice_type == 4:        # 守门：玩家为门将，对方后卫射门
        gk_ai, def_ai, shooter_ai, mod = 0, 0, _enemy_id(tier, 3), 0
    elif slice_type == 5:      # 点球：有对方门将，无后卫
        gk_ai, def_ai, shooter_ai, mod = _enemy_id(tier, 1), 0, 0, _tier_primary_modifier(tier, slice_type)
    else:                      # 进攻/角球/任意球/界外球
        gk_ai, def_ai, shooter_ai, mod = _enemy_id(tier, 1), _enemy_id(tier, 2), 0, _tier_primary_modifier(tier, slice_type)
    return {
        "AiProfileID": s["ai_profile_id"],
        "GoalkeeperAiID": gk_ai,
        "DefenderAiID": def_ai,
        "ShooterAiID": shooter_ai,
        "ModifierID": mod,
        "IsGuideAi": is_guide,
        "RewindRandom": 1,
        "OverrideReactionTimeMs": s["react_ms"],
    }


def _guide_instance_ai_cols(
    ai_profile_id: int,
    goalkeeper_ai_id: int,
    defender_ai_id: int,
    shooter_ai_id: int,
    override_reaction_time_ms: int,
) -> dict:
    return {
        "AiProfileID": ai_profile_id,
        "GoalkeeperAiID": goalkeeper_ai_id,
        "DefenderAiID": defender_ai_id,
        "ShooterAiID": shooter_ai_id,
        "ModifierID": 0,
        "IsGuideAi": 1,
        "RewindRandom": 1,
        "OverrideReactionTimeMs": override_reaction_time_ms,
    }


def _build_instance_library() -> list[dict]:
    """类型差异化库实例；旧 6 行(101-203)前置保留。
    切片胜利条件由 SliceType 决定(attack/free_kick/penalty/corner/throw_in→进球;
    goalkeep→不被进球),不在实例层覆盖。

    ID 编码：1{type_digit}{seq:02d}
      type_digit 1=attack 2=corner 3=free_kick 4=goalkeep 5=penalty 6=throw_in
      seq = (tier-1)*variant_count + variant
    e.g. 1101=attack tier1 v1, 1201=corner tier1 v1, 1601=throw_in tier1 v1。
    """
    legacy = [
        {"ID": 101, "SliceType": "attack", "PresetID": 1001, "Remark": "试训-进攻",
         **_guide_instance_ai_cols(1001, 2001, 0, 0, 1200)},
        {"ID": 102, "SliceType": "free_kick", "PresetID": 3001, "Remark": "试训-任意球",
         **_guide_instance_ai_cols(1001, 2001, 0, 0, 1200)},
        {"ID": 103, "SliceType": "penalty", "PresetID": 5001, "Remark": "试训-点球",
         **_guide_instance_ai_cols(1001, 2001, 0, 0, 1200)},
        {"ID": 201, "SliceType": "attack", "PresetID": 1002, "OverrideOperableAngle": 30, "Remark": "引导关1",
         **_guide_instance_ai_cols(1001, 2001, 0, 0, 1200)},
        {"ID": 202, "SliceType": "attack", "PresetID": 1003, "OverrideOperableAngle": 28, "Remark": "引导关2",
         **_guide_instance_ai_cols(1001, 2001, 2002, 0, 1200)},
        {"ID": 203, "SliceType": "goalkeep", "PresetID": 4001, "Remark": "引导关3-守门",
         **_guide_instance_ai_cols(1002, 0, 0, 2003, 900)},
    ]
    lib: list[dict] = []
    for stype in SLICE_TYPE_ORDER:
        pool = PRESET_POOL[stype]
        offset = PRESET_POOL_OFFSET.get(stype, 0)
        type_name = SLICE_TYPE_NAME[stype]
        variant_count = SLICE_TYPE_VARIANT_COUNT[stype]
        total = TIERS_TOTAL * variant_count
        eff_pool_size = len(pool) - offset
        full_cycles = total // eff_pool_size
        full_count = full_cycles * eff_pool_size
        tail_count = total - full_count
        for tier in range(1, TIERS_TOTAL + 1):
            for variant in range(1, variant_count + 1):
                seq = (tier - 1) * variant_count + variant
                # 完整循环段照常按顺序取；最后不足一轮的尾段右对齐到 pool 末尾，
                # 让后期 tier 用上更靠后的 preset、而非又回到开头。
                if seq <= full_count:
                    pool_idx = offset + (seq - 1) % eff_pool_size
                else:
                    tail_seq = seq - full_count  # 1..tail_count
                    pool_idx = offset + (eff_pool_size - tail_count) + (tail_seq - 1)
                preset_id = pool[pool_idx]
                iid = 1000 + stype * 100 + seq
                row: dict = {
                    "ID": iid,
                    "SliceType": type_name,
                    "PresetID": preset_id,
                    "OverrideOperableAngle": _instance_operable_angle(tier, stype),
                    "Modifiers": _instance_modifiers_json(tier, stype, variant),
                    "Remark": f"库 tier{tier} {type_name} v{variant}",
                }
                row.update(_instance_ai_cols(iid, stype))
                lib.append(row)
    return legacy + lib


def _enemy_band(tier: int) -> int:
    return 1 if tier <= 3 else 2 if tier <= 6 else 3 if tier <= 8 else 4


def _enemy_id(tier: int, role: int) -> int:
    """role: 1=keeper 2=defender 3=shooter。band1 复用旧 2001-2003。"""
    band = _enemy_band(tier)
    if band == 1:
        return {1: 2001, 2: 2002, 3: 2003}[role]
    return 2000 + band * 10 + role


def _build_enemy_ai() -> list[dict]:
    """12 行 = 4 难度带 × 3 角色；band1 复用旧 2001-2003(保引导切片引用)。"""
    rows = [
        {"ID": 2001, "Duty": PLAYER_AI_DUTY_ENUM["Goalkeeper"], "SaveWeight": 45, "LeftWeight": 40, "RightWeight": 40, "UpWeight": 20, "InterceptWeight": 0, "ClearanceWeight": 0, "KeeperCatchFail": 1, "OutOfBoundsFail": 0, "AnimationKey": "E04_GKDiveLeft", "Remark": "band1门将(tier1-3)"},
        {"ID": 2002, "Duty": PLAYER_AI_DUTY_ENUM["Defender"], "SaveWeight": 0, "LeftWeight": 0, "RightWeight": 0, "UpWeight": 0, "InterceptWeight": 35, "ClearanceWeight": 20, "KeeperCatchFail": 0, "OutOfBoundsFail": 1, "AnimationKey": "F03_Intercept", "Remark": "band1后卫(tier1-3)"},
        {"ID": 2003, "Duty": PLAYER_AI_DUTY_ENUM["Defender"], "SaveWeight": 0, "LeftWeight": 40, "RightWeight": 40, "UpWeight": 20, "InterceptWeight": 0, "ClearanceWeight": 0, "KeeperCatchFail": 0, "OutOfBoundsFail": 0, "AnimationKey": "D01_Shoot", "Remark": "band1守门切片射手(tier1-3)"},
    ]
    # band2/3/4 × keeper/def/shooter，权重随带收紧
    band_keeper = {2: (52, 44), 3: (60, 48), 4: (68, 52)}      # (SaveWeight, dir base)
    band_def = {2: 42, 3: 50, 4: 58}                            # InterceptWeight
    band_shooter = {2: 45, 3: 50, 4: 56}                        # dir base
    for band in (2, 3, 4):
        save_w, dir_w = band_keeper[band]
        rows.append({"ID": 2000 + band * 10 + 1, "Duty": PLAYER_AI_DUTY_ENUM["Goalkeeper"], "SaveWeight": save_w, "LeftWeight": dir_w, "RightWeight": dir_w, "UpWeight": 16, "InterceptWeight": 0, "ClearanceWeight": 0, "KeeperCatchFail": 1, "OutOfBoundsFail": 0, "AnimationKey": "E04_GKDiveLeft", "Remark": f"band{band}门将"})
        rows.append({"ID": 2000 + band * 10 + 2, "Duty": PLAYER_AI_DUTY_ENUM["Defender"], "SaveWeight": 0, "LeftWeight": 0, "RightWeight": 0, "UpWeight": 0, "InterceptWeight": band_def[band], "ClearanceWeight": 24, "KeeperCatchFail": 0, "OutOfBoundsFail": 1, "AnimationKey": "F03_Intercept", "Remark": f"band{band}后卫"})
        sw = band_shooter[band]
        rows.append({"ID": 2000 + band * 10 + 3, "Duty": PLAYER_AI_DUTY_ENUM["Defender"], "SaveWeight": 0, "LeftWeight": sw, "RightWeight": sw, "UpWeight": 18, "InterceptWeight": 0, "ClearanceWeight": 0, "KeeperCatchFail": 0, "OutOfBoundsFail": 0, "AnimationKey": "D01_Shoot", "Remark": f"band{band}守门切片射手"})
    return rows


def _build_ai_profiles() -> list[dict]:
    """10 档 AI 难度，单调递增；DeadCornerCanSave 固定 0。"""
    label = {"easy": "简单档", "normal": "普通档", "hard": "困难档"}
    rows = []
    for t in range(1, TIERS_TOTAL + 1):
        s = TIER[t]
        rows.append({
            "ID": s["ai_profile_id"], "Difficulty": s["difficulty"],
            "GoalkeeperSaveRate": s["gk_save"], "DefenderSuccessRate": s["def_success"],
            "ShooterSuccessRate": s["shooter_success"], "DeadCornerCanSave": 0,
            "ReactionTimeMs": s["react_ms"],
            "Remark": f"tier{t} {label[s['difficulty']]}(对手{s['opponent_star']}星)",
        })
    return rows


def _build_ai_modifiers() -> list[dict]:
    """7 行：保留 4001-4004，新增 4005/4006/4007。"""
    return [
        {"ID": 4001, "ModifierType": "moving_keeper", "Param1Key": "speed", "Param1Value": "1.0", "Param2Key": "range", "Param2Value": "2.5", "Param3Key": "start_offset", "Param3Value": "0.0", "Remark": "普通移动门将"},
        {"ID": 4002, "ModifierType": "moving_keeper", "Param1Key": "speed", "Param1Value": "1.5", "Param2Key": "range", "Param2Value": "3.5", "Param3Key": "start_offset", "Param3Value": "0.5", "Remark": "困难移动门将"},
        {"ID": 4003, "ModifierType": "no_aim_line", "Param1Key": "enabled", "Param1Value": "1", "Remark": "关闭辅助线"},
        {"ID": 4004, "ModifierType": "fixed_wall", "Param1Key": "enabled", "Param1Value": "1", "Remark": "任意球固定人墙"},
        {"ID": 4005, "ModifierType": "moving_keeper", "Param1Key": "speed", "Param1Value": "2.0", "Param2Key": "range", "Param2Value": "4.5", "Param3Key": "start_offset", "Param3Value": "0.5", "Remark": "极限移动门将"},
        {"ID": 4006, "ModifierType": "narrow_angle", "Param1Key": "shrink", "Param1Value": "0.7", "Remark": "收窄可操作夹角"},
        {"ID": 4007, "ModifierType": "random_dive", "Param1Key": "randomness", "Param1Value": "0.6", "Remark": "门将随机扑救方向"},
    ]


def _theme_team_pool(tier: int) -> list[int]:
    """该 tier 的 5 个主题对手球队 id（3001-3050）。"""
    base = 3000 + (tier - 1) * 5
    return [base + k for k in range(1, 6)]


def _build_theme_teams(lc: LcRegistry) -> list[dict]:
    """~50 主题对手展示球队(id 3001-3050)；仅展示，复用美术资源键轮换。"""
    regions = ["asia", "europe", "south_america", "africa", "north_america"]
    suffix = ["联", "FC", "竞技", "联合", "勇士"]
    rows = []
    for tier in range(1, TIERS_TOTAL + 1):
        theme = TIER_THEME[tier]
        for k in range(1, 6):
            tid = 3000 + (tier - 1) * 5 + k
            name = f"{theme}·{suffix[k - 1]}"
            kit_idx = ((tid - 1) % 12) + 1
            rows.append({
                "ID": tid,
                "NameLcKey": lc.add(lc_key("team", "name", str(tid)), name, f"TeamCfg/{tid}"),
                "Region": regions[(tier - 1) % len(regions)],
                "KitKey": f"WC_Kit_{kit_idx:02d}",
                "BadgeKey": f"WC_Badge_{kit_idx:02d}",
                "Remark": f"tier{tier}主题对手{k}",
            })
    return rows


def _slice_count(tier: int) -> int:
    if tier == 1:
        return 2
    if tier <= 3:
        return 3
    if tier <= 7:
        return 4
    return 5


def _compose_slice_list(level_in_round: int, tier: int) -> list[int]:
    """按类型权重轮换，引用库实例 id。
    整体占比：进攻最高、任意球次之；角球/界外球/点球低频，守门特殊调剂。
    实例 ID 编码：1{type_digit}{seq:02d}, seq=(tier-1)*variant_count + variant。"""
    n = _slice_count(tier)
    start = ((tier - 1) * LEVELS_PER_ROUND + level_in_round - 1) % len(WEIGHTED_SLICE_TYPE_ORDER)
    out: list[int] = []
    for k in range(n):
        stype = WEIGHTED_SLICE_TYPE_ORDER[(start + k) % len(WEIGHTED_SLICE_TYPE_ORDER)]
        variant_count = SLICE_TYPE_VARIANT_COUNT[stype]
        variant = ((tier + level_in_round + k - 2) % variant_count) + 1
        seq = (tier - 1) * variant_count + variant
        out.append(1000 + stype * 100 + seq)
    return out


def _build_seasons(lc: LcRegistry) -> list[dict]:
    """50 轮联赛(单 group=1)；NextSeason 链推进，最后一轮=0。"""
    rows = []
    for r in range(1, ROUNDS_TOTAL + 1):
        tier = math.ceil(r / 5)
        name = f"{TIER_THEME[tier]} 第{r}/{ROUNDS_TOTAL}轮"
        rows.append({
            "ID": r,
            "LeagueNameLcKey": lc.add(lc_key("season", "league_name", str(r)), name, f"SeasonCfg/{r}"),
            "NextSeason": (r + 1 if r < ROUNDS_TOTAL else 0),
            "ContractOfferCount": 3,
            "Remark": f"第{r}轮 tier{tier}",
        })
    return rows


LEVEL_SOURCE_FILE = OUT_DIR / "source-data" / "level.json"


def _build_levels(lc: LcRegistry) -> list[dict]:
    """500 关：从 source-data/level.json 加载已编排的关卡数据。
    源数据由策划在 xlsx 中编辑后，通过 source-data/extract_levels.py 提取生成。
    第1关为引导关(复用 201/202/203)。"""
    if not LEVEL_SOURCE_FILE.exists():
        raise FileNotFoundError(
            f"缺少关卡源数据: {LEVEL_SOURCE_FILE}。"
            f"可从 ActivitySoccer_preview.xlsx 重新提取: python source-data/extract_levels.py"
        )
    payload = json.loads(LEVEL_SOURCE_FILE.read_text(encoding="utf-8"))
    rows = []
    for row in payload["rows"]:
        sl = row["SliceList"]
        # SliceList 可能是 JSON 字符串或已解析的 list
        if isinstance(sl, list):
            sl_json = json.dumps(sl)
        else:
            sl_json = str(sl)
        is_tut = int(row["IsTutorial"] or 0)
        rid = int(row["ID"])
        season = int(row["SeasonID"])
        tier = math.ceil(season / 5)
        rows.append({
            "ID": rid,
            "IsTutorial": is_tut,
            "SliceList": sl_json,
            "AiProfileID": int(row["AiProfileID"]),
            "WinThreshold": int(row["WinThreshold"]),
            "DrawThreshold": int(row["DrawThreshold"]),
            "TicketCost": int(row["TicketCost"]),
            "OpponentTeamID": int(row["OpponentTeamID"]),
            "OpponentTeamStar": int(row["OpponentTeamStar"]),
            "SeasonID": season,
            "Remark": (
                "第1轮-引导关(含守门切片203)" if rid == 1
                else f"第{season}轮 tier{tier} 第{((rid - 1) % LEVELS_PER_ROUND) + 1}场"
            ),
        })
    return rows


def _preset_angle_cols(slice_type: int, operable_angle: float | None = None) -> dict:
    """Preset 自身 4 角度列跟随该站位的首脚夹角；逐 tier 收窄由实例 override 实现。
    点球/守门无扇形，span 置 0。"""
    if slice_type in (4, 5):  # goalkeep / penalty
        return {"AngleSpanMin": 0.0, "AngleSpanMax": 0.0, "AngleMaxCenterShift": 0.0, "AngleMargin": 0.0}
    angle_max = float(operable_angle if operable_angle is not None else SLICE_TYPE_ANGLE_RANGE[slice_type][0])
    angle_min = max(20.0, min(angle_max, angle_max - 18.0))
    return {
        "AngleSpanMin": round(angle_min, 1),
        "AngleSpanMax": round(angle_max, 1),
        # attack(1)/corner(2)/throw_in(6) 给较宽中心偏移与贴边余量；free_kick(3) 收紧
        "AngleMaxCenterShift": 5.0 if slice_type in (1, 2, 6) else 3.0,
        "AngleMargin": 6.0 if slice_type in (1, 2, 6) else 5.0,
    }


def _build_presets(lc: LcRegistry) -> list[dict]:
    """~18 摆位预设，跨 tier 复用。1/2/3/4 沿用原测试预设，新增侧别/难度变体。
    四角度列存最宽基线(tier1)，逐 tier 收窄由实例 OverrideOperableAngle 实现。"""

    def gk_attack(home_x: float, ball_z: float = -23.0) -> list[dict]:
        keeper_z = away_keeper_z("attack", ball_z)
        support_z = max(FIELD_Z_MID, ball_z - 7)
        weak_side_x = -home_x * 0.45
        box_def_z = min(-5.0, ball_z + 8)
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], home_x, 0, ball_z,
                        _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, home_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], home_x - 2, 0, support_z,
                        _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, home_x - 2, support_z)),
            player_init("home", 2, PLAYER_AI_DUTY_ENUM["Forward"], weak_side_x, 0, max(FIELD_Z_MID, ball_z + 2),
                        _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, weak_side_x, max(FIELD_Z_MID, ball_z + 2))),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(home_x, ball_z, 0, keeper_z)),
            player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], home_x - 4, 0, -10,
                        _face_toward(home_x, ball_z, home_x - 4, -10)),
            player_init("away", 2, PLAYER_AI_DUTY_ENUM["Defender"], home_x * 0.45, 0, box_def_z,
                        _face_toward(home_x, ball_z, home_x * 0.45, box_def_z)),
            player_init("away", 3, PLAYER_AI_DUTY_ENUM["Defender"], weak_side_x, 0, -7,
                        _face_toward(home_x, ball_z, weak_side_x, -7)),
        ]

    def free_kick_wall(ball_x: float, ball_z: float = -16.0) -> list[dict]:
        wall_z = min(-7.0, ball_z + PENALTY_FREE_RADIUS)
        wall_xs = (-3.0, -1.5, 0.0, 1.5)
        keeper_z = away_keeper_z("free_kick", ball_z)
        support_x = max(-14.0, min(14.0, ball_x + (4 if ball_x <= 0 else -4)))
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], ball_x, 0, ball_z,
                        _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, ball_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], support_x, 0, ball_z + 1.5,
                        _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, support_x, ball_z + 1.5)),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(ball_x, ball_z, 0, keeper_z)),
            *[
                player_init("away", idx, PLAYER_AI_DUTY_ENUM["Defender"], wx, 0, wall_z,
                            _face_toward(ball_x, ball_z, wx, wall_z))
                for idx, wx in enumerate(wall_xs, start=1)
            ],
        ]

    def corner_players(side_x: float) -> list[dict]:
        ball_z = -1.0
        keeper_z = away_keeper_z("corner")
        near_x = side_x * 0.35
        far_x = -side_x * 0.35
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], side_x, 0, ball_z,
                        _face_toward(0, 0, side_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], near_x, 0, -4.0,
                        _face_toward(side_x, ball_z, near_x, -4.0)),
            player_init("home", 2, PLAYER_AI_DUTY_ENUM["Forward"], 0, 0, -6.0,
                        _face_toward(side_x, ball_z, 0, -6.0)),
            player_init("home", 3, PLAYER_AI_DUTY_ENUM["Forward"], far_x, 0, -7.5,
                        _face_toward(side_x, ball_z, far_x, -7.5)),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(side_x, ball_z, 0, keeper_z)),
            player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], near_x * 0.8, 0, -3.0,
                        _face_toward(side_x, ball_z, near_x * 0.8, -3.0)),
            player_init("away", 2, PLAYER_AI_DUTY_ENUM["Defender"], 0, 0, -5.5,
                        _face_toward(side_x, ball_z, 0, -5.5)),
            player_init("away", 3, PLAYER_AI_DUTY_ENUM["Defender"], far_x * 0.8, 0, -7.0,
                        _face_toward(side_x, ball_z, far_x * 0.8, -7.0)),
            player_init("away", 4, PLAYER_AI_DUTY_ENUM["Defender"], -side_x * 0.15, 0, -9.0,
                        _face_toward(side_x, ball_z, -side_x * 0.15, -9.0)),
        ]

    def throw_in_players(side_x: float, ball_z: float = -16.0) -> list[dict]:
        throw_x = max(-FIELD_X_HALF, min(FIELD_X_HALF, side_x))
        recv_x = max(-16.0, min(16.0, throw_x - (3 if throw_x > 0 else -3)))
        recv_z = max(FIELD_Z_MID, ball_z + 4)
        keeper_z = away_keeper_z("throw_in")
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], throw_x, 0, ball_z,
                        _face_toward(recv_x, recv_z, throw_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], recv_x, 0, recv_z,
                        _face_toward(throw_x, ball_z, recv_x, recv_z)),
            player_init("home", 2, PLAYER_AI_DUTY_ENUM["Forward"], recv_x * 0.45, 0, recv_z - 5,
                        _face_toward(throw_x, ball_z, recv_x * 0.45, recv_z - 5)),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(throw_x, ball_z, 0, keeper_z)),
            player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], recv_x, 0, -10,
                        _face_toward(throw_x, ball_z, recv_x, -10)),
            player_init("away", 2, PLAYER_AI_DUTY_ENUM["Defender"], recv_x * 0.45, 0, min(-5, recv_z + 2),
                        _face_toward(throw_x, ball_z, recv_x * 0.45, min(-5, recv_z + 2))),
        ]

    def penalty_players() -> list[dict]:
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], 0, 0, PENALTY_SPOT[2], 0.0),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, 0, 180.0),
        ]

    def goalkeep_players() -> list[dict]:
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], 0, 0, -2, 180.0),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Defender"], 0, 0, -18, 0.0),
        ]

    def vec_to(ball_x: float, ball_z: float, target_x: float, target_z: float) -> str:
        dx, dz = target_x - ball_x, target_z - ball_z
        length = math.hypot(dx, dz) or 1.0
        return json.dumps({"x": round(dx / length, 3), "y": 0, "z": round(dz / length, 3)})

    def clamp_x(x: float, margin: float = 0.5) -> float:
        return max(-FIELD_X_HALF + margin, min(FIELD_X_HALF - margin, x))

    def attack_setup(ball_x: float, ball_z: float, target_x: float = 0.0, support_x: float | None = None) -> list[dict]:
        support_x = ball_x - 2 if support_x is None else support_x
        support_z = max(FIELD_Z_MID, ball_z - 7)
        defender_x = max(-FIELD_X_HALF, min(FIELD_X_HALF, ball_x * 0.55))
        defender_z = min(FIELD_Z_NEAR, max(PENALTY_AREA_Z_FAR, ball_z + 10))
        weak_side_x = max(-FIELD_X_HALF, min(FIELD_X_HALF, -ball_x * 0.55 if abs(ball_x) > 2 else ball_x + 6))
        box_mid_z = min(-4.0, ball_z + 8)
        box_side_z = min(-5.0, ball_z + 11)
        keeper_z = away_keeper_z("attack", ball_z)
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], ball_x, 0, ball_z,
                        _face_toward(target_x, 0, ball_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], support_x, 0, support_z,
                        _face_toward(target_x, 0, support_x, support_z)),
            player_init("home", 2, PLAYER_AI_DUTY_ENUM["Forward"], weak_side_x, 0, max(FIELD_Z_MID, ball_z + 1),
                        _face_toward(target_x, 0, weak_side_x, max(FIELD_Z_MID, ball_z + 1))),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(ball_x, ball_z, 0, keeper_z)),
            player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], defender_x, 0, defender_z,
                        _face_toward(ball_x, ball_z, defender_x, defender_z)),
            player_init("away", 2, PLAYER_AI_DUTY_ENUM["Defender"], max(-FIELD_X_HALF, min(FIELD_X_HALF, ball_x * 0.25)), 0, box_mid_z,
                        _face_toward(ball_x, ball_z, max(-FIELD_X_HALF, min(FIELD_X_HALF, ball_x * 0.25)), box_mid_z)),
            player_init("away", 3, PLAYER_AI_DUTY_ENUM["Defender"], weak_side_x * 0.55, 0, box_side_z,
                        _face_toward(ball_x, ball_z, weak_side_x * 0.55, box_side_z)),
        ]

    def corner_setup(side_x: float, target_x: float, target_z: float = 0.0) -> list[dict]:
        keeper_z = away_keeper_z("corner")
        ball_z = -1.0
        home2_x = -target_x if target_x else 3
        far_x = -side_x * 0.3
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], side_x, 0, ball_z,
                        _face_toward(target_x, target_z, side_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], target_x, 0, -6,
                        _face_toward(side_x, ball_z, target_x, -6)),
            player_init("home", 2, PLAYER_AI_DUTY_ENUM["Forward"], home2_x, 0, -7,
                        _face_toward(side_x, ball_z, home2_x, -7)),
            player_init("home", 3, PLAYER_AI_DUTY_ENUM["Forward"], far_x, 0, -4.5,
                        _face_toward(side_x, ball_z, far_x, -4.5)),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(side_x, ball_z, 0, keeper_z)),
            player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], target_x * 0.5, 0, -8,
                        _face_toward(side_x, ball_z, target_x * 0.5, -8)),
            player_init("away", 2, PLAYER_AI_DUTY_ENUM["Defender"], target_x, 0, -4,
                        _face_toward(side_x, ball_z, target_x, -4)),
            player_init("away", 3, PLAYER_AI_DUTY_ENUM["Defender"], home2_x * 0.5, 0, -6,
                        _face_toward(side_x, ball_z, home2_x * 0.5, -6)),
            player_init("away", 4, PLAYER_AI_DUTY_ENUM["Defender"], far_x * 0.5, 0, -9,
                        _face_toward(side_x, ball_z, far_x * 0.5, -9)),
        ]

    def throw_in_setup(side_x: float, ball_z: float, target_x: float, target_z: float) -> list[dict]:
        throw_x = max(-FIELD_X_HALF, min(FIELD_X_HALF, side_x))
        support_x = max(-16.0, min(16.0, target_x * 0.5))
        keeper_z = away_keeper_z("throw_in")
        defender_z = min(-3, target_z + 4)
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], throw_x, 0, ball_z,
                        _face_toward(target_x, target_z, throw_x, ball_z)),
            player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], target_x, 0, target_z,
                        _face_toward(throw_x, ball_z, target_x, target_z)),
            player_init("home", 2, PLAYER_AI_DUTY_ENUM["Forward"], support_x, 0, target_z - 6,
                        _face_toward(throw_x, ball_z, support_x, target_z - 6)),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, keeper_z,
                        _face_toward(throw_x, ball_z, 0, keeper_z)),
            player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], target_x, 0, defender_z,
                        _face_toward(throw_x, ball_z, target_x, defender_z)),
            player_init("away", 2, PLAYER_AI_DUTY_ENUM["Defender"], support_x, 0, min(-5, target_z - 2),
                        _face_toward(throw_x, ball_z, support_x, min(-5, target_z - 2))),
        ]

    def goalkeep_setup(shot_x: float, shot_z: float) -> list[dict]:
        return [
            player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], 0, 0, -2,
                        _face_toward(shot_x, shot_z, 0, -2)),
            player_init("away", 0, PLAYER_AI_DUTY_ENUM["Defender"], shot_x, 0, shot_z,
                        _face_toward(0, -2, shot_x, shot_z)),
        ]

    def align_ball_with_owner(
        stype: str,
        ball_pos: str,
        owner: int,
        players: list[dict],
    ) -> tuple[str, list[dict]]:
        if stype == "goalkeep":
            return ball_pos, players
        ball = json.loads(ball_pos)
        owner_player = next(
            player for player in players
            if player["team"] == "home" and int(player["idx"]) == int(owner)
        )
        forward_x, forward_z = _forward_from_facing(float(owner_player["facing"]))
        if stype in {"penalty", "corner"}:
            owner_player["pos"]["x"] = _round_coord(float(ball["x"]) - forward_x * BALL_CONTROL_DISTANCE)
            owner_player["pos"]["z"] = _round_coord(float(ball["z"]) - forward_z * BALL_CONTROL_DISTANCE)
            ball["x"] = _round_coord(float(ball["x"]))
            ball["z"] = _round_coord(float(ball["z"]))
            return json.dumps(ball), players
        ball["x"] = _round_coord(float(owner_player["pos"]["x"]) + forward_x * BALL_CONTROL_DISTANCE)
        ball["z"] = _round_coord(float(owner_player["pos"]["z"]) + forward_z * BALL_CONTROL_DISTANCE)
        return json.dumps(ball), players

    def normalize_free_kick_wall(ball_pos: str, target: str | None, players: list[dict]) -> list[dict]:
        ball = json.loads(ball_pos)
        target_pos = json.loads(target) if target else {"x": GOAL_CENTER_X, "z": GOAL_CENTER_Z}
        ball_x, ball_z = float(ball["x"]), float(ball["z"])
        target_x, target_z = float(target_pos["x"]), float(target_pos["z"])
        dx, dz = target_x - ball_x, target_z - ball_z
        length = math.hypot(dx, dz) or 1.0
        ux, uz = dx / length, dz / length
        wall_distance = min(max(5.6, length * 0.45), max(5.6, length - 1.0))
        center_x = ball_x + ux * wall_distance
        center_z = ball_z + uz * wall_distance
        perp_x, perp_z = -uz, ux
        defenders = [
            player for player in players
            if player["team"] == "away" and int(player["duty"]) == PLAYER_AI_DUTY_ENUM["Defender"]
        ]
        wall_gap = 1.5  # 留出四舍五入余量，同时保证内侧人墙能遮挡射线。
        offsets = [
            (idx - (len(defenders) - 1) / 2) * wall_gap
            for idx in range(len(defenders))
        ]
        for defender, offset in zip(defenders, offsets):
            x = clamp_x(center_x + perp_x * offset)
            z = max(FIELD_Z_FAR, min(FIELD_Z_NEAR, center_z + perp_z * offset))
            defender["pos"]["x"] = _round_coord(x)
            defender["pos"]["z"] = _round_coord(z)
            defender["facing"] = _face_toward(ball_x, ball_z, x, z)
        return players

    def normalize_corner_receivers(
        ball_pos: str,
        ball_vec: str,
        target: str | None,
        owner: int,
        players: list[dict],
    ) -> list[dict]:
        if not target:
            return players
        ball = json.loads(ball_pos)
        vector = json.loads(ball_vec)
        target_pos = json.loads(target)
        ball_x, ball_z = float(ball["x"]), float(ball["z"])
        target_x, target_z = float(target_pos["x"]), float(target_pos["z"])
        goalward_target = target_z > ball_z + 1.0 or float(vector.get("z", 0)) > 0.3
        receivers = [
            player for player in players
            if player["team"] == "home" and int(player["idx"]) != int(owner)
        ]
        if not receivers:
            return players
        side = -1.0 if ball_x < 0 else 1.0
        anchor_z = max(target_z - 0.8, ball_z - 0.8) if goalward_target else target_z
        planned = [
            (target_x, anchor_z),
            (target_x - side * 3.0, anchor_z - 1.4),
            (target_x + side * 3.2, anchor_z - 1.8),
            (target_x - side * 5.6, anchor_z - 3.2),
        ]
        for receiver, (x, z) in zip(sorted(receivers, key=lambda p: int(p["idx"])), planned):
            x = clamp_x(x)
            z = max(FIELD_Z_FAR, min(FIELD_Z_NEAR, z))
            receiver["pos"]["x"] = _round_coord(x)
            receiver["pos"]["z"] = _round_coord(z)
            receiver["facing"] = _face_toward(ball_x, ball_z, x, z)
        return players

    def normalize_attack_receivers(tags: str, target: str | None, owner: int, players: list[dict]) -> list[dict]:
        if not target:
            return players
        try:
            tag_set = set(json.loads(tags))
        except json.JSONDecodeError:
            tag_set = set()
        if not (tag_set & {"tap_in", "rebound"}):
            return players
        target_pos = json.loads(target)
        target_x, target_z = float(target_pos["x"]), float(target_pos["z"])
        receivers = [
            player for player in players
            if player["team"] == "home" and int(player["idx"]) != int(owner)
        ]
        if not receivers:
            return players
        receiver = sorted(receivers, key=lambda p: int(p["idx"]))[-1]
        x = clamp_x(target_x)
        z = max(FIELD_Z_FAR, min(FIELD_Z_NEAR, target_z - 2.0))
        receiver["pos"]["x"] = _round_coord(x)
        receiver["pos"]["z"] = _round_coord(z)
        receiver["facing"] = _face_toward(target_x, target_z, x, z)
        return players

    def normalize_throw_in_receivers(target: str | None, owner: int, players: list[dict]) -> list[dict]:
        if not target:
            return players
        target_pos = json.loads(target)
        target_x, target_z = float(target_pos["x"]), float(target_pos["z"])
        receivers = [
            player for player in players
            if player["team"] == "home" and int(player["idx"]) != int(owner)
        ]
        if not receivers:
            return players
        receiver = sorted(receivers, key=lambda p: int(p["idx"]))[0]
        receiver["pos"]["x"] = _round_coord(clamp_x(target_x))
        receiver["pos"]["z"] = _round_coord(max(FIELD_Z_FAR, min(FIELD_Z_NEAR, target_z)))
        owner_player = next(
            player for player in players
            if player["team"] == "home" and int(player["idx"]) == int(owner)
        )
        receiver["facing"] = _face_toward(
            float(owner_player["pos"]["x"]),
            float(owner_player["pos"]["z"]),
            receiver["pos"]["x"],
            receiver["pos"]["z"],
        )
        return players

    def normalize_playable_spacing(
        stype: str,
        tags: str,
        ball_pos: str,
        ball_vec: str,
        target: str | None,
        owner: int,
        players: list[dict],
    ) -> list[dict]:
        if stype == "free_kick":
            return normalize_free_kick_wall(ball_pos, target, players)
        if stype == "corner":
            return normalize_corner_receivers(ball_pos, ball_vec, target, owner, players)
        if stype == "attack":
            return normalize_attack_receivers(tags, target, owner, players)
        if stype == "throw_in":
            return normalize_throw_in_receivers(target, owner, players)
        return players

    def reference_rows() -> list[dict]:
        """把截图复刻的 29 个参考 preset 纳入正式 preset 配置池。
        ID 通过 REFERENCE_PRESET_ID_REMAP 映射到新分段(1015+/2011+/3010+/4005/5006/6009)。"""
        if not REFERENCE_PRESETS_FILE.exists():
            raise FileNotFoundError(f"缺少参考切片配置源: {REFERENCE_PRESETS_FILE}")
        payload = json.loads(REFERENCE_PRESETS_FILE.read_text(encoding="utf-8"))
        rows: list[dict] = []
        for item in payload["rows"]:
            cfg = dict(item["official_like_row"])
            old_pid = int(cfg["ID"])
            pid = REFERENCE_PRESET_ID_REMAP[old_pid]
            cfg["ID"] = pid
            players = json.loads(cfg["PlayersInit"])
            cfg["BallPos"], players = align_ball_with_owner(
                str(cfg["SliceType"]),
                str(cfg["BallPos"]),
                int(cfg["BallOwner"]),
                players,
            )
            players = normalize_playable_spacing(
                str(cfg["SliceType"]),
                str(cfg["Tags"]),
                str(cfg["BallPos"]),
                str(cfg["BallVector"]),
                cfg.get("TargetPoint"),
                int(cfg["BallOwner"]),
                players,
            )
            ball_z = float(json.loads(cfg["BallPos"])["z"])
            for player_cfg in players:
                if (
                    player_cfg["team"] == "away"
                    and int(player_cfg["duty"]) == PLAYER_AI_DUTY_ENUM["Goalkeeper"]
                ):
                    player_cfg["pos"]["z"] = away_keeper_z(str(cfg["SliceType"]), ball_z)
            cfg["PlayersInit"] = players_init_json(players)
            type_payload = json.loads(cfg["TypePayload"])
            if cfg["SliceType"] == "free_kick":
                type_payload["wall_count"] = sum(
                    1
                    for player_cfg in players
                    if (
                        player_cfg["team"] == "away"
                        and int(player_cfg["duty"]) == PLAYER_AI_DUTY_ENUM["Defender"]
                    )
                )
            cfg["TypePayload"] = json.dumps(type_payload, ensure_ascii=False)
            cfg["NameLcKey"] = lc.add(
                lc_key("preset", "ref", str(old_pid)),
                str(item["name"]),
                f"SlicePresetCfg/{pid}/ref:{item['image']}",
            )
            cfg["Remark"] = f"{cfg['Remark']}；来源=参考截图复刻；已纳入正式 preset 池"
            rows.append(cfg)
        return rows

    preset1_players = [
        player_init("home", 0, PLAYER_AI_DUTY_ENUM["Forward"], 12, 0, -13.0,
                    _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, 12, -13.0)),
        player_init("home", 1, PLAYER_AI_DUTY_ENUM["Forward"], 10, 0, -20,
                    _face_toward(GOAL_CENTER_X, GOAL_CENTER_Z, 10, -20)),
        player_init("away", 0, PLAYER_AI_DUTY_ENUM["Goalkeeper"], 0, 0, away_keeper_z("attack", -13.0),
                    _face_toward(12, -13.0, 0, away_keeper_z("attack", -13.0))),
        player_init("away", 1, PLAYER_AI_DUTY_ENUM["Defender"], 8, 0, -8,
                    _face_toward(12, -13.0, 8, -8)),
    ]

    # (id, slice_type, name, tags, ball_pos, ball_vector, ball_owner, players, fov, target, op_angle, type_payload, rec_modes)
    specs = [
        (1, "attack", "右路单刀", '["side","easy"]', pos_json(12, 0, -13), '{"x":0,"y":0,"z":1}', 0, preset1_players, 45, pos_json(0, 0, 0), 52.0, '{"keeper_weight":5000,"angle":52}', '["draw_line","slingshot"]'),
        (5, "attack", "左路单刀", '["side"]', pos_json(-12, 0, -13), '{"x":0,"y":0,"z":1}', 0, gk_attack(-12, -13), 45, pos_json(0, 0, 0), 52.0, '{"keeper_weight":5000,"angle":52}', '["draw_line","slingshot"]'),
        (6, "attack", "中路突破", '["center"]', pos_json(0, 0, -12), '{"x":0,"y":0,"z":1}', 0, gk_attack(0, -12), 44, pos_json(0, 0, 0), 50.0, '{"keeper_weight":5200,"angle":50}', '["draw_line","slingshot"]'),
        (16, "attack", "中路吊射", '["center","lob"]', pos_json(0, 0, -18), '{"x":0,"y":0,"z":1}', 0, gk_attack(0, -18), 46, pos_json(0, 1.5, 0), 44.0, '{"keeper_weight":5400,"angle":44}', '["draw_line","slingshot"]'),
        (2, "free_kick", "中路任意球", '["center"]', pos_json(0, 0, -16), '{"x":0,"y":0,"z":1}', 0, free_kick_wall(0, -16), 42, pos_json(0, 1.8, 0), 32.0, '{"wall_count":4,"keeper_weight":4500}', '["draw_line","slingshot"]'),
        (7, "free_kick", "左侧任意球", '["side"]', pos_json(-10, 0, -14), '{"x":0,"y":0,"z":1}', 0, free_kick_wall(-10, -14), 42, pos_json(-2, 1.8, 0), 34.0, '{"wall_count":4,"keeper_weight":4500}', '["draw_line","slingshot"]'),
        (8, "free_kick", "右侧任意球", '["side"]', pos_json(10, 0, -14), '{"x":0,"y":0,"z":1}', 0, free_kick_wall(10, -14), 42, pos_json(2, 1.8, 0), 34.0, '{"wall_count":4,"keeper_weight":4500}', '["draw_line","slingshot"]'),
        (17, "free_kick", "弧线任意球", '["center","curve"]', pos_json(4, 0, -17), vec_to(4, -17, -3, 0), 0, free_kick_wall(4, -17), 41, pos_json(-3, 2.0, 0), 30.0, '{"wall_count":4,"keeper_weight":4800}', '["draw_line"]'),
        (3, "penalty", "标准点球", '["penalty"]', penalty_ball_pos(), '{"x":0,"y":0,"z":1}', 0, penalty_players(), 40, pos_json(0, 0.5, 0), 0.0, '{"keeper_dirs":[2500,2500,2500,2500]}', '["draw_line","slingshot"]'),
        (9, "penalty", "加压点球", '["penalty","hard"]', penalty_ball_pos(), '{"x":0,"y":0,"z":1}', 0, penalty_players(), 40, pos_json(0, 0.5, 0), 0.0, '{"keeper_dirs":[2000,3000,3000,2000]}', '["draw_line","slingshot"]'),
        (10, "corner", "左角球", '["corner","left"]', corner_ball_pos("left"), '{"x":1,"y":0,"z":1}', 0, corner_players(CORNER_LEFT_BALL[0]), 48, pos_json(0, 2.0, 0), 38.0, '{"first_point_weight":5000}', '["draw_line"]'),
        (11, "corner", "右角球", '["corner","right"]', corner_ball_pos("right"), '{"x":-1,"y":0,"z":1}', 0, corner_players(CORNER_RIGHT_BALL[0]), 48, pos_json(0, 2.0, 0), 38.0, '{"first_point_weight":5000}', '["draw_line"]'),
        (18, "corner", "后点包抄", '["corner","far"]', corner_ball_pos("right"), '{"x":-1,"y":0,"z":1}', 0, corner_players(CORNER_RIGHT_BALL[0]), 48, pos_json(-6, 2.0, 0), 34.0, '{"first_point_weight":4500,"far_post":1}', '["draw_line"]'),
        (12, "throw_in", "左界外球", '["throw_in","left"]', pos_json(-18, 0, -16), '{"x":1,"y":0,"z":0}', 0, throw_in_players(-20, -16), 44, pos_json(-10, 0, -12), 34.0, '{"second_attack":1}', '["draw_line"]'),
        (13, "throw_in", "右界外球", '["throw_in","right"]', pos_json(18, 0, -16), '{"x":-1,"y":0,"z":0}', 0, throw_in_players(20, -16), 44, pos_json(10, 0, -12), 34.0, '{"second_attack":1}', '["draw_line"]'),
        (4, "goalkeep", "基础守门", '["gk"]', pos_json(0, 0, -18), '{"x":0,"y":0,"z":1}', 0, goalkeep_players(), 50, None, 0.0, '{"shot_dirs":[3000,3000,4000],"reaction_ms":2500}', '["draw_line"]'),
        (14, "goalkeep", "大范围守门", '["gk","wide"]', pos_json(0, 0, -18), '{"x":0,"y":0,"z":1}', 0, goalkeep_players(), 52, None, 0.0, '{"shot_dirs":[3500,3500,3000],"reaction_ms":2200}', '["draw_line"]'),
        (15, "goalkeep", "近距扑点", '["gk","penalty"]', pos_json(0, 0, -11), '{"x":0,"y":0,"z":1}', 0, goalkeep_players(), 50, None, 0.0, '{"shot_dirs":[4000,4000,2000],"reaction_ms":1800}', '["draw_line"]'),
        (19, "attack", "右肋斜插", '["side","diagonal"]', pos_json(8, 0, -15), vec_to(8, -15, -1, 0), 0, attack_setup(8, -15, -1, 5), 45, pos_json(-1, 0.2, 0), 48.0, '{"keeper_weight":5200,"diagonal":1}', '["draw_line","slingshot"]'),
        (20, "attack", "左肋斜插", '["side","diagonal"]', pos_json(-8, 0, -15), vec_to(-8, -15, 1, 0), 0, attack_setup(-8, -15, 1, -5), 45, pos_json(1, 0.2, 0), 48.0, '{"keeper_weight":5200,"diagonal":1}', '["draw_line","slingshot"]'),
        (21, "attack", "禁区弧顶远射", '["center","long_shot"]', pos_json(0, 0, -12), vec_to(0, -12, 0, 0), 0, attack_setup(0, -12, 0, -5), 46, pos_json(0, 1.0, 0), 46.0, '{"keeper_weight":5600,"long_shot":1}', '["draw_line","slingshot"]'),
        (22, "attack", "右路内切射门", '["side","cut_in"]', pos_json(13, 0, -14), vec_to(13, -14, -2, 0), 0, attack_setup(13, -14, -2, 8), 45, pos_json(-2, 0.5, 0), 44.0, '{"keeper_weight":5400,"cut_in":1}', '["draw_line","slingshot"]'),
        (23, "attack", "左路内切射门", '["side","cut_in"]', pos_json(-13, 0, -14), vec_to(-13, -14, 2, 0), 0, attack_setup(-13, -14, 2, -8), 45, pos_json(2, 0.5, 0), 44.0, '{"keeper_weight":5400,"cut_in":1}', '["draw_line","slingshot"]'),
        (24, "attack", "倒三角回做", '["assist","cutback"]', pos_json(14, 0, -5), vec_to(14, -5, 0, -8), 0, attack_setup(14, -5, 0, 2), 47, pos_json(0, 0, -8), 42.0, '{"keeper_weight":5000,"cutback":1}', '["draw_line","slingshot"]'),
        (25, "attack", "门前抢点", '["center","tap_in"]', pos_json(3, 0, -8), vec_to(3, -8, 0, 0), 0, attack_setup(3, -8, 0, -3), 43, pos_json(0, 0.4, 0), 40.0, '{"keeper_weight":6200,"tap_in":1}', '["draw_line","slingshot"]'),
        (26, "attack", "禁区外吊射", '["center","lob","long_shot"]', pos_json(-3, 0, -18), vec_to(-3, -18, 0, 0), 0, attack_setup(-3, -18, 0, 4), 48, pos_json(0, 2.2, 0), 44.0, '{"keeper_weight":5800,"lob":1}', '["draw_line","slingshot"]'),
        (27, "attack", "二点补射", '["center","rebound"]', pos_json(5, 0, -7), vec_to(5, -7, -1, 0), 0, attack_setup(5, -7, -1, 0), 44, pos_json(-1, 0.6, 0), 40.0, '{"keeper_weight":6000,"rebound":1}', '["draw_line","slingshot"]'),
        (28, "attack", "横向摆脱射门", '["center","dribble"]', pos_json(-5, 0, -14), vec_to(-5, -14, 1, 0), 0, attack_setup(-5, -14, 1, -9), 46, pos_json(1, 0.4, 0), 46.0, '{"keeper_weight":5500,"dribble":1}', '["draw_line","slingshot"]'),
        (29, "free_kick", "近距中路任意球", '["center","close"]', pos_json(0, 0, -11), vec_to(0, -11, 0, 0), 0, free_kick_wall(0, -11), 43, pos_json(0, 1.6, 0), 24.0, '{"wall_count":4,"keeper_weight":5200,"close":1}', '["draw_line","slingshot"]'),
        (30, "free_kick", "远距中路任意球", '["center","far"]', pos_json(0, 0, -22), vec_to(0, -22, 0, 0), 0, free_kick_wall(0, -22), 44, pos_json(0, 2.0, 0), 36.0, '{"wall_count":4,"keeper_weight":4300,"far":1}', '["draw_line","slingshot"]'),
        (31, "free_kick", "右侧绕墙低射", '["side","low"]', pos_json(8, 0, -15), vec_to(8, -15, -2, 0), 0, free_kick_wall(8, -15), 42, pos_json(-2, 0.4, 0), 28.0, '{"wall_count":4,"keeper_weight":4800,"low":1}', '["draw_line","slingshot"]'),
        (32, "free_kick", "左侧绕墙低射", '["side","low"]', pos_json(-8, 0, -15), vec_to(-8, -15, 2, 0), 0, free_kick_wall(-8, -15), 42, pos_json(2, 0.4, 0), 28.0, '{"wall_count":4,"keeper_weight":4800,"low":1}', '["draw_line","slingshot"]'),
        (33, "free_kick", "传中任意球", '["side","cross"]', pos_json(13, 0, -19), vec_to(13, -19, -3, 0), 0, free_kick_wall(13, -19), 45, pos_json(-3, 2.2, 0), 34.0, '{"wall_count":4,"keeper_weight":4200,"cross":1}', '["draw_line"]'),
        (34, "penalty", "左下角点球", '["penalty","low_left"]', penalty_ball_pos(), '{"x":-0.35,"y":0,"z":1}', 0, penalty_players(), 40, pos_json(-2.8, 0.2, 0), 0.0, '{"keeper_dirs":[3500,1800,2500,2200]}', '["draw_line","slingshot"]'),
        (35, "penalty", "右下角点球", '["penalty","low_right"]', penalty_ball_pos(), '{"x":0.35,"y":0,"z":1}', 0, penalty_players(), 40, pos_json(2.8, 0.2, 0), 0.0, '{"keeper_dirs":[1800,3500,2500,2200]}', '["draw_line","slingshot"]'),
        (36, "penalty", "半高点球", '["penalty","mid_high"]', penalty_ball_pos(), '{"x":0,"y":0,"z":1}', 0, penalty_players(), 40, pos_json(0, 1.8, 0), 0.0, '{"keeper_dirs":[2400,2400,3200,2000]}', '["draw_line","slingshot"]'),
        (37, "corner", "左前点角球", '["corner","left","near"]', corner_ball_pos("left"), vec_to(-17, -1, -5, 0), 0, corner_setup(CORNER_LEFT_BALL[0], -5, 0), 48, pos_json(-5, 1.7, 0), 36.0, '{"first_point_weight":6200,"near_post":1}', '["draw_line"]'),
        (38, "corner", "右前点角球", '["corner","right","near"]', corner_ball_pos("right"), vec_to(17, -1, 5, 0), 0, corner_setup(CORNER_RIGHT_BALL[0], 5, 0), 48, pos_json(5, 1.7, 0), 36.0, '{"first_point_weight":6200,"near_post":1}', '["draw_line"]'),
        (39, "corner", "左后点高球", '["corner","left","far","high"]', corner_ball_pos("left"), vec_to(-17, -1, 6, 0), 0, corner_setup(CORNER_LEFT_BALL[0], 6, 0), 49, pos_json(6, 2.4, 0), 34.0, '{"first_point_weight":4200,"far_post":1,"high":1}', '["draw_line"]'),
        (40, "corner", "右后点高球", '["corner","right","far","high"]', corner_ball_pos("right"), vec_to(17, -1, -6, 0), 0, corner_setup(CORNER_RIGHT_BALL[0], -6, 0), 49, pos_json(-6, 2.4, 0), 34.0, '{"first_point_weight":4200,"far_post":1,"high":1}', '["draw_line"]'),
        (41, "corner", "短角球配合", '["corner","short"]', corner_ball_pos("left"), vec_to(-17, -1, -12, -6), 0, corner_setup(CORNER_LEFT_BALL[0], -12, -6), 47, pos_json(-12, 0, -6), 40.0, '{"short_corner":1}', '["draw_line"]'),
        (42, "corner", "低平扫门前", '["corner","low_cross"]', corner_ball_pos("right"), vec_to(17, -1, -2, -2), 0, corner_setup(CORNER_RIGHT_BALL[0], -2, -2), 47, pos_json(-2, 0.4, -2), 36.0, '{"low_cross":1}', '["draw_line"]'),
        (43, "corner", "禁区混战二点", '["corner","scramble"]', corner_ball_pos("left"), vec_to(-17, -1, 1, -5), 0, corner_setup(CORNER_LEFT_BALL[0], 1, -5), 50, pos_json(1, 1.0, -5), 38.0, '{"scramble":1}', '["draw_line"]'),
        (44, "throw_in", "左侧近端接应", '["throw_in","left","near"]', pos_json(-18, 0, -10), vec_to(-18, -10, -12, -8), 0, throw_in_setup(-18, -10, -12, -8), 44, pos_json(-12, 0, -8), 34.0, '{"near_support":1}', '["draw_line"]'),
        (45, "throw_in", "右侧近端接应", '["throw_in","right","near"]', pos_json(18, 0, -10), vec_to(18, -10, 12, -8), 0, throw_in_setup(18, -10, 12, -8), 44, pos_json(12, 0, -8), 34.0, '{"near_support":1}', '["draw_line"]'),
        (46, "throw_in", "左侧远端转移", '["throw_in","left","switch"]', pos_json(-18, 0, -23), vec_to(-18, -23, 5, -16), 0, throw_in_setup(-18, -23, 5, -16), 46, pos_json(5, 0, -16), 36.0, '{"switch_side":1}', '["draw_line"]'),
        (47, "throw_in", "右侧远端转移", '["throw_in","right","switch"]', pos_json(18, 0, -23), vec_to(18, -23, -5, -16), 0, throw_in_setup(18, -23, -5, -16), 46, pos_json(-5, 0, -16), 36.0, '{"switch_side":1}', '["draw_line"]'),
        (48, "throw_in", "左路快速反击", '["throw_in","left","counter"]', pos_json(-18, 0, -25), vec_to(-18, -25, -8, -18), 0, throw_in_setup(-18, -25, -8, -18), 45, pos_json(-8, 0, -18), 35.0, '{"counter":1}', '["draw_line"]'),
        (49, "throw_in", "右路快速反击", '["throw_in","right","counter"]', pos_json(18, 0, -25), vec_to(18, -25, 8, -18), 0, throw_in_setup(18, -25, 8, -18), 45, pos_json(8, 0, -18), 35.0, '{"counter":1}', '["draw_line"]'),
        (50, "goalkeep", "左侧低球扑救", '["gk","left","low"]', pos_json(-5, 0, -16), '{"x":0,"y":0,"z":1}', 0, goalkeep_setup(-5, -16), 50, None, 0.0, '{"shot_dirs":[4500,2500,2000],"reaction_ms":2000}', '["draw_line"]'),
    ]

    rows = []
    type_id = {v: k for k, v in SLICE_TYPE_NAME.items()}
    type_desc = {
        "attack": "进攻射门基础站位",
        "free_kick": "任意球+人墙基础站位",
        "penalty": "点球基础站位",
        "corner": "角球传中基础站位",
        "throw_in": "界外球二次进攻基础站位",
        "goalkeep": "守门扑救基础站位",
    }
    for (old_pid, stype, name, tags, ball_pos, ball_vec, owner, players, fov, target, op_angle, payload, rec) in specs:
        pid = MANUAL_PRESET_ID_REMAP[old_pid]
        ball_pos, players = align_ball_with_owner(stype, ball_pos, owner, players)
        players = normalize_playable_spacing(stype, tags, ball_pos, ball_vec, target, owner, players)
        target_desc = target if target else "无固定目标点"
        row = {
            "ID": pid, "SliceType": stype,
            "NameLcKey": lc.add(lc_key("preset", "name", str(old_pid)), name, f"SlicePresetCfg/{pid}"),
            "Tags": tags, "BallPos": ball_pos, "BallVector": ball_vec, "BallOwner": owner,
            "PlayersInit": players_init_json(players), "CameraFov": fov, "TargetPoint": target,
            "OperableAngle": op_angle, "TypePayload": payload, "RecommendedModes": rec,
            "Remark": (
                f"{name}；{type_desc[stype]}；球点={ball_pos}；方向={ball_vec}；"
                f"目标={target_desc}；控球home[{owner}]；推荐操作={rec}"
            ),
        }
        row.update(_preset_angle_cols(type_id[stype], op_angle))
        rows.append(row)
    rows.extend(reference_rows())
    # 按 ID 升序输出：手工 + 参考混合编号后，同 SliceType 段连续(1xxx/2xxx/...)。
    rows.sort(key=lambda r: r["ID"])
    return rows


def build_workbook(lc: LcRegistry) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # --- 3.1 创角 ---
    make_sheet(
        wb,
        "ActvSoccerCharacterCfg",
        c(
            id_col("int", "角色ID character_id"),
            ("AppearanceKey", "string", "外观资源键"),
            ("DisplayPower", "int", "展示战力(仅表现)"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "AppearanceKey": "Assets/k1/K1D1/Res/Models/Soccer/Footballplayer_1.prefab", "DisplayPower": 10, "Remark": "角色A"},
            {"ID": 2, "AppearanceKey": "Assets/k1/K1D1/Res/Models/Soccer/Footballplayer_2.prefab", "DisplayPower": 10, "Remark": "角色B"},
            {"ID": 3, "AppearanceKey": "Assets/k1/K1D1/Res/Models/Soccer/Footballplayer_3.prefab", "DisplayPower": 10, "Remark": "角色C"},
            {"ID": 4, "AppearanceKey": "Assets/k1/K1D1/Res/Models/Soccer/Footballplayer_4.prefab", "DisplayPower": 10, "Remark": "角色D"},
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerNationalityCfg",
        c(
            id_col("int", "国籍ID"),
            ("NameLcKey", "string", "国籍名称→ActvSoccerLanguageCfg"),
            ("Region", "string", "所属地区"),
            ("ContractPool", "int[]", "首签合同池(关联contract_id)"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "NameLcKey": lc.add(lc_key("nationality", "name", "1"), "中国", "NationalityCfg/1"), "Region": "asia", "ContractPool": "[1,101,102]", "Remark": "默认合同1+地区合同101/102"},
            {"ID": 2, "NameLcKey": lc.add(lc_key("nationality", "name", "2"), "德国", "NationalityCfg/2"), "Region": "europe", "ContractPool": "[11,103,104]", "Remark": ""},
            {"ID": 3, "NameLcKey": lc.add(lc_key("nationality", "name", "3"), "巴西", "NationalityCfg/3"), "Region": "south_america", "ContractPool": "[12,105,106]", "Remark": ""},
            {"ID": 4, "NameLcKey": lc.add(lc_key("nationality", "name", "4"), "阿根廷", "NationalityCfg/4"), "Region": "south_america", "ContractPool": "[13,105,107]", "Remark": ""},
            {"ID": 5, "NameLcKey": lc.add(lc_key("nationality", "name", "5"), "法国", "NationalityCfg/5"), "Region": "europe", "ContractPool": "[11,103,108]", "Remark": ""},
            {"ID": 6, "NameLcKey": lc.add(lc_key("nationality", "name", "6"), "西班牙", "NationalityCfg/6"), "Region": "europe", "ContractPool": "[11,103,108]", "Remark": ""},
            {"ID": 7, "NameLcKey": lc.add(lc_key("nationality", "name", "7"), "葡萄牙", "NationalityCfg/7"), "Region": "europe", "ContractPool": "[11,103,108]", "Remark": ""},
            {"ID": 8, "NameLcKey": lc.add(lc_key("nationality", "name", "8"), "比利时", "NationalityCfg/8"), "Region": "europe", "ContractPool": "[11,103,108]", "Remark": ""},
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerTutorialCfg",
        c(
            id_col("int", "试训步骤序号"),
            ("SliceType", "string", "切片类型"),
            ("SliceInstanceID", "int", "关联切片实例(测试用)"),
            ("ForcedOrder", "bool", "强制顺序"),
            ("DescLcKey", "string", "步骤描述→ActvSoccerLanguageCfg"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "SliceType": "attack", "SliceInstanceID": 101, "ForcedOrder": 1, "DescLcKey": lc.add(lc_key("tutorial", "desc", "1"), "进攻教学", "TutorialCfg/1")},
            {"ID": 2, "SliceType": "free_kick", "SliceInstanceID": 102, "ForcedOrder": 1, "DescLcKey": lc.add(lc_key("tutorial", "desc", "2"), "任意球教学", "TutorialCfg/2")},
            {"ID": 3, "SliceType": "penalty", "SliceInstanceID": 103, "ForcedOrder": 1, "DescLcKey": lc.add(lc_key("tutorial", "desc", "3"), "点球教学", "TutorialCfg/3")},
        ],
    )

    build_guide_step_rows(lc)
    copy_sheet_from_xlsx(wb, "ActvSoccerGuideStepCfg", GUIDE_STEP_SOURCE)

    # --- 3.2 切片 ---
    make_sheet(
        wb,
        "ActvSoccerSlicePresetCfg",
        c(
            id_col("int", "preset_id"),
            ("SliceType", "string", "切片类型"),
            ("NameLcKey", "string", "预设名→ActvSoccerLanguageCfg"),
            ("Tags", "string[]", "标签"),
            ("BallPos", "ext", "球位置", P_VEC3, '{"x":0,"y":0,"z":0}'),
            ("BallVector", "ext", "球方向", P_VEC3, '{"x":0,"y":0,"z":1}'),
            ("BallOwner", "int", "控球球员索引"),
            ("PlayersInit", "ext[]", f"球员站位+duty+朝向({PLAYER_AI_DUTY_ENUM_COMMENT})", V_PLAYER_INIT, PLAYER_INIT_DEFAULT),
            ("CameraFov", "float", "相机FOV"),
            ("TargetPoint", "ext", "目标点", P_VEC3, '{"x":0,"y":0,"z":0}'),
            ("OperableAngle", "float", "可操作夹角(兼容字段=AngleSpanMax默认)"),
            ("AngleSpanMin", "float", "可操作夹角宽度下限(°)"),
            ("AngleSpanMax", "float", "可操作夹角宽度上限(°)"),
            ("AngleMaxCenterShift", "float", "扇形中心相对接球方向最大偏移(°)"),
            ("AngleMargin", "float", "合法目标贴边余量(°)"),
            ("TypePayload", "ext", "type_payload默认", V_TYPE_PAYLOAD, '{"keeper_weight":5000,"angle":35}'),
            ("RecommendedModes", "string[]", "建议模式"),
            ("Remark", "string", "设计备注:说明该preset的切片类型、球点/方向、目标点、控球球员和推荐操作；用于策划排查与实例复用"),
        ),
        _build_presets(lc),
    )

    make_sheet(
        wb,
        "ActvSoccerSliceInstanceCfg",
        c(
            id_col("int", "slice_instance_id"),
            ("SliceType", "string", "切片类型"),
            ("PresetID", "int", "preset_id"),
            ("OverrideOperableAngle", "float", "覆盖可操作夹角(0=不覆盖)"),
            ("Modifiers", "ext[]", "切片机制", V_MODIFIER, '[{"id":"moving_keeper","params":{"speed":1.0}}]'),
            ("AiProfileID", "int", "难度档→ActvSoccerAiProfileCfg"),
            ("GoalkeeperAiID", "int", "门将AI"),
            ("DefenderAiID", "int", "防守AI"),
            ("ShooterAiID", "int", "射手AI"),
            ("ModifierID", "int", "机制→ActvSoccerAiModifierCfg"),
            ("IsGuideAi", "bool", "引导关AI"),
            ("RewindRandom", "bool", "回溯后重随机"),
            ("OverrideReactionTimeMs", "int", "覆盖反应时间(0=默认)"),
            ("Remark", "string", "备注"),
        ),
        _build_instance_library(),
    )

    make_sheet(
        wb,
        "ActvSoccerHapticCfg",
        c(
            id_col("int", "编号"),
            ("EventID", "string", "事件"),
            ("Category", "string", "core_operation/key_result"),
            ("Intensity", "string", "light/medium/heavy"),
            ("Pattern", "string", "tick/double/continuous"),
            ("DurationMs", "int", "时长ms"),
            ("MinIntervalMs", "int", "节流ms"),
            ("EnabledDefault", "bool", "默认开启"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "EventID": "charge_locked", "Category": "core_operation", "Intensity": "light", "Pattern": "tick", "DurationMs": 30, "MinIntervalMs": 200, "EnabledDefault": 1},
            {"ID": 2, "EventID": "ball_released", "Category": "core_operation", "Intensity": "medium", "Pattern": "tick", "DurationMs": 40, "MinIntervalMs": 200, "EnabledDefault": 1},
            {"ID": 3, "EventID": "slice_success", "Category": "key_result", "Intensity": "heavy", "Pattern": "double", "DurationMs": 80, "MinIntervalMs": 300, "EnabledDefault": 1, "Remark": "进球"},
            {"ID": 4, "EventID": "saved", "Category": "key_result", "Intensity": "medium", "Pattern": "tick", "DurationMs": 50, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"ID": 5, "EventID": "out_of_bounds", "Category": "key_result", "Intensity": "medium", "Pattern": "tick", "DurationMs": 50, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"ID": 6, "EventID": "gk_timeout", "Category": "key_result", "Intensity": "heavy", "Pattern": "tick", "DurationMs": 60, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"ID": 7, "EventID": "gk_wrong_judge", "Category": "key_result", "Intensity": "medium", "Pattern": "double", "DurationMs": 70, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"ID": 8, "EventID": "gk_save", "Category": "key_result", "Intensity": "heavy", "Pattern": "double", "DurationMs": 80, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"ID": 9, "EventID": "hit_post", "Category": "key_result", "Intensity": "medium", "Pattern": "double", "DurationMs": 60, "MinIntervalMs": 300, "EnabledDefault": 1},
            {"ID": 10, "EventID": "rewind_reset", "Category": "key_result", "Intensity": "light", "Pattern": "continuous", "DurationMs": 120, "MinIntervalMs": 500, "EnabledDefault": 1},
        ],
    )

    # --- 接球决策 (2026-06-15 first-touch design) ---
    # ReceiveDecision 5001-5004 / PlayerStyle 5101-5104 / FirstTouchAnim 5201+
    make_sheet(
        wb,
        "ActvSoccerReceiveDecisionCfg",
        c(
            id_col("int", "接球决策配置ID"),
            ("Style", "string", "默认球员风格", "", "Balanced/Playmaker/Dribbler/TargetMan"),
            ("DecisionMinMs", "int", "最早决策时间;预计触球前多少ms开始允许生成决策"),
            ("DecisionMaxMs", "int", "最晚决策时间;预计触球前多少ms内必须已有决策"),
            ("SafeDistance", "float", "安全距离;最近防守人大于该距离视为低压(m)"),
            ("HighPressureDistance", "float", "高压距离;最近防守人小于该距离视为高压(m)"),
            ("ForwardProbeDistance", "float", "前方空间探测距离(m)"),
            ("SideProbeDistance", "float", "左右空间探测距离(m)"),
            ("BackwardProbeDistance", "float", "身后空间探测距离(m)"),
            ("HighBallHeight", "float", "高空球高度阈值;超过后表现层优先胸停/头球(m)"),
            ("FastBallSpeed", "float", "高速来球速度阈值;超过后提高停球权重(m/s)"),
            ("StopWeight", "int", "停球权重(100=标准)"),
            ("PushForwardWeight", "int", "顺势向前领球权重(100=标准)"),
            ("PushSideWeight", "int", "左右领球权重(100=标准)"),
            ("HalfTurnWeight", "int", "半转身权重(100=标准)"),
            ("ShieldWeight", "int", "护球权重(100=标准)"),
            ("OneTouchPassWeight", "int", "一脚传球权重;玩家可操作接球点不自动出球(100=标准)"),
            ("OneTouchShotWeight", "int", "一脚射门权重;玩家可操作接球点不自动出球(100=标准)"),
            ("SpaceGainWeight", "int", "评分项:空间收益权重(100=标准)"),
            ("GoalProgressWeight", "int", "评分项:向球门推进权重(100=标准)"),
            ("SafetyWeight", "int", "评分项:安全性权重(100=标准)"),
            ("FlowWeight", "int", "评分项:是否利于下一步动作(100=标准)"),
            ("StyleBonusWeight", "int", "评分项:球员风格加成(100=标准)"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 5001, "Style": "Balanced", "DecisionMinMs": 300, "DecisionMaxMs": 1000, "SafeDistance": 3.5, "HighPressureDistance": 1.6, "ForwardProbeDistance": 5.0, "SideProbeDistance": 3.5, "BackwardProbeDistance": 2.5, "HighBallHeight": 0.9, "FastBallSpeed": 16.0, "StopWeight": 100, "PushForwardWeight": 100, "PushSideWeight": 100, "HalfTurnWeight": 100, "ShieldWeight": 100, "OneTouchPassWeight": 100, "OneTouchShotWeight": 100, "SpaceGainWeight": 100, "GoalProgressWeight": 100, "SafetyWeight": 100, "FlowWeight": 100, "StyleBonusWeight": 100, "Remark": "通用默认"},
            {"ID": 5002, "Style": "Playmaker", "DecisionMinMs": 300, "DecisionMaxMs": 1000, "SafeDistance": 3.5, "HighPressureDistance": 1.7, "ForwardProbeDistance": 4.5, "SideProbeDistance": 3.5, "BackwardProbeDistance": 2.5, "HighBallHeight": 0.9, "FastBallSpeed": 16.0, "StopWeight": 90, "PushForwardWeight": 90, "PushSideWeight": 90, "HalfTurnWeight": 100, "ShieldWeight": 80, "OneTouchPassWeight": 140, "OneTouchShotWeight": 100, "SpaceGainWeight": 90, "GoalProgressWeight": 110, "SafetyWeight": 100, "FlowWeight": 130, "StyleBonusWeight": 130, "Remark": "组织核心,倾向一脚传球"},
            {"ID": 5003, "Style": "Dribbler", "DecisionMinMs": 300, "DecisionMaxMs": 1000, "SafeDistance": 3.2, "HighPressureDistance": 1.5, "ForwardProbeDistance": 5.5, "SideProbeDistance": 4.0, "BackwardProbeDistance": 2.5, "HighBallHeight": 0.9, "FastBallSpeed": 16.0, "StopWeight": 80, "PushForwardWeight": 140, "PushSideWeight": 130, "HalfTurnWeight": 120, "ShieldWeight": 90, "OneTouchPassWeight": 80, "OneTouchShotWeight": 110, "SpaceGainWeight": 130, "GoalProgressWeight": 120, "SafetyWeight": 90, "FlowWeight": 120, "StyleBonusWeight": 130, "Remark": "突破手,倾向顺势领球"},
            {"ID": 5004, "Style": "TargetMan", "DecisionMinMs": 300, "DecisionMaxMs": 1000, "SafeDistance": 3.0, "HighPressureDistance": 1.4, "ForwardProbeDistance": 4.0, "SideProbeDistance": 3.0, "BackwardProbeDistance": 3.0, "HighBallHeight": 0.9, "FastBallSpeed": 15.0, "StopWeight": 120, "PushForwardWeight": 80, "PushSideWeight": 80, "HalfTurnWeight": 110, "ShieldWeight": 150, "OneTouchPassWeight": 100, "OneTouchShotWeight": 120, "SpaceGainWeight": 80, "GoalProgressWeight": 100, "SafetyWeight": 140, "FlowWeight": 110, "StyleBonusWeight": 130, "Remark": "支点,倾向护球和背身处理"},
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerPlayerStyleCfg",
        c(
            id_col("int", "球员风格配置ID"),
            ("Style", "string", "风格枚举", "", "Balanced/Playmaker/Dribbler/TargetMan"),
            ("ReceiveDecisionCfgID", "int", "接球决策配置→ActvSoccerReceiveDecisionCfg"),
            ("PassBias", "int", "传球倾向(100=标准)"),
            ("DribbleBias", "int", "盘带/领球倾向(100=标准)"),
            ("ShieldBias", "int", "护球倾向(100=标准)"),
            ("ShotBias", "int", "射门倾向(100=标准)"),
            ("RiskBias", "int", "风险倾向;越高越敢做高收益低安全动作(100=标准)"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 5101, "Style": "Balanced", "ReceiveDecisionCfgID": 5001, "PassBias": 100, "DribbleBias": 100, "ShieldBias": 100, "ShotBias": 100, "RiskBias": 100, "Remark": "通用;DEF兜底"},
            {"ID": 5102, "Style": "Playmaker", "ReceiveDecisionCfgID": 5002, "PassBias": 140, "DribbleBias": 90, "ShieldBias": 80, "ShotBias": 100, "RiskBias": 105, "Remark": "德布劳内/莫德里奇类;MID兜底"},
            {"ID": 5103, "Style": "Dribbler", "ReceiveDecisionCfgID": 5003, "PassBias": 80, "DribbleBias": 140, "ShieldBias": 90, "ShotBias": 110, "RiskBias": 120, "Remark": "梅西/姆巴佩类;FWD兜底"},
            {"ID": 5104, "Style": "TargetMan", "ReceiveDecisionCfgID": 5004, "PassBias": 100, "DribbleBias": 80, "ShieldBias": 150, "ShotBias": 120, "RiskBias": 90, "Remark": "哈兰德/凯恩类"},
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerFirstTouchAnimCfg",
        c(
            id_col("int", "表现配置ID"),
            ("Action", "string", "第一脚逻辑动作", "", "Stop/PushForward/PushLeft/PushRight/HalfTurnLeft/HalfTurnRight/Shield/OneTouchPass/OneTouchShot"),
            ("BallHeightType", "string", "来球高度分类", "", "Ground/High"),
            ("BallDirectionType", "string", "来球方向分类", "", "Front/Left/Right/Back"),
            ("BodyDirectionType", "string", "身体朝向分类", "", "FaceBall/BackToBall/SideToBall"),
            ("PressureLevel", "string", "压力等级", "", "Low/Medium/High/Any"),
            ("StateKey", "string", "对应ActvSoccerCharacterStateCfg.StateKey"),
            ("AnimKey", "string", "美术动作Key;优先匹配已有动作表"),
            ("Priority", "int", "多条命中时取优先级高的"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 5201, "Action": "Stop", "BallHeightType": "Ground", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Any", "StateKey": "Control", "AnimKey": "B01_ReceiveBall", "Priority": 100, "Remark": "地面正面停球"},
            {"ID": 5202, "Action": "Stop", "BallHeightType": "High", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Any", "StateKey": "Control", "AnimKey": "B03_ReceiveAir", "Priority": 100, "Remark": "高空球停球,可复用B01"},
            {"ID": 5203, "Action": "PushForward", "BallHeightType": "Ground", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Low", "StateKey": "Control", "AnimKey": "B02_DribbleTouch", "Priority": 110, "Remark": "顺势领球推进"},
            {"ID": 5204, "Action": "PushLeft", "BallHeightType": "Ground", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Low", "StateKey": "Control", "AnimKey": "B02_DribbleTouch", "Priority": 100, "Remark": "左侧领球,表现可镜像"},
            {"ID": 5205, "Action": "PushRight", "BallHeightType": "Ground", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Low", "StateKey": "Control", "AnimKey": "B02_DribbleTouch", "Priority": 100, "Remark": "右侧领球,表现可镜像"},
            {"ID": 5206, "Action": "HalfTurnLeft", "BallHeightType": "Ground", "BallDirectionType": "Back", "BodyDirectionType": "BackToBall", "PressureLevel": "Medium", "StateKey": "TurnLeft", "AnimKey": "A05_TurnLeft", "Priority": 100, "Remark": "背身半转身"},
            {"ID": 5207, "Action": "HalfTurnRight", "BallHeightType": "Ground", "BallDirectionType": "Back", "BodyDirectionType": "BackToBall", "PressureLevel": "Medium", "StateKey": "TurnRight", "AnimKey": "A06_TurnRight", "Priority": 100, "Remark": "背身半转身镜像"},
            {"ID": 5208, "Action": "Shield", "BallHeightType": "Ground", "BallDirectionType": "Back", "BodyDirectionType": "BackToBall", "PressureLevel": "High", "StateKey": "Control", "AnimKey": "B01_ReceiveBall", "Priority": 120, "Remark": "高压背身护球,后续可替换专用护球动作"},
            {"ID": 5209, "Action": "OneTouchPass", "BallHeightType": "Ground", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Low", "StateKey": "Pass", "AnimKey": "C03_OneTouchPass", "Priority": 120, "Remark": "一脚传球"},
            {"ID": 5210, "Action": "OneTouchShot", "BallHeightType": "Ground", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Low", "StateKey": "Kick", "AnimKey": "D01_Shoot", "Priority": 100, "Remark": "一脚射门,可复用射门动作"},
            {"ID": 5211, "Action": "OneTouchShot", "BallHeightType": "High", "BallDirectionType": "Front", "BodyDirectionType": "FaceBall", "PressureLevel": "Any", "StateKey": "Kick", "AnimKey": "D05_Volley", "Priority": 110, "Remark": "高空球一脚处理,后续可扩展头球/倒挂金钩"},
        ],
    )

    # --- 3.3 关卡 ---
    make_sheet(
        wb,
        "ActvSoccerLevelCfg",
        c(
            id_col("int", "level_id"),
            ("IsTutorial", "bool", "引导关"),
            ("SliceList", "int[]", "切片实例序列"),
            ("AiProfileID", "int", "AI档位"),
            ("WinThreshold", "int", "胜利阈值"),
            ("DrawThreshold", "int", "平局阈值"),
            ("TicketCost", "int", "门票消耗"),
            ("OpponentTeamID", "int", "对手球队(队名/队服/队标)"),
            ("OpponentTeamStar", "int", "对手球队星级(球员属性计算依据)"),
            ("SeasonID", "int", "所属联赛轮次(=SeasonCfg.ID)"),
            ("Remark", "string", "备注"),
        ),
        _build_levels(lc),
    )

    # --- FSM/BT/敌人AI (2026-06-09 design) ---
    make_sheet(
        wb,
        "ActvSoccerAiProfileCfg",
        c(
            id_col("int", "AI难度档ID"),
            ("Difficulty", "string", "easy/normal/hard"),
            ("GoalkeeperSaveRate", "int", "门将扑救成功率%"),
            ("DefenderSuccessRate", "int", "防守成功率%"),
            ("ShooterSuccessRate", "int", "对手射门成功率%"),
            ("DeadCornerCanSave", "int", "死角可扑(固定0)"),
            ("ReactionTimeMs", "int", "默认反应时间ms"),
            ("Remark", "string", "备注"),
        ),
        _build_ai_profiles(),
    )

    make_sheet(
        wb,
        "ActvSoccerEnemyAiCfg",
        c(
            id_col("int", "敌人AI配置ID"),
            ("Duty", "int", "球员AI职责(程序枚举PlayerAiDuty)", "", PLAYER_AI_DUTY_ENUM_COMMENT),
            ("SaveWeight", "int", "扑救权重"),
            ("LeftWeight", "int", "左方向权重"),
            ("RightWeight", "int", "右方向权重"),
            ("UpWeight", "int", "上方向权重"),
            ("InterceptWeight", "int", "拦截权重"),
            ("ClearanceWeight", "int", "解围出界权重"),
            ("KeeperCatchFail", "int", "门将接球直接失败"),
            ("OutOfBoundsFail", "int", "踢出界外失败"),
            ("AnimationKey", "string", "默认表现动作"),
            ("Remark", "string", "备注"),
        ),
        _build_enemy_ai(),
    )

    make_sheet(
        wb,
        "ActvSoccerAiModifierCfg",
        c(
            id_col("int", "机制ID"),
            ("ModifierType", "string", "机制类型"),
            ("Param1Key", "string", "参数1键"),
            ("Param1Value", "string", "参数1值"),
            ("Param2Key", "string", "参数2键"),
            ("Param2Value", "string", "参数2值"),
            ("Param3Key", "string", "参数3键"),
            ("Param3Value", "string", "参数3值"),
            ("Remark", "string", "备注"),
        ),
        _build_ai_modifiers(),
    )

    make_sheet(
        wb,
        "ActvSoccerSliceFlowCfg",
        c(
            id_col("int", "切片流程ID"),
            ("SliceType", "string", "切片类型"),
            ("ForceOperationMode", "string", "强制操作模式"),
            ("WaitInputTimeMs", "int", "等待限时(0=无限)"),
            ("NeedReplayClick", "bool", "回放需手动结束"),
            ("EnableRewind", "bool", "允许回溯"),
            ("SuccessReplayKey", "string", "成功回放key"),
            ("FailReplayKey", "string", "失败回放key"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 5001, "SliceType": "attack", "ForceOperationMode": "none", "WaitInputTimeMs": 0, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "Celebrate", "FailReplayKey": "Fail", "Remark": "进攻切片"},
            {"ID": 5002, "SliceType": "free_kick", "ForceOperationMode": "none", "WaitInputTimeMs": 0, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "Celebrate", "FailReplayKey": "Fail", "Remark": "任意球"},
            {"ID": 5003, "SliceType": "penalty", "ForceOperationMode": "none", "WaitInputTimeMs": 0, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "Celebrate", "FailReplayKey": "Fail", "Remark": "点球"},
            {"ID": 5004, "SliceType": "goalkeep", "ForceOperationMode": "draw_line", "WaitInputTimeMs": 900, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "SaveSuccess", "FailReplayKey": "TimeoutFeedback", "Remark": "守门强制划线"},
            {"ID": 5005, "SliceType": "tutorial_attack", "ForceOperationMode": "none", "WaitInputTimeMs": 0, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "Celebrate", "FailReplayKey": "Fail", "Remark": "引导进攻(试训模式可切换)"},
            {"ID": 5006, "SliceType": "corner", "ForceOperationMode": "none", "WaitInputTimeMs": 0, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "Celebrate", "FailReplayKey": "Fail", "Remark": "角球P1"},
            {"ID": 5007, "SliceType": "throw_in", "ForceOperationMode": "none", "WaitInputTimeMs": 0, "NeedReplayClick": 1, "EnableRewind": 1, "SuccessReplayKey": "Celebrate", "FailReplayKey": "Fail", "Remark": "界外球P1"},
        ],
    )

    # --- 3.4 养成 ---
    make_sheet(
        wb,
        "ActvSoccerFameGrowthLevelCfg",
        c(
            id_col("int", "编号"),
            ("Level", "int", "知名度等级"),
            ("ExpRequired", "int", "升级所需知名度经验"),
            ("ContractStarLicReward", "int", "合同星级许可奖励(0=无,2-5=license_id)"),
            ("PlayerRating", "int", "主角评分(仅知名度线投放)"),
            ("TitleLcKey", "string", "称号→ActvSoccerLanguageCfg"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "Level": 1, "ExpRequired": 0, "ContractStarLicReward": 0, "PlayerRating": 10, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "1"), "业余球员", "FameGrowthLevelCfg/1")},
            {"ID": 2, "Level": 2, "ExpRequired": 100, "ContractStarLicReward": 0, "PlayerRating": 15, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "2"), "新秀", "FameGrowthLevelCfg/2"), "Remark": ""},
            {"ID": 3, "Level": 3, "ExpRequired": 150, "ContractStarLicReward": 2, "PlayerRating": 20, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "3"), "专业球员", "FameGrowthLevelCfg/3"), "Remark": "3级→2星许可"},
            {"ID": 4, "Level": 4, "ExpRequired": 180, "ContractStarLicReward": 2, "PlayerRating": 25, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "4"), "明星", "FameGrowthLevelCfg/4"), "Remark": ""},
            {"ID": 5, "Level": 5, "ExpRequired": 200, "ContractStarLicReward": 3, "PlayerRating": 30, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "5"), "大师", "FameGrowthLevelCfg/5"), "Remark": "文档示例:知名度5级→3星许可"},
            {"ID": 6, "Level": 6, "ExpRequired": 240, "ContractStarLicReward": 4, "PlayerRating": 35, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "6"), "世界级", "FameGrowthLevelCfg/6"), "Remark": "6级→4星许可"},
            {"ID": 7, "Level": 7, "ExpRequired": 300, "ContractStarLicReward": 5, "PlayerRating": 40, "TitleLcKey": lc.add(lc_key("growth", "fame_title", "7"), "传奇", "FameGrowthLevelCfg/7"), "Remark": "7级→5星许可(tier9-10合同池可达)"},
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerLifeGrowthLevelCfg",
        c(
            id_col("int", "编号"),
            ("Level", "int", "生活等级"),
            ("ExpRequired", "int", "升级消耗金币"),
            ("RewardFame", "int", "升级奖励知名度(0=无)"),
            ("ContractStarLicReward", "int", "合同星级许可奖励(0=无,2-5=license_id)"),
            ("TicketCap", "int", "门票上限"),
            ("TicketRecoverMin", "int", "门票恢复间隔分钟"),
            ("FreeRewind", "int", "免费回溯次数"),
            ("ExtraRound", "int", "额外联赛轮次"),
            ("QualityShow", "int", "品质等级展示(配合UI色块)"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "Level": 1, "ExpRequired": 0, "ContractStarLicReward": 0, "TicketCap": 80, "TicketRecoverMin": 30, "QualityShow": 1, "Remark": ""},
            {"ID": 2, "Level": 2, "ExpRequired": 100, "ContractStarLicReward": 0, "TicketCap": 200, "TicketRecoverMin": 28, "FreeRewind": 1, "QualityShow": 1, "Remark": "升级消耗金币100"},
            {"ID": 3, "Level": 3, "ExpRequired": 150, "ContractStarLicReward": 2, "TicketCap": 220, "TicketRecoverMin": 26, "QualityShow": 2, "Remark": "3级→2星许可"},
            {"ID": 4, "Level": 4, "ExpRequired": 200, "RewardFame": 10, "ContractStarLicReward": 3, "TicketCap": 250, "TicketRecoverMin": 25, "ExtraRound": 5, "QualityShow": 2, "Remark": "文档示例:生活4级→3星许可"},
            {"ID": 5, "Level": 5, "ExpRequired": 300, "ContractStarLicReward": 4, "TicketCap": 280, "TicketRecoverMin": 22, "QualityShow": 3, "Remark": "5级→4星许可"},
            {"ID": 6, "Level": 6, "ExpRequired": 400, "RewardFame": 15, "ContractStarLicReward": 4, "TicketCap": 300, "TicketRecoverMin": 20, "FreeRewind": 2, "QualityShow": 3, "Remark": "6级→4星许可(维持)"},
            {"ID": 7, "Level": 7, "ExpRequired": 550, "RewardFame": 20, "ContractStarLicReward": 5, "TicketCap": 320, "TicketRecoverMin": 18, "FreeRewind": 2, "ExtraRound": 10, "QualityShow": 4, "Remark": "7级→5星许可(tier9-10合同池可达)"},
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerTeamCfg",
        c(
            id_col("int", "team_id"),
            ("NameLcKey", "string", "球队名→ActvSoccerLanguageCfg"),
            ("Region", "string", "地区(纯展示)"),
            ("KitKey", "string", "队服资源"),
            ("BadgeKey", "string", "队标资源"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 101, "NameLcKey": lc.add(lc_key("team", "name", "101"), "红星联", "TeamCfg/101"), "Region": "asia", "KitKey": "WC_Kit_01", "BadgeKey": "WC_Badge_01", "Remark": "默认展示球队"},
            {"ID": 102, "NameLcKey": lc.add(lc_key("team", "name", "102"), "莱茵青年", "TeamCfg/102"), "Region": "europe", "KitKey": "WC_Kit_02", "BadgeKey": "WC_Badge_02", "Remark": ""},
            {"ID": 103, "NameLcKey": lc.add(lc_key("team", "name", "103"), "桑巴之星", "TeamCfg/103"), "Region": "south_america", "KitKey": "WC_Kit_03", "BadgeKey": "WC_Badge_03", "Remark": ""},
            {"ID": 104, "NameLcKey": lc.add(lc_key("team", "name", "104"), "蓝白雄鹰", "TeamCfg/104"), "Region": "south_america", "KitKey": "WC_Kit_04", "BadgeKey": "WC_Badge_04", "Remark": ""},
            {"ID": 201, "NameLcKey": lc.add(lc_key("team", "name", "201"), "海港FC", "TeamCfg/201"), "Region": "asia", "KitKey": "WC_Kit_05", "BadgeKey": "WC_Badge_05", "Remark": "中国动态池"},
            {"ID": 202, "NameLcKey": lc.add(lc_key("team", "name", "202"), "东方之鹰", "TeamCfg/202"), "Region": "asia", "KitKey": "WC_Kit_06", "BadgeKey": "WC_Badge_06", "Remark": "中国动态池"},
            {"ID": 203, "NameLcKey": lc.add(lc_key("team", "name", "203"), "北欧狼", "TeamCfg/203"), "Region": "europe", "KitKey": "WC_Kit_07", "BadgeKey": "WC_Badge_07", "Remark": "联赛对手展示"},
            {"ID": 204, "NameLcKey": lc.add(lc_key("team", "name", "204"), "阿根廷神鹰", "TeamCfg/204"), "Region": "south_america", "KitKey": "WC_Kit_08", "BadgeKey": "WC_Badge_08", "Remark": "文档合同示例"},
            {"ID": 205, "NameLcKey": lc.add(lc_key("team", "name", "205"), "桑巴红魔", "TeamCfg/205"), "Region": "south_america", "KitKey": "WC_Kit_09", "BadgeKey": "WC_Badge_09", "Remark": ""},
            {"ID": 206, "NameLcKey": lc.add(lc_key("team", "name", "206"), "南美风暴", "TeamCfg/206"), "Region": "south_america", "KitKey": "WC_Kit_10", "BadgeKey": "WC_Badge_10", "Remark": ""},
            {"ID": 207, "NameLcKey": lc.add(lc_key("team", "name", "207"), "潘帕斯之翼", "TeamCfg/207"), "Region": "south_america", "KitKey": "WC_Kit_11", "BadgeKey": "WC_Badge_11", "Remark": ""},
            {"ID": 208, "NameLcKey": lc.add(lc_key("team", "name", "208"), "高卢雄鸡", "TeamCfg/208"), "Region": "europe", "KitKey": "WC_Kit_12", "BadgeKey": "WC_Badge_12", "Remark": "12队服/队标"},
            *_build_theme_teams(lc),
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerContractStarLicCfg",
        c(
            id_col("int", "license_id(=最高可出星级)"),
            ("AllowedTeamStars", "int[]", "允许出现的合同星级列表"),
            ("StarWeights", "int[]", "与AllowedTeamStars等长、按下标对应权重"),
            ("Remark", "string", "备注"),
        ),
        _contract_star_license_rows(),
    )

    make_sheet(
        wb,
        "ActvSoccerContractCfg",
        c(
            id_col("int", "contract_id"),
            ("TeamID", "int", "关联展示球队→TeamCfg"),
            ("TeamStar", "int", "合同星级(唯一生效星级)"),
            ("PayFinish", "int", "周薪/完赛待遇"),
            ("PayGoal", "int", "进球待遇"),
            ("PayAssist", "int", "助攻待遇"),
            ("PayFame", "int", "名气待遇"),
            ("SeasonGoal", "ext[]", "赛季目标", V_SEASON_GOAL, '[{"type":"rank","threshold":12,"settle_at":"season_end"}]'),
            ("SeasonReward", "ext[]", "目标奖励", P_TYIDVAL, '[{"typ":"vm","id":11151001,"val":84}]'),
            ("SignReward", "ext[]", "签约即时奖励", P_TYIDVAL, '[]'),
            ("GrantFameLevel", "int", "抽样门槛-知名度等级"),
            ("GrantLifeLevel", "int", "抽样门槛-生活等级"),
            ("GrantScene", "string", "first_sign/league_finish"),
            ("Remark", "string", "备注"),
        ),
        [
            *_first_sign_contract_rows(),
            *_league_finish_contract_rows(),
        ],
    )

    # --- 3.5 积分赛 ---
    make_sheet(
        wb,
        "ActvSoccerSeasonCfg",
        c(
            id_col("int", "season_id"),
            ("LeagueNameLcKey", "string", "联赛名称→ActvSoccerLanguageCfg"),
            ("NextSeason", "int", "后置联赛(0=系列结束)"),
            ("ContractOfferCount", "int", "换约候选份数(0=读全局常量)"),
            ("Remark", "string", "备注"),
        ),
        _build_seasons(lc),
    )

    # --- 3.6 淘汰赛 ---
    make_sheet(
        wb,
        "ActvSoccerKnockoutCfg",
        c(
            id_col("int", "编号"),
            ("OpenLeagueLevel", "int", "开放所需积分赛关卡L"),
            ("TeamSizeMax", "int", "队伍人数上限M"),
            ("QualifierGroupCount", "int", "海选分组数G"),
            ("QualifyPerGroup", "int", "每组晋级K"),
            ("KnockoutGroupCount", "int", "单淘汰分组数(8组)"),
            ("MaxRounds", "int", "单淘汰轮次R"),
            ("BotPlayerRating", "int", "补位球员评分"),
            ("ReadyTime", "string", "每日开赛前准备时长(支持h/m后缀)"),
            ("BetTime", "string", "每日竞猜窗口时长"),
            ("MatchTime", "string", "每日比赛时长"),
            ("Remark", "string", "备注"),
        ),
        [
            {
                "ID": 1,
                "OpenLeagueLevel": KNOCKOUT_OPEN_LEVEL,
                "TeamSizeMax": 5,
                "QualifierGroupCount": 16,
                "QualifyPerGroup": 4,
                "KnockoutGroupCount": 8,
                "MaxRounds": 6,
                "BotPlayerRating": 10,
                "ReadyTime": "30m",
                "BetTime": "23h",
                "MatchTime": "30m",
                "Remark": "最终文档:64强分8组,海选16组取前4",
            },
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerKnockoutPhaseCfg",
        c(
            id_col("int", "编号"),
            ("PhaseLcKey", "string", "阶段名→ActvSoccerLanguageCfg"),
            ("PhaseKey", "string", "阶段键"),
            ("StartTime", "string", "开始UTC"),
            ("EndTime", "string", "结束UTC"),
            ("DayContentLcKey", "string", "当日内容→ActvSoccerLanguageCfg"),
            ("BetOpen", "bool", "开放竞猜"),
            ("Remark", "string", "备注"),
        ),
        [
            {"ID": 1, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "team_build"), "组队期", "KnockoutPhaseCfg/1"), "PhaseKey": "team_build", "StartTime": "2026-07-10 00:00:00", "EndTime": "2026-07-11 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "1"), "发起/加入组队、命名队名队标、审批", "KnockoutPhaseCfg/1"), "BetOpen": 0},
            {"ID": 2, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "3"), "海选", "KnockoutPhaseCfg/2"), "PhaseKey": "qualifier", "StartTime": "2026-07-12 00:00:00", "EndTime": "2026-07-12 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "3"), "1天内筛出64强", "KnockoutPhaseCfg/2"), "BetOpen": 0, "Remark": "不开竞猜"},
            {"ID": 3, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "4"), "正赛第1轮", "KnockoutPhaseCfg/3"), "PhaseKey": "knockout_r1", "StartTime": "2026-07-13 00:00:00", "EndTime": "2026-07-13 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "4"), "8组各一轮(64强)", "KnockoutPhaseCfg/3"), "BetOpen": 1},
            {"ID": 4, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "5"), "正赛第2轮", "KnockoutPhaseCfg/4"), "PhaseKey": "knockout_r2", "StartTime": "2026-07-14 00:00:00", "EndTime": "2026-07-14 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "5"), "8组各一轮(32强)", "KnockoutPhaseCfg/4"), "BetOpen": 1},
            {"ID": 5, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "6"), "正赛第3轮", "KnockoutPhaseCfg/5"), "PhaseKey": "knockout_r3", "StartTime": "2026-07-15 00:00:00", "EndTime": "2026-07-15 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "6"), "8组各一轮(16强)", "KnockoutPhaseCfg/5"), "BetOpen": 1},
            {"ID": 6, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "7"), "正赛第4轮", "KnockoutPhaseCfg/6"), "PhaseKey": "knockout_r4", "StartTime": "2026-07-16 00:00:00", "EndTime": "2026-07-16 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "7"), "8组决出组冠军→8强", "KnockoutPhaseCfg/6"), "BetOpen": 1},
            {"ID": 7, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "8"), "正赛第5轮", "KnockoutPhaseCfg/7"), "PhaseKey": "knockout_r5", "StartTime": "2026-07-17 00:00:00", "EndTime": "2026-07-17 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "8"), "8强淘汰赛(4强)", "KnockoutPhaseCfg/7"), "BetOpen": 1},
            {"ID": 8, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "9"), "决赛", "KnockoutPhaseCfg/8"), "PhaseKey": "knockout_final", "StartTime": "2026-07-18 00:00:00", "EndTime": "2026-07-18 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "9"), "冠亚军决赛", "KnockoutPhaseCfg/8"), "BetOpen": 1},
            {"ID": 9, "PhaseLcKey": lc.add(lc_key("knockout", "phase_name", "10"), "展示期", "KnockoutPhaseCfg/9"), "PhaseKey": "showcase", "StartTime": "2026-07-19 00:00:00", "EndTime": "2026-07-19 23:59:59", "DayContentLcKey": lc.add(lc_key("knockout", "day_content", "10"), "冠军展示+兑换商店可用", "KnockoutPhaseCfg/9"), "BetOpen": 0},
        ],
    )

    # --- 3.7 竞猜赔率与投注档位 ---
    make_sheet(
        wb,
        "ActvSoccerBetMultiplierCfg",
        c(
            id_col("int", "编号"),
            ("winRate_int", "int", "胜率万分值(主队;0..10000)"),
            ("powerRate_int", "int", "战力比万分值(主队/客队;3000..20000)"),
            ("oddsLeft_int", "int", "主队赔率万分值"),
            ("oddsRight_int", "int", "客队赔率万分值"),
        ),
        [
            {
                "ID": i + 1,
                "winRate_int": wr,
                "powerRate_int": pr,
                "oddsLeft_int": odds_l,
                "oddsRight_int": odds_r,
            }
            for i, (wr, pr, odds_l, odds_r) in enumerate(BET_MULTIPLIER_GRID)
        ],
    )

    make_sheet(
        wb,
        "ActvSoccerBetStakeTierCfg",
        c(
            id_col("int", "编号"),
            ("TierKey", "string", "档位键(free/stake_N)"),
            ("Stake", "int", "投注竞猜币(免费=0)"),
            ("HitPayout", "int", "命中固定派彩(0=按倍率)"),
            ("Sort", "int", "排序"),
            ("Remark", "string", "备注"),
        ),
        BET_STAKE_TIER_ROWS,
    )

    return align_to_latest_activity_soccer_schema(wb)


def export_summary(sheets: list[str], lc_rows: list[dict], const_rows: list[dict]) -> None:
    summary = {
        "file": "ActivitySoccer_preview.xlsx",
        "language_file": "ActivitySoccerLanguage.xlsx",
        "sources": SOURCE_DOCS,
        "sheets": sheets,
        "language_sheet": "ActvSoccerLanguageCfg",
        "language_entry_count": len(lc_rows),
        "sheets_program_only": list(SHEETS_PROGRAM_ONLY),
        "sheets_common_platform": {
            "ActvSoccerBattlePassCfg": "BattlePassNew / ActivityBattlePass",
            "ActvSoccerExchangeShopCfg": "ExchangeShopItemCfg",
            "ActvSoccerGiftCfg": "GiftCfg / D2GiftCfg",
            "ActvSoccerRankSectionCfg": "ActivityRank / ActvRankSectionCfg",
        },
        "appendix_only": list(APPENDIX_ONLY),
        "appendix_in_match_items": IN_MATCH_ITEMS,
        "appendix_const_rows": const_rows,
        "bet_program_const_map": {
            "ChampionDefaultOdds": "ActvSoccer_bet_initial_display_odds (17500→1.75)",
            "ChampionMaxOdds": "ActvSoccer_bet_max_odds (500000→50)",
            "ChampionMinOdds": "ActvSoccer_bet_min_odds / bet_constraint_return_rate (12000→1.2)",
            "ChampionPowerRatioMax": "ActvSoccer_bet_power_ratio_max (20000→2.0)",
            "ChampionPowerRatioMin": "ActvSoccer_bet_power_ratio_min (3000→0.3)",
            "ChampionOddsCfg": "ActvSoccerBetMultiplierCfg（同表）",
            "oneMatchPreFightTimes": "ActvSoccer_bet_sim_count (=10)",
            "maxPreFightBattlesPerTick": "ActvSoccer_bet_max_prefight_battles_per_tick (=50)",
            "roundFightInterval": "ActvSoccer_bet_round_fight_interval_sec (=1)",
            "preFightTimeout": "ActvSoccer_bet_prefight_timeout_sec (=30)",
        },
        "bet_config_gaps": [
            "爆冷规则(ChampionBaseOdds/UpSet*/体验扰动): 本期暂不接入,无活动配置",
            "bet_coin_purchase_limit / bet_daily_free_amount: 活动经济,程序ConstCommon无对应",
            "bet_display_odds_tick_sec: 主案UI动态展示,程序controller无直接键",
        ],
        "const_merge_target": "dataconfig/ConstConfig.xlsx / ConstConfigCfg",
        "const_entry_count": len(const_rows),
        "const_cfgid_range": f"{ACTV_SOCCER_CONST_CFGID_BASE}-{ACTV_SOCCER_CONST_CFGID_BASE + len(const_rows) - 1}",
        "sheet_groups": {
            "玩法基础": [
                "ActvSoccerCharacterCfg", "ActvSoccerNationalityCfg", "ActvSoccerTutorialCfg",
                "ActvSoccerGuideStepCfg",
                "ActvSoccerSlicePresetCfg", "ActvSoccerSliceInstanceCfg",
                "ActvSoccerLevelCfg", "ActvSoccerSeasonCfg",
            ],
            "FSM_BT_AI": [
                "ActvSoccerAiProfileCfg", "ActvSoccerEnemyAiCfg",
                "ActvSoccerAiModifierCfg", "ActvSoccerSliceFlowCfg",
            ],
            "养成与合同": [
                "ActvSoccerFameGrowthLevelCfg", "ActvSoccerLifeGrowthLevelCfg",
                "ActvSoccerTeamCfg", "ActvSoccerContractStarLicCfg", "ActvSoccerContractCfg",
            ],
            "淘汰赛": [
                "ActvSoccerKnockoutCfg", "ActvSoccerKnockoutPhaseCfg",
            ],
            "竞猜": [
                "ActvSoccerBetMultiplierCfg", "ActvSoccerBetStakeTierCfg",
            ],
            "体验与反馈": [
                "ActvSoccerHapticCfg",
            ],
            "接球决策": [
                "ActvSoccerReceiveDecisionCfg", "ActvSoccerPlayerStyleCfg",
                "ActvSoccerFirstTouchAnimCfg",
            ],
        },
        "id_cross_ref": {
            "AiProfile": "1001-1010=tier1-10难度档(easy/normal/hard),单调递增,DeadCorner固定0",
            "PlayerAiDuty": "1=Goalkeeper守门员,2=Defender后卫(对方非门将),3=Forward前锋(我方含玩家)",
            "EnemyAi": "2001-2003=band1(tier1-3);20X1/20X2/20X3=bandX门将/后卫/射手(X=2,3,4)",
            "Modifier": "4001/4002/4005=移动门将(普通/困难/极限),4003=无辅助线,4004=固定人墙,4006=收窄夹角,4007=随机扑救",
            "ReceiveDecision": "5001-5004=接球决策风格(Balanced/Playmaker/Dribbler/TargetMan)",
            "PlayerStyle": "5101-5104=球员风格(Balanced/Playmaker/Dribbler/TargetMan);CharacterCfg.PlayerStyleCfgID引用",
            "FirstTouchAnim": "5201+=第一脚触球表现映射(Action×BallHeight×Direction×Pressure)",
            "SliceInstance": "库实例id=tier*100+type*10+variant(type1-6; attack 1-5,free_kick 1-4,corner 1-3,penalty/throw_in/goalkeep 1-2);101-203=试训/引导;AI字段已并入本表",
            "Level": "1-500=50轮×10关;Group=轮次;第1关=引导关;淘汰赛round15起(level141)开放",
            "Season": "1-50轮单group=1;NextSeason链推进;总轮次=count(同group)",
            "BetMultiplier": "385行=程序ChampionOddsCfg二维网格;按(WinRate, PowerRate)查行→OddsLeft/OddsRight万分值",
            "BetStakeTier": "6档:free(命中+5)+50/100/150/200/300",
        },
        "test_flow": [
            "创角→试训切片101/102/103 (SliceInstance内置easy档AI字段)",
            "引导关201/202/203 (进攻+助攻+守门射手)",
            "正式关301移动门将 (Modifier4001) / 302点球",
            "困难关复用301 (Profile1003+Modifier4002)",
            "切片FSM: SliceFlowCfg按类型读流程; 角色动画映射见客户端CharacterStateCfg(不进策划xlsx)",
        ],
        "notes": [
            "球员AI职责仅三档:守门员/后卫(对方非门将)/前锋(我方含玩家);程序枚举PlayerAiDuty(1/2/3)",
            "PlayersInit: SoccerPlayerInit_V 含 facing(°) 初始朝向；回溯重置恢复摆位+facing",
            "不进策划xlsx: SliceTypeDef/PlayerAiDutyEnum/BetMatch/CharacterState(客户端+美术)",
            "知名度结算: 读current_contract.PayFinish/PayGoal/PayAssist/PayFame，无ActvSoccerFameGainRuleCfg",
            "淘汰赛演算: match_simulation_rule写在主案规则节，无ActvSoccerMatchSimulationCfg",
            "竞猜币投放: 通用GiftCfg.Reward(typ:vm bet_coin)；每日免费见常量bet_daily_free_amount",
            "竞猜赔率: BetMultiplierCfg对照表+附录常量bet_*；BetMatch运行时生成",
            "通行证/兑换商店/礼包/排名: 走项目通用配套表，不在ActivitySoccer.xlsx重复维护",
            "局内道具与玩法常量: 见配置表结构文档附录；常量合并 dataconfig/ConstConfig.xlsx",
            "知名度等级表可拆分为更细档位(行数>当前测试7档)；生活等级保持小量级不参与主角评分",
            "AI难度只控成功率; 死角球见ConstConfig ActvSoccer_dead_corner_can_save",
            "单切片AI字段已并入ActvSoccerSliceInstanceCfg,移动门将归属切片级Modifier",
            "回溯后RewindRandom=1,种子含rewind_count",
            "参数叠加: preset→instance→ai_profile→modifier",
            "JSON字段使用ext/ext[]类型,第5行标注proto(如TypIDVal_P_cspb/SoccerModifier_V)",
            "ext/ext[]数据行不得留空: ext默认{}, ext[]默认[]; 有列示例时优先用第7行示例作默认值",
            "仅含单个参数时用int/float/string等基础类型,不用ext",
            "第1行读取端: 能确定仅前端c/仅后端s,拿不准或双端用cs; 见SHEET_DEFAULT_READ/READ_OVERRIDES",
            "展示文案字段用*LcKey(string)引用语言表ID,格式ActvSoccer_{category}_{semantic}_{seq}",
        ],
        "const_keys": [row["Constant"] for row in const_rows],
        "lc_id_format": "ActvSoccer_{category}_{semantic}_{seq}",
        "lc_fields": [
            "NameLcKey", "TitleLcKey", "LeagueNameLcKey", "PhaseLcKey", "DayContentLcKey", "DescLcKey",
            "DialogueLcKey",
        ],
        "ext_proto_map": {
            "TypIDVal_P_cspb": ["SeasonReward", "SignReward", "FreeReward", "PaidReward", "Reward"],
            "PositionTuple_P": ["BallPos", "BallVector", "TargetPoint"],
            "SoccerTypePayload_V": ["TypePayload"],
            "SoccerModifier_V": ["Modifiers"],
            "SoccerPlayerInit_V": ["PlayersInit(team,idx,duty,pos,facing)"],
            "SoccerSeasonGoal_V": ["SeasonGoal"],
        },
        "flattened_fields": {
            "DefaultCameraFov": "原DefaultCamera.fov",
            "CameraFov": "原Camera.fov",
            "OverrideOperableAngle": "原Overrides.operable_angle",
            "ActvSoccerFameGrowthLevelCfg/ActvSoccerLifeGrowthLevelCfg": "原ActvSoccerGrowthLevelCfg按养成线拆分",
            "RewardFame/ContractStarLicReward/TicketCap等": "原LevelUpReward/LevelUpEffect拆列",
            "PlayerRating": "仅知名度等级表投放；读当前知名度档位行.PlayerRating；生活等级不参与",
            "effective_license_id": "max(知名度许可,生活许可) OR取较高",
            "ContractPool": "首签关联contract_id(原TeamPool)",
            "TeamCfg无Star": "星级仅在ContractCfg.TeamStar",
            "ActvSoccerContractStarLicCfg": "联赛换约星级权重(2-5星许可);同星级内均匀抽合同",
            "ActvSoccerContractCfg": "读取端cs(合同内容客户端展示)",
            "GrantScene": "首签走ContractPool;联赛换约走许可表+合同池",
            "pending_contract_choices": "服务端生成待选合同列表(玩家存档)",
            "GrantFameLevel/GrantLifeLevel": "原GrantLevelReq",
            "FameGain": "比赛结算读ContractCfg.PayFinish/PayGoal/PayAssist/PayFame，无独立FameGainRuleCfg",
            "MatchSimulation": "海选/单淘汰演算见主案match_simulation_rule规则节，无独立MatchSimulationCfg",
            "BetCoinSource": "竞猜币礼包投放见通用GiftCfg.Reward；每日免费见ActvSoccer_bet_daily_free_amount",
            "BetOdds": "mult=min(constraint_return_rate/win_rate,max_odds);ChampionOddsCfg=ActvSoccerBetMultiplierCfg",
            "CommonPlatform": "BP/兑换商店/礼包/排名见项目通用表；局内道具与常量见配置表结构附录",
            "LevelCfg.Group": "所属联赛轮次=SeasonCfg.ID",
            "SeasonCfg.Group": "联赛系列;总轮次=count(同Group);小关卡由LevelCfg.Group关联",
            "SeasonCfg.NextSeason": "后置联赛(原UnlockPrevSeason反向)",
        },
    }
    actual_sheets = set(sheets)
    rendered_sheets: set[str] = set()
    filtered_groups: dict[str, list[str]] = {}
    for group_name, group_sheets in summary["sheet_groups"].items():
        filtered = [sheet_name for sheet_name in group_sheets if sheet_name in actual_sheets]
        if filtered:
            filtered_groups[group_name] = filtered
            rendered_sheets.update(filtered)
    remaining_sheets = [sheet_name for sheet_name in sheets if sheet_name not in rendered_sheets]
    if remaining_sheets:
        filtered_groups["Other"] = remaining_sheets
    summary["sheet_groups"] = filtered_groups

    (OUT_DIR / "test-config-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def main() -> None:
    lc = LcRegistry()
    const_rows = actv_soccer_const_rows()
    wb = build_workbook(lc)
    lc_wb = build_language_workbook(lc)

    target = OUTPUT_FILE
    try:
        wb.save(target)
    except PermissionError:
        target = OUT_DIR / "ActivitySoccer_preview.generated.xlsx"
        wb.save(target)
        print(f"WARN: {OUTPUT_FILE} 被占用，已写入 {target}")

    lc_target = OUTPUT_LC_FILE
    try:
        lc_wb.save(lc_target)
    except PermissionError:
        lc_target = OUT_DIR / "ActivitySoccerLanguage.generated.xlsx"
        lc_wb.save(lc_target)
        print(f"WARN: {OUTPUT_LC_FILE} 被占用，已写入 {lc_target}")

    export_summary(wb.sheetnames, lc.rows, const_rows)
    print(f"Wrote {target}")
    print(f"Wrote {lc_target} ({len(lc.rows)} language entries)")
    print(f"Appendix: {len(IN_MATCH_ITEMS)} in-match items, {len(const_rows)} const keys (→ ConstConfig.xlsx)")
    print(f"Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
