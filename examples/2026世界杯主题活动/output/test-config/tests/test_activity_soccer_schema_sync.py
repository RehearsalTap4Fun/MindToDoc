# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
GENERATOR = ROOT / "output" / "test-config" / "generate_activity_soccer_test_config.py"
CANONICAL_XLSX = Path("C:/Project/K1Dataconfig/dataconfig/ActivitySoccer.xlsx")


def _load_generator():
    spec = importlib.util.spec_from_file_location("generate_activity_soccer_test_config", GENERATOR)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _schema_from_workbook(workbook):
    schema = {}
    for ws in workbook.worksheets:
        columns = []
        for col_idx in range(1, ws.max_column + 1):
            field = ws.cell(3, col_idx).value
            if field is None or str(field).strip() == "":
                continue
            columns.append(
                {
                    "read": ws.cell(1, col_idx).value or "",
                    "type": ws.cell(2, col_idx).value or "",
                    "field": field,
                    "server": ws.cell(4, col_idx).value or "",
                    "ext": ws.cell(5, col_idx).value or "",
                }
            )
        schema[ws.title] = columns
    return schema


def test_generated_schema_matches_latest_dataconfig_xlsx():
    assert CANONICAL_XLSX.exists(), f"missing canonical xlsx: {CANONICAL_XLSX}"

    generator = _load_generator()
    generated = generator.build_workbook(generator.LcRegistry())
    latest = load_workbook(CANONICAL_XLSX, read_only=True, data_only=True)
    try:
        assert generated.sheetnames == latest.sheetnames
        assert _schema_from_workbook(generated) == _schema_from_workbook(latest)
    finally:
        latest.close()
